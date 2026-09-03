"""The number on the customer's document is the number we use for the part.

KODE BARANG on an exported quotation came from a linked catalogue product,
and fell back to the line's position when there wasn't one. A quotation built
from a price request links no product — every line of it is written from the
request — so the fallback was all anyone ever saw: the first line of every
quotation we sent read "001", and the SKU the request had issued, the same one
the catalogue row carries, never left the building.

So the line carries its SKU now and both exports print it. The two exports
matter equally and share one resolver: a part that is "TSE-2026-0042" on the
PDF and "001" on the spreadsheet is a support call nobody can answer.

The other half is the quotations already sent. Their price request still knows
each line's SKU, matched by line number, so the migration puts it back rather
than leaving those documents numbered by position forever — which is checked
here against a row deliberately emptied first, because a backfill that only
works on data that never needed it is not a backfill.
"""
import asyncio, os, sys, uuid
os.environ.update(DATABASE_URL="postgresql+asyncpg://postgres@127.0.0.1:55432/transmisi_test",
    APP_ENV="dev", DEMO_SEED_PASSWORD="test-pass-123",
    STORAGE_LOCAL_DIR="/tmp/storage_test", JWT_SECRET="e2e-test-secret")
sys.path.insert(0, "/home/user/Gay-boy-/backend")
import httpx, logging; logging.disable(logging.INFO)
TAG = uuid.uuid4().hex[:6]
PASS, FAIL = [], []
def check(n, c, d=""):
    (PASS if c else FAIL).append(n); print(("  PASS " if c else "  FAIL ")+n+(f"  [{d}]" if d and not c else ""))
def J(r):
    try: return r.json()
    except Exception: return {"_": r.text[:200]}
def why(r):
    b = J(r)
    return str(b.get("detail")
               or (b.get("errors") or [{}])[0].get("message", "")).lower()


def xlsx_text(blob: bytes) -> str:
    """Everything a reader would see in the sheet, as one string."""
    from io import BytesIO
    from openpyxl import load_workbook
    ws = load_workbook(BytesIO(blob)).active
    out = []
    for row in ws.iter_rows(values_only=True):
        out.extend(str(v) for v in row if v is not None)
    return "\n".join(out)


def pdf_text(blob: bytes) -> str:
    """The text a reader would see. Reportlab ASCII85s its content streams,
    so pulling the words out by hand is not worth doing badly."""
    from io import BytesIO
    from pypdf import PdfReader
    return "\n".join((pg.extract_text() or "") for pg in PdfReader(BytesIO(blob)).pages)


async def main():
    from app.scripts.seed import ensure_schema; await ensure_schema()
    from app.main import app
    c = httpx.AsyncClient(transport=httpx.ASGITransport(app=app),
                          base_url="http://t/api/v1", timeout=180)

    async def login(e):
        r = await c.post("/auth/login", json={"email": e, "password": "test-pass-123"})
        return {"Authorization": f"Bearer {r.json()['access_token']}"}
    d = await login("director@demo.local")
    s1 = await login("sales1@demo.local")
    pur = await login("purchasing@demo.local")

    # ══ a request with a SKU on it ═══════════════════════════════════════
    print("\n── a part with a number ──")
    cust = J(await c.post("/customers", headers=s1, json={
        "company_name": f"PT Kode {TAG}", "industry": "mining"}))["id"]
    MY_SKU = f"SKU-{TAG.upper()}"
    pr = J(await c.post("/price-requests", headers=s1, json={
        "customer_id": cust,
        "items": [{"description": f"Conveyor Chain {TAG}", "qty": 4,
                   "uom": "pcs", "category": "conveyor_chain", "sku": MY_SKU},
                  {"description": f"Sprocket {TAG}", "qty": 2, "uom": "pcs",
                   "category": "sprocket"}]}))
    await c.post(f"/price-requests/{pr['id']}/submit", headers=s1)
    pr = J(await c.get(f"/price-requests/{pr['id']}", headers=s1))
    issued = pr["items"][1].get("sku")
    check("the request carries the SKU sales had", pr["items"][0]["sku"] == MY_SKU,
          str(pr["items"][0].get("sku")))
    check("...and issued one for the line that had none", bool(issued), str(issued))

    await c.post(f"/price-requests/{pr['id']}/price", headers=pur, json={
        "items": [{"line_no": 1, "cost_price": 500_000, "basis": "unit"},
                  {"line_no": 2, "cost_price": 200_000, "basis": "unit"}]})
    await c.post(f"/price-requests/{pr['id']}/approve", headers=d, json={
        "items": [{"line_no": 1, "sell_price": 1_000_000, "basis": "unit"},
                  {"line_no": 2, "sell_price": 400_000, "basis": "unit"}]})

    # ══ it reaches the quotation ═════════════════════════════════════════
    print("\n── and the quotation built from it ──")
    q = J(await c.post(f"/quotations/from-price-request/{pr['id']}", headers=s1))
    q_id = q["id"]
    skus = [i.get("sku") for i in q["items"]]
    check("the quotation's lines carry the part numbers", skus == [MY_SKU, issued],
          str(skus))

    # ══ both exports print it ════════════════════════════════════════════
    print("\n── on both things the customer can be sent ──")
    r = await c.get(f"/quotations/{q_id}/export.pdf", headers=s1)
    check("the PDF builds", r.status_code == 200, f"{r.status_code} {r.text[:120]}")
    text = pdf_text(r.content)
    check("...printing the SKU as KODE BARANG", MY_SKU in text, text[:400])
    check("...for every line, not just the one sales typed",
          issued in text, text[:400])
    check("...and no longer numbering the parts by position",
          "001" not in text.replace(MY_SKU, "").replace(issued or "~", ""),
          "a positional number is still on the document")

    r = await c.get(f"/quotations/{q_id}/export.xlsx", headers=s1)
    check("the spreadsheet builds", r.status_code == 200,
          f"{r.status_code} {r.text[:120]}")
    sheet = xlsx_text(r.content)
    check("...with a column for the part number", "Product no." in sheet,
          sheet[:300])
    check("...carrying the same numbers as the PDF",
          MY_SKU in sheet and (issued or "") in sheet, sheet[:400])

    # ══ editing a line must not blank it ═════════════════════════════════
    print("\n── and it survives the document being edited ──")
    body = {
        "items": [{"line_no": i["line_no"], "source": "custom",
                   "sku": i.get("sku"), "description": i["description"],
                   "spec": {}, "qty": float(i["qty"]) + 1, "uom": i["uom"],
                   "unit_price": float(i["unit_price"]), "cost_estimate": 0}
                  for i in q["items"]],
    }
    r = await c.patch(f"/quotations/{q_id}", headers=d, json=body)
    check("a quantity can be corrected", r.status_code == 200,
          f"{r.status_code} {why(r)}")
    q2 = J(await c.get(f"/quotations/{q_id}", headers=d))
    check("...without blanking the part numbers",
          [i.get("sku") for i in q2["items"]] == [MY_SKU, issued],
          str([i.get("sku") for i in q2["items"]]))

    # A change to the request has to bring the number with it, or a synced
    # quotation and a freshly built one would print different documents.
    r = await c.patch(f"/price-requests/{pr['id']}", headers=d, json={
        "items": [{"line_no": 1, "description": f"Conveyor Chain {TAG} rev B",
                   "qty": 4, "uom": "pcs", "category": "conveyor_chain"},
                  {"line_no": 2, "description": f"Sprocket {TAG}", "qty": 2,
                   "uom": "pcs", "category": "sprocket"}]})
    check("the request can be changed", r.status_code == 200,
          f"{r.status_code} {why(r)}")
    q3 = J(await c.get(f"/quotations/{q_id}", headers=d))
    check("...and the synced quotation still knows the part numbers",
          [i.get("sku") for i in q3["items"]] == [MY_SKU, issued],
          str([i.get("sku") for i in q3["items"]]))

    # ══ a quotation typed from scratch ═══════════════════════════════════
    print("\n── a quotation with no catalogue behind it ──")
    # Director-only: sales files a price request, and the quotation is built
    # from it. A typed-from-scratch one is the only case with no SKU behind it.
    r = await c.post("/quotations", headers=d, json={
        "customer_id": cust, "variant": "detailed",
        "items": [{"line_no": 1, "source": "custom",
                   "description": f"One-off {TAG}", "spec": {}, "qty": 1,
                   "uom": "pcs", "unit_price": 250_000, "cost_estimate": 0}]})
    check("a quotation typed from scratch saves", r.status_code == 201,
          f"{r.status_code} {why(r)}")
    plain = J(r)
    check("...with no part number, having no catalogue behind it",
          plain["items"][0].get("sku") in (None, ""),
          str(plain["items"][0].get("sku")))
    r = await c.get(f"/quotations/{plain['id']}/export.pdf", headers=d)
    check("...and still exports", r.status_code == 200, str(r.status_code))
    check("...falling back to the position, which is right when there is none",
          "001" in pdf_text(r.content), "no fallback number on the document")

    # ══ the documents already sent ═══════════════════════════════════════
    print("\n── quotations built before the line carried one ──")
    from app.core.db import SessionLocal, engine
    from sqlalchemy import select as _sel, text as _text
    from app.models.quotation import QuotationItem as _QI
    async with SessionLocal() as db:
        rows = (await db.scalars(
            _sel(_QI).where(_QI.quotation_id == uuid.UUID(q_id)))).all()
        for row in rows:
            row.sku = None                      # as an older document looks
        await db.commit()
    q4 = J(await c.get(f"/quotations/{q_id}", headers=d))
    check("the older shape has no part numbers on its lines",
          all(i.get("sku") in (None, "") for i in q4["items"]),
          str([i.get("sku") for i in q4["items"]]))

    from app.scripts.seed import ensure_schema as _schema
    await _schema()
    q5 = J(await c.get(f"/quotations/{q_id}", headers=d))
    check("the migration puts them back from the price request",
          [i.get("sku") for i in q5["items"]] == [MY_SKU, issued],
          str([i.get("sku") for i in q5["items"]]))
    r = await c.get(f"/quotations/{q_id}/export.pdf", headers=s1)
    check("...so an old document exports with real numbers too",
          MY_SKU in pdf_text(r.content), "still numbered by position")

    await c.aclose()
    print(f"\n{len(PASS)} passed, {len(FAIL)} failed")
    if FAIL:
        print("FAILED: " + "; ".join(FAIL)); sys.exit(1)


asyncio.run(main())

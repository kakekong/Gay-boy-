"""Which address goes on the printed document is asked, not assumed.

A customer holds three addresses and they are genuinely different places: the
office that signs the paperwork, the site the goods are delivered to, and the
one registered for tax. Which of them belongs on a document is a fact about
*that document*, not about the customer — a quotation for work at a plant in
Gresik should carry the plant, while the invoice for it goes to head office in
Jakarta.

The quotation export used to hard-code `company_address` with no way to say
otherwise. The customer PO sheet already asked, so the shapes had drifted:
two lists of addresses, two ideas of what "delivery" meant, and only one of
them included the tax address at all. Both now read the same module.

What this driver actually pins:

  the choice is offered   every address the customer holds is listed, empty
                          ones included, so somebody can see that "Alamat
                          Pengiriman" exists and is blank

  the choice is honoured  asking for the delivery address puts the delivery
                          address on the sheet — checked by reading the bytes
                          back out of the generated file, not by trusting the
                          parameter was accepted

  an empty one falls back a document with the wrong heading is recoverable;
                          one with no destination on it gets sent back

  nobody else's contact   a contact id belonging to another customer is
                          refused rather than ignored, because a stranger's
                          name and phone number on a quotation reaches the
                          customer before it reaches us
"""
import asyncio, os, sys, uuid
os.environ.update(DATABASE_URL="postgresql+asyncpg://postgres@127.0.0.1:55432/transmisi_test",
    APP_ENV="dev", DEMO_SEED_PASSWORD="test-pass-123",
    STORAGE_LOCAL_DIR="/tmp/storage_test", JWT_SECRET="e2e-test-secret")
sys.path.insert(0, "/home/user/Gay-boy-/backend")
import httpx, logging; logging.disable(logging.INFO)
PASS, FAIL = [], []
def check(n, c, d=""):
    (PASS if c else FAIL).append(n); print(("  PASS " if c else "  FAIL ")+n+(f"  [{d}]" if d and not c else ""))
def J(r):
    try: return r.json()
    except Exception: return {"_": r.text[:200]}


def xlsx_text(data: bytes) -> str:
    """Every string in a workbook, so an address can be looked for."""
    import io
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = []
    for name in wb.sheetnames:
        for row in wb[name].iter_rows(values_only=True):
            out += [str(c) for c in row if c is not None]
    wb.close()
    return "\n".join(out)


def pdf_text(data: bytes) -> str:
    """The text of a PDF, properly extracted.

    Hand-rolling this does not work: the content streams are compressed, the
    text is split across drawing operators, and a search over the raw bytes
    finds nothing even when the words are plainly on the page. That produced a
    driver that failed on behaviour which was in fact correct, including the
    default this change did not touch — so it uses a real parser.
    """
    import io
    from pypdf import PdfReader
    return "\n".join((pg.extract_text() or "") for pg in PdfReader(io.BytesIO(data)).pages)


def pdf_has(data: bytes, needle: str) -> bool:
    """Is this text on the page? Compared with whitespace and case removed,
    because a PDF line break can land anywhere."""
    import re
    squash = lambda s: re.sub(r"[^a-z0-9]", "", s.lower())
    return squash(needle) in squash(pdf_text(data))


async def main():
    from app.scripts.seed import ensure_schema; await ensure_schema()
    from app.main import app
    c = httpx.AsyncClient(transport=httpx.ASGITransport(app=app),
                          base_url="http://t/api/v1", timeout=120)

    async def login(e):
        r = await c.post("/auth/login", json={"email": e, "password": "test-pass-123"})
        return {"Authorization": f"Bearer {r.json()['access_token']}"}

    tag = uuid.uuid4().hex[:5]
    d = await login("director@demo.local")
    s1 = await login("sales1@demo.local")
    pur = await login("purchasing@demo.local")

    OFFICE = f"Jl. Kantor Pusat 1, Jakarta {tag}"
    DELIV = f"Jl. Pabrik Gresik 99, Jawa Timur {tag}"
    TAX = f"Jl. Pajak Terdaftar 7, Jakarta {tag}"

    cust = J(await c.post("/customers", headers=s1, json={
        "company_name": f"PT Tiga Alamat {tag}", "industry": "sugar",
        "company_address": OFFICE, "delivery_address": DELIV,
        "tax_address": TAX, "pic_name": "Pak Budi", "pic_position": "Purchasing",
        "phone": "031-991234", "email": "budi@example.co.id"}))
    cid = cust["id"]
    other = J(await c.post("/customers", headers=s1, json={
        "company_name": f"PT Lain {tag}", "industry": "cement"}))["id"]
    stranger = J(await c.post(f"/customers/{other}/contacts", headers=s1, json={
        "name": f"Orang Lain {tag}", "position": "Manager"}))
    mine = J(await c.post(f"/customers/{cid}/contacts", headers=s1, json={
        "name": f"Ibu Sari {tag}", "position": "Site Engineer",
        "phone": "0812-555", "email": "sari@example.co.id"}))

    # A quotation, through the flow that actually produces one.
    pr = J(await c.post("/price-requests", headers=s1, json={
        "customer_id": cid,
        "items": [{"description": f"Chain {tag}", "qty": 10, "uom": "meter"}]}))["id"]
    await c.post(f"/price-requests/{pr}/submit", headers=s1)
    await c.post(f"/price-requests/{pr}/price", headers=pur,
                 json={"items": [{"line_no": 1, "cost_price": 1_000_000, "basis": "unit"}]})
    await c.post(f"/price-requests/{pr}/approve", headers=d,
                 json={"items": [{"line_no": 1, "sell_price": 2_000_000, "basis": "unit"}]})
    q = J(await c.post(f"/quotations/from-price-request/{pr}", headers=s1))
    qid = q["id"]

    # ══ the choice is offered ════════════════════════════════════════════════
    print("\n── what the picker is given ──")
    opts = J(await c.get(f"/quotations/{qid}/pdf-options", headers=s1))
    keys = [a["key"] for a in opts["addresses"]]
    check("all three kinds of address are offered",
          keys == ["office", "delivery", "tax"], str(keys))
    by_key = {a["key"]: a for a in opts["addresses"]}
    check("...with the real text on each",
          by_key["office"]["address"] == OFFICE
          and by_key["delivery"]["address"] == DELIV
          and by_key["tax"]["address"] == TAX, str(by_key)[:200])
    check("...and an Indonesian label for the picker",
          by_key["delivery"]["label_id"] == "Alamat Pengiriman",
          str(by_key["delivery"]))
    check("the customer is named, so you know what you are printing",
          opts["customer_name"] == f"PT Tiga Alamat {tag}", str(opts.get("customer_name")))
    names = [p["name"] for p in opts["pics"]]
    check("the contacts are offered too", "Pak Budi" in names
          and f"Ibu Sari {tag}" in names, str(names))
    check("...and the customer's own PIC comes first, with no id",
          opts["pics"][0]["name"] == "Pak Budi" and opts["pics"][0]["id"] is None,
          str(opts["pics"][0]))
    check("a quotation defaults to the office", opts["default_address"] == "office",
          opts["default_address"])

    # ══ the choice is honoured — read back out of the file ═══════════════════
    print("\n── the address actually lands on the document ──")
    r = await c.get(f"/quotations/{qid}/export.pdf", headers=s1,
                    params={"address": "delivery"})
    check("the PDF export accepts an address", r.status_code == 200, str(r.status_code))
    check("...and the delivery address is on it", pdf_has(r.content, DELIV), "not found")
    check("...instead of the office one", not pdf_has(r.content, OFFICE), "office is on it")

    r = await c.get(f"/quotations/{qid}/export.pdf", headers=s1,
                    params={"address": "tax"})
    check("the tax address prints when asked for", pdf_has(r.content, TAX), "not found")

    r = await c.get(f"/quotations/{qid}/export.pdf", headers=s1)
    check("with nothing asked for, the office address prints",
          pdf_has(r.content, OFFICE), "not found")

    r = await c.get(f"/quotations/{qid}/export.xlsx", headers=s1,
                    params={"address": "delivery"})
    check("the Excel export accepts one too", r.status_code == 200, str(r.status_code))
    text = xlsx_text(r.content)
    check("...and carries the delivery address", DELIV in text, text[:200])
    check("...labelled as the delivery address, not just 'Address'",
          "Alamat Pengiriman" in text, text[:200])

    # ══ the contact ══════════════════════════════════════════════════════════
    print("\n── who it is addressed to ──")
    r = await c.get(f"/quotations/{qid}/export.pdf", headers=s1,
                    params={"address": "office", "contact_id": mine["id"]})
    check("a chosen contact is printed", pdf_has(r.content, f"Ibu Sari {tag}"), "not found")
    r = await c.get(f"/quotations/{qid}/export.pdf", headers=s1,
                    params={"contact_id": stranger["id"]})
    check("another customer's contact is refused, not printed",
          r.status_code == 400, str(r.status_code))
    r = await c.get(f"/quotations/{qid}/export.xlsx", headers=s1,
                    params={"contact_id": stranger["id"]})
    check("...on the Excel export as well", r.status_code == 400, str(r.status_code))

    # ══ an address the customer has not filled in ════════════════════════════
    print("\n── when the address asked for is empty ──")
    bare = J(await c.post("/customers", headers=s1, json={
        "company_name": f"PT Hanya Kantor {tag}", "industry": "mining",
        "company_address": OFFICE}))["id"]
    pr2 = J(await c.post("/price-requests", headers=s1, json={
        "customer_id": bare,
        "items": [{"description": f"Roller {tag}", "qty": 1, "uom": "pcs"}]}))["id"]
    await c.post(f"/price-requests/{pr2}/submit", headers=s1)
    await c.post(f"/price-requests/{pr2}/price", headers=pur,
                 json={"items": [{"line_no": 1, "cost_price": 500_000, "basis": "unit"}]})
    await c.post(f"/price-requests/{pr2}/approve", headers=d,
                 json={"items": [{"line_no": 1, "sell_price": 900_000, "basis": "unit"}]})
    q2 = J(await c.post(f"/quotations/from-price-request/{pr2}", headers=s1))["id"]

    o2 = J(await c.get(f"/quotations/{q2}/pdf-options", headers=s1))
    empty = {a["key"]: a["address"] for a in o2["addresses"]}
    check("an address the customer lacks is still listed, blank",
          empty["delivery"] == "" and empty["tax"] == "", str(empty))
    r = await c.get(f"/quotations/{q2}/export.pdf", headers=s1,
                    params={"address": "delivery"})
    check("asking for it falls back to the office rather than printing nothing",
          r.status_code == 200 and pdf_has(r.content, OFFICE), str(r.status_code))

    r = await c.get(f"/quotations/{qid}/export.pdf", headers=s1,
                    params={"address": "nonsense"})
    check("a nonsense address does not break the download",
          r.status_code == 200 and pdf_has(r.content, OFFICE), str(r.status_code))

    # ══ the customer PO sheet, which shares the module ═══════════════════════
    print("\n── the customer PO sheet uses the same list ──")
    await c.post(f"/quotations/{qid}/submit", headers=d)
    await c.post(f"/quotations/{qid}/won", headers=d)
    po = J(await c.post("/customer-pos", headers=s1, json={
        "customer_id": cid, "quotation_id": qid, "number": f"PO-ADDR-{tag}",
        "po_date": "2026-08-07",
        "items": [{"description": f"Chain {tag}", "qty": 10, "uom": "meter",
                   "unit_price": 2_000_000}], "is_downpayment": False}))
    po_id = po.get("id")
    if po_id:
        po_opts = J(await c.get(f"/customer-pos/{po_id}/pdf-options", headers=s1))
        check("the PO offers the same three addresses",
              [a["key"] for a in po_opts["addresses"]] == ["office", "delivery", "tax"],
              str([a["key"] for a in po_opts.get("addresses", [])]))
        check("...including the tax address it never used to offer",
              any(a["key"] == "tax" and a["address"] == TAX
                  for a in po_opts["addresses"]), str(po_opts["addresses"])[:160])
        check("...but defaults to delivery, as a delivery sheet should",
              po_opts["default_address"] == "delivery", str(po_opts.get("default_address")))
        r = await c.get(f"/customer-pos/{po_id}/export.pdf", headers=s1,
                        params={"ship_to": "tax"})
        check("and it can now print against the tax address",
              r.status_code == 200 and pdf_has(r.content, TAX), str(r.status_code))
        r = await c.get(f"/customer-pos/{po_id}/export.pdf", headers=s1,
                        params={"ship_to": "nonsense"})
        check("...while a nonsense one is still refused there",
              r.status_code == 400, str(r.status_code))
        # leave the approval queue as we found it
        await c.post(f"/customer-pos/{po_id}/reject", headers=d,
                     json={"notes": "test fixture"})

    # ══ scoping is unchanged ═════════════════════════════════════════════════
    print("\n── and the options endpoint is scoped like the export ──")
    s2 = await login("sales2@demo.local")
    r = await c.get(f"/quotations/{qid}/pdf-options", headers=s2)
    check("another rep cannot read this quotation's print options",
          r.status_code == 403, str(r.status_code))

    await c.aclose()
    print(f"\n{len(PASS)} passed, {len(FAIL)} failed")
    if FAIL:
        sys.exit(1)


asyncio.run(main())

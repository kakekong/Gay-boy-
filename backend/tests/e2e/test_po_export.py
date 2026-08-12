"""The supplier PO as a document you can send.

Asked for: the purchase order needs to convert to PDF and Excel.

It was the only document in the chain that existed nowhere but on a screen —
the quotation, the order confirmation and the invoice all print, so a PO going
to a vendor was being retyped into an email. Two formats because they get used
differently: the PDF is what the supplier receives, the spreadsheet is what
gets pasted into a stock sheet or a payment schedule.

The sheet has to carry the things a vendor actually acts on: which of their
addresses the goods leave from (the warehouse, when they gave one — that is
the gate the truck pulls up to), who to ask for, where to deliver, and the PO
number to quote back on the surat jalan.
"""
import asyncio, os, sys, uuid
os.environ.update(DATABASE_URL="postgresql+asyncpg://postgres@127.0.0.1:55432/transmisi_test",
    APP_ENV="dev", DEMO_SEED_PASSWORD="test-pass-123",
    STORAGE_LOCAL_DIR="/tmp/storage_test", JWT_SECRET="e2e-test-secret")
sys.path.insert(0, "/home/user/Gay-boy-/backend")
import io, httpx, logging; logging.disable(logging.INFO)
PASS, FAIL = [], []
def check(n, c, d=""):
    (PASS if c else FAIL).append(n); print(("  PASS " if c else "  FAIL ")+n+(f"  [{d}]" if d and not c else ""))
def J(r):
    try: return r.json()
    except Exception: return {"_": r.text[:200]}


def pdf_text(data: bytes) -> str:
    from pypdf import PdfReader
    return "\n".join((p.extract_text() or "") for p in PdfReader(io.BytesIO(data)).pages)


def xlsx_text(data: bytes) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data))
    return "\n".join(str(c.value) for ws in wb.worksheets
                     for row in ws.iter_rows() for c in row if c.value is not None)


async def main():
    from app.scripts.seed import ensure_schema; await ensure_schema()
    from app.main import app
    c = httpx.AsyncClient(transport=httpx.ASGITransport(app=app),
                          base_url="http://t/api/v1", timeout=120)
    tag = uuid.uuid4().hex[:5]

    async def login(e):
        r = await c.post("/auth/login", json={"email": e, "password": "test-pass-123"})
        return {"Authorization": f"Bearer {r.json()['access_token']}"}
    d = await login("director@demo.local")
    pur = await login("purchasing@demo.local")
    s1 = await login("sales1@demo.local")

    WARE = "Gudang A3, Jl. Industri Selatan 9, Cikarang"
    sup = J(await c.post("/purchasing/suppliers", headers=pur, json={
        "name": f"PT Cetak PO {tag}", "category": "fabrication",
        "company_address": "Kantor pusat, Jl. Sudirman 1, Jakarta",
        "warehouse_address": WARE, "phone": "021-5550000",
        "contacts": [{"name": f"Pak Yanto {tag}", "phone": "0812-4444-5555",
                      "email": f"yanto{tag}@pemasok.co.id", "is_primary": True}]}))["id"]

    # a real project to hang the PO on
    projects = J(await c.get("/operation/projects", headers=d))
    projects = projects if isinstance(projects, list) else projects.get("data", [])
    check("there is a project to raise a PO against", bool(projects), str(projects)[:120])
    proj = projects[0]["id"]

    po = J(await c.post("/purchasing/po", headers=d, json={
        "supplier_id": sup, "project_id": proj, "po_date": "2026-08-11",
        "quoted_lead_days": 30,
        "items": [
            {"description": f"CHAIN C-2122 {tag}", "qty": 40, "uom": "meter",
             "unit_price": 1_800_000, "amount": 72_000_000},
            {"description": f"SPROCKET 24T {tag}", "qty": 4, "uom": "pcs",
             "unit_price": 1_000_000, "amount": 4_000_000},
        ],
        "total": 76_000_000}))
    check("the PO is raised", bool(po.get("id")), str(po)[:180])
    po_id, po_no = po["id"], po["number"]

    # ══ the PDF ══════════════════════════════════════════════════════════════
    print("\n── the sheet that goes to the vendor ──")
    r = await c.get(f"/purchasing/po/{po_id}/export.pdf", headers=pur)
    check("it prints", r.status_code == 200 and r.content[:4] == b"%PDF",
          f"{r.status_code} {r.content[:20]}")
    txt = pdf_text(r.content)
    check("...naming the supplier", f"PT CETAK PO {tag}".upper() in txt.upper(),
          txt[:300])
    check("...quoting the PO number back", po_no in txt, txt[:300])
    check("...with the address the goods leave from, not the office",
          "Industri Selatan 9" in txt and "Sudirman 1" not in txt, txt[:400])
    check("...the person to ask for", f"Pak Yanto {tag}" in txt, txt[:400])
    check("...where to deliver", "SHIP TO" in txt.upper(), txt[:500])
    check("...both lines", f"CHAIN C-2122 {tag}" in txt and f"SPROCKET 24T {tag}" in txt,
          txt[:600])
    check("...and the total", "76.000.000" in txt, txt[:600])
    check("it is one page", r.content.count(b"/Type /Page") <= 3, "多")

    # ══ the spreadsheet ══════════════════════════════════════════════════════
    print("\n── the sheet that gets pasted somewhere ──")
    r = await c.get(f"/purchasing/po/{po_id}/export.xlsx", headers=pur)
    check("it exports", r.status_code == 200 and r.content[:2] == b"PK",
          f"{r.status_code} {r.content[:8]}")
    x = xlsx_text(r.content)
    check("...carrying the number and supplier",
          po_no in x and f"PT Cetak PO {tag}" in x, x[:300])
    check("...every line", f"CHAIN C-2122 {tag}" in x and f"SPROCKET 24T {tag}" in x,
          x[:400])
    check("...the lead time", "30" in x, x[:400])
    check("...and the total as a number, not a string",
          "76000000" in x.replace(".0", ""), x[-300:])

    # ══ who may print it ═════════════════════════════════════════════════════
    print("\n── and who may take a copy ──")
    r = await c.get(f"/purchasing/po/{po_id}/export.pdf", headers=s1)
    check("sales cannot print a purchase order", r.status_code == 403,
          str(r.status_code))
    r = await c.get(f"/purchasing/po/{po_id}/export.xlsx", headers=s1)
    check("...in either format", r.status_code == 403, str(r.status_code))
    r = await c.get(f"/purchasing/po/{po_id}/export.pdf", headers=d)
    check("the director can", r.status_code == 200, str(r.status_code))
    r = await c.get(f"/purchasing/po/{uuid.uuid4()}/export.pdf", headers=pur)
    check("a PO that does not exist is a 404", r.status_code == 404,
          str(r.status_code))

    # ══ the language, and the money ══════════════════════════════════════════
    # This is the one document that leaves the company in English. It goes to
    # a factory in Jiangsu; a sheet headed KEPADA — PEMASOK with the quantities
    # under NAMA BARANG is unreadable to the person filling the order.
    print("\n── it is written for the vendor, not for us ──")
    for word in ("PURCHASE ORDER", "TO — SUPPLIER", "PO DETAILS", "SHIP TO",
                 "TERMS", "DESCRIPTION", "UNIT PRICE", "AMOUNT", "TOTAL"):
        check(f'it says "{word}"', word in txt.upper(), txt[:600])
    def context(hay: str, needle: str) -> str:
        """Where the word actually is — a failure that prints 600 characters
        of letterhead tells you nothing about which line is still Indonesian."""
        i = hay.upper().find(needle.upper())
        return "—" if i < 0 else repr(hay[max(0, i - 60):i + 60])

    # The *labels* have to be English. A vendor's own Indonesian address or
    # email is their business and must not fail this — which is why these are
    # the printed headings, not bare words that can appear in supplier data.
    for word in ("KEPADA — PEMASOK", "INFORMASI PO", "NAMA BARANG",
                 "HARGA SATUAN", "KIRIM KE", "SYARAT", "PROYEK", "TANGGAL",
                 "KETERANGAN", "Hormat kami"):
        check(f'...and never "{word}"', word.upper() not in txt.upper(),
              context(txt, word))
    check("...counting days in English, not hari",
          "days" in txt and " hari" not in txt, txt[:600])

    # A price with no currency beside it is the one thing that costs money to
    # get wrong: a vendor quoting dollars reads 1.800.000 as their own number.
    print("\n── every price says what it is denominated in ──")
    check("the currency is stated in its own right", "CURRENCY" in txt.upper(),
          txt[:600])
    check("...on the money columns", "UNIT PRICE (IDR)" in txt.upper()
          and "AMOUNT (IDR)" in txt.upper(), txt[:700])
    check("...and again on the total line", "TOTAL IDR" in txt.upper(), txt[:700])
    x0 = xlsx_text((await c.get(f"/purchasing/po/{po_id}/export.xlsx",
                                headers=pur)).content)
    check("the spreadsheet says so too",
          "Currency" in x0 and "IDR" in x0 and "Unit price (IDR)" in x0, x0[:400])

    # An order to an overseas vendor, in their currency. Dollars are written
    # 1,800.00 — printing them the Indonesian way would read as 1.8 million.
    print("\n── an order in dollars ──")
    usd_po = J(await c.post("/purchasing/po", headers=d, json={
        "supplier_id": sup, "project_id": proj, "currency": "usd",
        "items": [{"description": f"CHAIN C-2122 {tag}", "qty": 40, "uom": "meter",
                   "unit_price": 120.5, "amount": 4820}],
        "total": 4820}))
    check("a PO can be raised in another currency",
          (usd_po.get("currency") or "") == "USD", str(usd_po.get("currency")))
    u = pdf_text((await c.get(f"/purchasing/po/{usd_po['id']}/export.pdf",
                              headers=pur)).content)
    check("...the columns say USD", "UNIT PRICE (USD)" in u.upper(), u[:600])
    check("...and never IDR", "IDR" not in u.upper(), u[:600])
    check("...the figures are written the way dollars are written",
          "4,820.00" in u and "4.820" not in u, u[:700])
    check("...including the unit price's cents", "120.50" in u, u[:700])

    r = await c.patch(f"/purchasing/po/{usd_po['id']}", headers=d,
                      json={"currency": "cny"})
    check("the currency can be corrected afterwards", r.status_code == 200,
          f"{r.status_code} {J(r)}"[:140])
    u2 = pdf_text((await c.get(f"/purchasing/po/{usd_po['id']}/export.pdf",
                               headers=pur)).content)
    check("...and the printed order follows", "UNIT PRICE (CNY)" in u2.upper(),
          u2[:600])

    # a supplier with no warehouse address falls back to the office
    print("\n── a supplier who only gave one address ──")
    sup2 = J(await c.post("/purchasing/suppliers", headers=pur, json={
        "name": f"PT Satu Alamat {tag}",
        "company_address": "Ruko Puri Niaga blok K"}))["id"]
    po2 = J(await c.post("/purchasing/po", headers=d, json={
        "supplier_id": sup2, "project_id": proj,
        "items": [{"description": "bolt", "qty": 1, "uom": "pcs",
                   "unit_price": 1000, "amount": 1000}], "total": 1000}))
    r = await c.get(f"/purchasing/po/{po2['id']}/export.pdf", headers=pur)
    check("it still prints", r.status_code == 200, str(r.status_code))
    check("...using the office address", "Puri Niaga" in pdf_text(r.content),
          pdf_text(r.content)[:300])

    # ══ the address we ask them to deliver to ════════════════════════════════
    # It used to be a string literal in the endpoint — an address nobody at the
    # company had confirmed, printed on every order as the place to send goods.
    # It comes from configuration now, and says it is unset rather than
    # inventing one, because a wrong address here is a container in the wrong
    # town.
    print("\n── where we ask them to deliver ──")
    from app.core.config import settings as _s
    ship = pdf_text((await c.get(f"/purchasing/po/{po_id}/export.pdf",
                                 headers=pur)).content)
    if _s.COMPANY_WAREHOUSE_ADDRESS.strip():
        check("the configured warehouse address is what prints",
              _s.COMPANY_WAREHOUSE_ADDRESS.strip().split(",")[0] in ship,
              ship[:500])
    else:
        check("with none configured it admits so instead of inventing one",
              "not set" in ship.lower(), ship[:500])
    check("...and no hard-coded address survives in the code",
          "Jl. Raya Serang KM 16" not in ship, ship[:500])

    await c.aclose()
    print(f"\n{len(PASS)} passed, {len(FAIL)} failed")
    if FAIL:
        sys.exit(1)


asyncio.run(main())

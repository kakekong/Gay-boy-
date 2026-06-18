"""Generic PDF / Excel export for tabular data (reports + KPIs).

A 'section' is a titled table:
    {"name": "P&L", "headers": ["Metric", "Value"], "rows": [["Revenue", "Rp 1.000.000"], ...]}

Both renderers take a document title + a list of sections and return raw
bytes. Used by the Reports and KPI export endpoints so they share one
consistent layout (mirrors the quotation export styling).
"""

from datetime import UTC, datetime
from io import BytesIO


def _fmt_cell(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        # Whole numbers render without trailing .0
        return f"{v:,.2f}".rstrip("0").rstrip(".") if v % 1 else f"{int(v):,}"
    return str(v)


def render_pdf(title: str, sections: list[dict]) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        HRFlowable,
        KeepTogether,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    # Narrow metric tables (≤3 columns, e.g. KPIs) read far better in
    # portrait; wide tables (reports) still get landscape so columns fit.
    max_cols = max((len(s.get("headers") or []) for s in sections), default=2)
    portrait = max_cols <= 3
    pagesize = A4 if portrait else landscape(A4)
    margin = 16 * mm
    page_w = (210 if portrait else 297) * mm
    usable = page_w - 2 * margin

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=pagesize,
        leftMargin=margin, rightMargin=margin,
        topMargin=margin, bottomMargin=margin,
        title=title,
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontSize=20,
                        alignment=0, spaceAfter=0,
                        textColor=colors.HexColor("#111827"))
    subtitle = ParagraphStyle("subtitle", parent=styles["BodyText"],
                              fontSize=11, textColor=colors.HexColor("#374151"),
                              spaceBefore=2)
    section_style = ParagraphStyle(
        "section", parent=styles["BodyText"], fontSize=12,
        fontName="Helvetica-Bold", textColor=colors.HexColor("#1f2937"),
        spaceBefore=10, spaceAfter=5,
    )
    label = ParagraphStyle(
        "label", parent=styles["BodyText"], textColor=colors.grey,
        fontSize=8.5, leading=11,
    )
    flow: list = []

    flow.append(Paragraph("Transmisi Eng", h1))
    flow.append(Paragraph(title, subtitle))
    flow.append(Paragraph(
        "Generated " + datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC"), label,
    ))
    flow.append(Spacer(1, 3 * mm))
    flow.append(HRFlowable(width="100%", thickness=1,
                           color=colors.HexColor("#e5e7eb")))
    flow.append(Spacer(1, 4 * mm))

    for sec in sections:
        headers = sec.get("headers") or []
        ncol = len(headers) or 1
        # Tidy, left-aligned widths: a wide first (label) column, the rest
        # share the remainder. Cap so a 2-column table doesn't stretch across
        # the whole page and look empty.
        if ncol == 1:
            widths = [min(usable, 120 * mm)]
        elif ncol == 2:
            widths = [min(usable * 0.62, 120 * mm), min(usable * 0.30, 45 * mm)]
        else:
            first = usable * 0.32
            rest = (usable - first) / (ncol - 1)
            widths = [first] + [rest] * (ncol - 1)

        data = [headers] + [[_fmt_cell(c) for c in row] for row in sec["rows"]]
        if len(data) == 1:
            data.append(["—"] * ncol)

        tbl = Table(data, colWidths=widths, repeatRows=1)
        tbl.hAlign = "LEFT"
        tbl.setStyle(TableStyle([
            # Header row
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 9.5),
            # Body
            ("FONT", (0, 1), (-1, -1), "Helvetica", 9),
            ("TEXTCOLOR", (0, 1), (-1, -1), colors.HexColor("#111827")),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#f3f4f6")]),
            ("LINEBELOW", (0, 1), (-1, -1), 0.3, colors.HexColor("#e5e7eb")),
            ("BOX", (0, 0), (-1, -1), 0.6, colors.HexColor("#d1d5db")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 9),
            ("RIGHTPADDING", (0, 0), (-1, -1), 9),
        ]))
        # Keep a section heading glued to its table so it never gets orphaned
        # at the bottom of a page.
        flow.append(KeepTogether([
            Paragraph(sec["name"], section_style),
            tbl,
            Spacer(1, 5 * mm),
        ]))

    doc.build(flow)
    return buf.getvalue()


def _safe_cell(v):
    """openpyxl is picky about value types — coerce Decimal / date /
    datetime / unknown objects to safe primitives or strings."""
    from datetime import date as _date
    from datetime import datetime as _dt
    from decimal import Decimal
    if v is None:
        return ""
    if isinstance(v, (int, float, str, bool)):
        return v
    if isinstance(v, Decimal):
        return float(v)
    if isinstance(v, (_date, _dt)):
        return v.isoformat()
    return str(v)


def render_xlsx(title: str, sections: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    first = True
    bold = Font(bold=True)
    right = Alignment(horizontal="right")
    used_titles: set[str] = set()

    for sec in sections:
        ws = wb.active if first else wb.create_sheet()
        first = False
        base = (sec["name"] or "Sheet")[:31].replace("/", "-").replace("\\", "-")
        title_safe = base
        n = 2
        while title_safe in used_titles:
            suffix = f" {n}"
            title_safe = (base[: 31 - len(suffix)] + suffix)
            n += 1
        used_titles.add(title_safe)
        ws.title = title_safe

        ws["A1"] = "Transmisi Eng"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = title
        ws["A2"].font = bold
        ws["A3"] = "Generated " + datetime.now(UTC).strftime("%Y-%m-%d %H:%M UTC")

        header_font = Font(bold=True, color="FFFFFF")
        header_fill_dark = PatternFill("solid", fgColor="1F2937")

        row = 5
        headers = sec["headers"]
        header_row = row
        col_max: dict[int, int] = {}
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=_safe_cell(h))
            cell.font = header_font
            cell.fill = header_fill_dark
            col_max[col_idx] = len(str(h))
        row += 1
        for r in sec["rows"]:
            for col_idx, v in enumerate(r, start=1):
                cell = ws.cell(row=row, column=col_idx, value=_safe_cell(v))
                if col_idx > 1 and isinstance(v, (int, float)):
                    cell.alignment = right
                col_max[col_idx] = max(
                    col_max.get(col_idx, 0), len(str(_safe_cell(v)))
                )
            row += 1

        # Freeze the header row and size columns to fit the widest cell.
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        for col_idx, m in col_max.items():
            width = max(12, min(48, m + 4))
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

from datetime import UTC, datetime
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, UploadFile, status
from fastapi.responses import JSONResponse, Response
from pydantic import BaseModel
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.core.approval import (
    decide,
    evaluate_discount,
    request_approval,
)
from app.core.audit import record as audit_record
from app.core.db import get_db
from app.core.deps import get_current_user
from app.core.permissions import (
    Role, can_approve_quotation, require, require_min, sales_may_see,
    sales_scope,
)
from app.models.account import Account
from app.models.approval import ApprovalRequest, ApprovalStatus
from app.models.crm import Activity, Customer, CustomerContact, Reminder
from app.models.quotation import Product, Quotation, QuotationItem
from app.models.user import User
from app.schemas.quotation import (
    QuotationAccountLinks,
    QuotationCreate,
    QuotationDecide,
    QuotationOut,
    QuotationUpdate,
)
from app.services import ledger, quote_import
from app.services.numbering import next_quotation_number

router = APIRouter(
    # Internal-only surface. External portal accounts (customer /
    # supplier, hierarchy tier 0) must never reach the CRM, pricing,
    # calendar or notification data — they have /portal/* instead.
    dependencies=[Depends(require_min(Role.SALES))]
)

# The ledger / chart-of-accounts is back-office only — sales never see it.
_ledger_viewer = require(Role.ADMIN, Role.FINANCE, Role.MANAGER, Role.DIRECTOR)


def _line_shape(it) -> tuple:
    """One line, reduced to the things a change to it would mean something.

    Compared as text for the numbers because the stored side is `Decimal` and
    the posted side is `float`, and `Decimal("40.0000") != 40.0`.
    """
    get = (lambda k: getattr(it, k, None)) if not isinstance(it, dict) else it.get
    return (
        (get("description") or "").strip(),
        f"{float(get('qty') or 0):.4f}",
        (get("uom") or "").strip().lower(),
        f"{float(get('unit_price') or 0):.2f}",
        str(get("product_id") or ""),
    )


def _changed(q: Quotation, field: str, value) -> bool:
    """Is this submitted field actually different from what is stored?

    Anything unrecognised counts as changed — a new field should behave the
    way it always did until someone teaches this function about it, rather
    than being silently dropped.
    """
    if field == "items":
        if value is None:
            return False
        current = [_line_shape(i) for i in sorted(q.items, key=lambda x: x.line_no)]
        incoming = [_line_shape(dict(i) if not isinstance(i, dict) else i)
                    for i in value]
        return current != incoming
    if field == "number":
        return (value or "").strip() != (q.number or "")
    if field in ("discount_pct", "tax_pct"):
        return abs(float(value or 0) - float(getattr(q, field) or 0)) > 1e-9
    if field == "contact_id":
        cur = q.contact_id
        return (str(value) if value else None) != (str(cur) if cur else None)
    if field in ("variant", "notes", "valid_until"):
        return value != getattr(q, field)
    return True


async def _apply_quotation_changes(db: AsyncSession, q: Quotation, changes: dict) -> None:
    """Apply a queued edit payload (JSON-shaped) to a quotation + recalc.

    Used by the direct director PATCH path and by the approvals engine when
    the director approves a non-director's edit to an approved quotation.
    `changes` comes from QuotationUpdate.model_dump(mode="json") so UUIDs
    and dates arrive as strings.
    """
    from datetime import date as _date

    for field in ("contact_id", "variant", "discount_pct", "tax_pct", "notes"):
        if field in changes:
            v = changes[field]
            if field == "contact_id" and isinstance(v, str):
                v = UUID(v)
            setattr(q, field, v)
    if "valid_until" in changes:
        v = changes["valid_until"]
        q.valid_until = _date.fromisoformat(v) if isinstance(v, str) else v
    new_items = changes.get("items")
    if new_items is not None:
        for it in list(q.items):
            await db.delete(it)
        await db.flush()
        items = []
        for i in new_items:
            row = dict(i)
            pid = row.get("product_id")
            if isinstance(pid, str):
                row["product_id"] = UUID(pid)
            row.pop("line_total", None)
            row.pop("id", None)
            items.append(QuotationItem(quotation_id=q.id, **row))
        db.add_all(items)
    else:
        items = list(q.items)
    _recalc(q, items)
    await db.flush()


def _recalc(q: Quotation, items: list[QuotationItem]) -> None:
    subtotal = sum(float(it.qty) * float(it.unit_price) for it in items)
    discount_amount = subtotal * float(q.discount_pct) / 100.0
    after_discount = subtotal - discount_amount
    tax = after_discount * float(q.tax_pct) / 100.0
    q.subtotal = subtotal
    q.discount_amount = discount_amount
    q.total = after_discount + tax
    for it in items:
        it.line_total = float(it.qty) * float(it.unit_price)


@router.post("", response_model=QuotationOut, status_code=201)
async def create_quotation(
    payload: QuotationCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    # Sales can no longer type a quotation out of thin air — the intended
    # flow is PR → purchasing costs → director sets sell + approves →
    # quotation auto-generates via /quotations/from-price-request/{pr_id}.
    # Director/manager/admin still have the direct path so they can file
    # a one-off quote when a customer's already negotiated a fixed price
    # off-system.
    if Role(user.role) == Role.SALES:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "Sales can't create a quotation directly. File a price request "
            "first — the quotation is generated automatically once the "
            "director approves the sell price on that PR.",
        )
    # Number is auto-generated from the fixed company token, but the user may
    # supply a custom full number (editable when needed) or just override the
    # token segment.
    custom = (payload.number or "").strip()
    if custom:
        clash = await db.scalar(
            select(func.count(Quotation.id)).where(Quotation.number == custom)
        )
        if clash:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                f"Quotation number '{custom}' already exists.",
            )
        number = custom
    else:
        number = await next_quotation_number(db, token=payload.number_token)
    # Owned by the account's rep, not by whoever filled the form in. The
    # director writing a one-off quote for somebody's customer is doing it on
    # that rep's behalf; leaving it in the director's name meant the rep could
    # not edit their own deal.
    _cust = await db.get(Customer, payload.customer_id)
    q = Quotation(
        number=number,
        customer_id=payload.customer_id,
        contact_id=payload.contact_id,
        variant=payload.variant,
        sales_pic_id=(_cust.sales_pic_id if _cust else None) or user.id,
        discount_pct=payload.discount_pct,
        tax_pct=payload.tax_pct,
        valid_until=payload.valid_until,
        notes=payload.notes,
        status="draft",
        created_by=user.id, updated_by=user.id,
    )
    db.add(q)
    await db.flush()
    items = [
        QuotationItem(quotation_id=q.id, **i.model_dump())
        for i in payload.items
    ]
    db.add_all(items)
    _recalc(q, items)
    await db.flush()
    return await _load(q.id, db)


@router.post("/from-price-request/{pr_id}", response_model=QuotationOut, status_code=201)
async def create_from_price_request(
    pr_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Build a quotation straight from an approved Price Request.

    Sales never types a price here — every line's unit price is the selling
    price the director set on the approved form, and the resulting quotation
    is locked to those prices.
    """
    from app.models.price_request import PriceRequest
    pr = await db.get(PriceRequest, pr_id)
    if not pr or pr.is_deleted:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Price request not found")
    # This is the step that was actually broken: a price request the director
    # raised names the director, so the rep who owns the customer could not
    # turn it into a quotation — the whole point of raising it.
    _cust = await db.get(Customer, pr.customer_id) if pr.customer_id else None
    if not sales_may_see(user, pr.sales_pic_id,
                         _cust.sales_pic_id if _cust else None):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    if pr.status != "approved":
        raise HTTPException(status.HTTP_409_CONFLICT,
                            "Only an approved price request can become a quotation")
    if pr.quotation_id:
        raise HTTPException(status.HTTP_409_CONFLICT,
                            "This price request already has a quotation")

    # PR notes carry internal `[purchasing]` / `[director]` side-channel
    # lines that must not bleed into the customer-facing quotation. Filter
    # them out at the copy point so the quote only inherits the sales-
    # authored notes that were meant for the customer.
    from app.api.v1.endpoints.price_requests import strip_internal_notes
    q = Quotation(
        number=await next_quotation_number(db),
        customer_id=pr.customer_id,
        # Same rule as everywhere else: the account's rep owns it. The price
        # request's own rep is the fallback, then whoever pressed the button.
        sales_pic_id=((_cust.sales_pic_id if _cust else None)
                      or pr.sales_pic_id or user.id),
        price_request_id=pr.id,
        discount_pct=0, tax_pct=11,
        notes=strip_internal_notes(pr.notes),
        status="draft",
        created_by=user.id, updated_by=user.id,
    )
    db.add(q)
    await db.flush()
    items = [
        QuotationItem(
            quotation_id=q.id,
            line_no=it.get("line_no") or (i + 1),
            source="custom",
            description=it.get("description") or "",
            qty=float(it.get("qty") or 0),
            uom=it.get("uom") or "pcs",
            unit_price=float(it.get("sell_price") or 0),
            cost_estimate=float(it.get("cost_price") or 0),
        )
        for i, it in enumerate(pr.items or [])
    ]
    db.add_all(items)
    _recalc(q, items)
    pr.quotation_id = q.id
    await db.flush()
    await audit_record(db, actor=user, action="create_from_price_request",
                       entity="quotation", entity_id=q.id,
                       after={"price_request_id": str(pr.id)})
    return await _load(q.id, db)


@router.post("/parse-upload")
async def parse_upload(
    file: UploadFile,
    _user: User = Depends(get_current_user),
):
    """Parse an uploaded Excel/CSV (or PDF/image) into draft line items so the
    new-quotation form can be auto-filled. Nothing is persisted — the caller
    reviews and edits before creating the quotation."""
    MAX_BYTES = 5 * 1024 * 1024
    blob = await file.read()
    if len(blob) > MAX_BYTES:
        raise HTTPException(status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, "File too large (max 5 MB)")
    return await quote_import.parse_upload(
        blob, filename=file.filename, content_type=file.content_type
    )


@router.patch("/{q_id}", response_model=QuotationOut)
async def update_quotation(
    q_id: UUID,
    payload: QuotationUpdate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Edit a quotation after creation.

    The owner (or director) may edit. Only draft / rejected quotations are
    fully editable — once the numbers have been submitted for approval or
    sent to the customer, they're locked.

    Exception: valid_until and notes can be updated at any stage before the
    deal is closed (won / lost / cancelled). Those two are meta the customer
    frequently renegotiates and don't affect pricing, so blocking them on a
    submitted / approved / sent / won quote would just force sales to file
    a bogus 'rejected' state to get back into edit mode.
    """
    q = await _load(q_id, db)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Quotation not found")
    if not await _may_see(db, user, q):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")

    data = payload.model_dump(exclude_unset=True)
    # Drop the fields the caller sent that are already what they say.
    #
    # The edit form posts the whole quotation every time, so a sales rep who
    # only typed a note still sends the line items back untouched. Every rule
    # below asks "are the items in this payload?" when it means "did the items
    # change", and the two answers stopped agreeing the moment the form got a
    # notes box. Adding a note to a price-request-backed quotation was refused
    # with "line prices are fixed and can't be edited" — about prices nobody
    # touched — and on an approved one it quietly became an approval request in
    # the director's queue instead.
    #
    # Deciding this once, here, fixes both, and keeps them fixed for any other
    # caller that posts a full document back.
    data = {k: v for k, v in data.items() if _changed(q, k, v)}
    if not data:
        # Save pressed with nothing altered. Not an error, and not worth an
        # audit line or an approval request either.
        return await _load(q.id, db)

    # Fields that don't change pricing/approval — allowed even after submit.
    _META_FIELDS = {"valid_until", "notes"}
    _CLOSED_STATES = {"won", "lost", "cancelled"}
    is_meta_only = bool(data) and all(k in _META_FIELDS for k in data)

    if q.status in _CLOSED_STATES or q.status == "superseded":
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"A '{q.status}' quotation can't be edited — the deal is closed.",
        )
    if q.status not in ("draft", "rejected") and not is_meta_only:
        if q.status not in ("approved", "sent"):
            # pending_approval: unsubmit first — editing under the
            # director's nose would decide a different document.
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                f"A '{q.status}' quotation is locked for pricing edits. "
                "Only valid_until / notes can be updated at this stage.",
            )
        # Approved/sent quotations CAN take pricing edits, but they need
        # the director's sign-off again — the numbers were already
        # approved once. Non-directors file the change; it applies when
        # the director approves in /approvals. Directors apply directly.
        if Role(user.role) != Role.DIRECTOR:
            # Only the fields that really differ, so the director is asked to
            # approve the change rather than a re-post of the whole document.
            queued = {k: v for k, v in
                      payload.model_dump(mode="json", exclude_unset=True).items()
                      if k in data}
            queued.pop("number", None)  # number stays a direct meta edit
            existing = await db.scalar(
                select(ApprovalRequest).where(
                    ApprovalRequest.target_type == "quotation_edit",
                    ApprovalRequest.target_id == q.id,
                    ApprovalRequest.status == ApprovalStatus.PENDING.value,
                )
            )
            if existing:
                existing.payload = {"action": "update", "changes": queued}
                existing.requested_by = user.id
            else:
                await request_approval(
                    db,
                    target_type="quotation_edit",
                    target_id=q.id,
                    requested_by=user.id,
                    required_role=Role.DIRECTOR,
                    reason=f"Edit approved quotation {q.number}",
                    payload={"action": "update", "changes": queued},
                )
            await audit_record(db, actor=user, action="edit_requested",
                               entity="quotation", entity_id=q.id,
                               after={"changes": sorted(queued)})
            await db.flush()
            return JSONResponse(
                status_code=status.HTTP_202_ACCEPTED,
                content={
                    "status": "pending_approval",
                    "message": "Edit sent to the director for approval — "
                               "the quotation updates once they approve.",
                },
            )

    # Prices on a price-request-backed quotation are fixed by the director's
    # approved form — sales can adjust meta (validity, notes) but not the lines.
    if q.price_request_id and Role(user.role) == Role.SALES and data.get("items") is not None:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "Line prices are fixed by the approved price request and can't be edited.",
        )

    # A custom number must stay unique.
    if data.get("number"):
        new_number = data["number"].strip()
        if new_number and new_number != q.number:
            clash = await db.scalar(
                select(func.count(Quotation.id)).where(
                    Quotation.number == new_number, Quotation.id != q.id
                )
            )
            if clash:
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    f"Quotation number '{new_number}' already exists.",
                )
            q.number = new_number
    data.pop("number", None)

    new_items = data.pop("items", None)

    for field in ("contact_id", "variant", "discount_pct", "tax_pct", "valid_until", "notes"):
        if field in data:
            setattr(q, field, data[field])

    if new_items is not None:
        # Replace the line items wholesale.
        #
        # `data` came from model_dump, so these are plain dicts — calling
        # .model_dump() on them raised AttributeError and returned a 500.
        # It survived unnoticed because every other route to this line was
        # blocked first: an unchanged payload is pruned before it gets here,
        # a price-request-backed quotation is refused outright, and an
        # approved one is queued for the director instead. What was left was
        # a real line change on a plain draft, which only its owner can make
        # — and until the scope fix, a rep rarely owned one.
        for it in list(q.items):
            await db.delete(it)
        await db.flush()
        items = []
        for i in new_items:
            row = dict(i)
            pid = row.get("product_id")
            if isinstance(pid, str):
                row["product_id"] = UUID(pid)
            row.pop("line_total", None)
            row.pop("id", None)
            items.append(QuotationItem(quotation_id=q.id, **row))
        db.add_all(items)
    else:
        items = list(q.items)

    q.updated_by = user.id
    _recalc(q, items)
    await db.flush()
    await audit_record(db, actor=user, action="update", entity="quotation",
                       entity_id=q.id)
    return await _load(q.id, db)


class QuotationReassign(BaseModel):
    sales_pic_id: UUID
    note: str | None = None


@router.post("/{q_id}/reassign", response_model=QuotationOut)
async def reassign_quotation(
    q_id: UUID,
    payload: QuotationReassign,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require(Role.DIRECTOR)),
):
    """Put a different sales rep's name on one quotation.

    Handing over a whole customer already moves every document with it. This
    is the finer instrument: one deal covered by somebody else while its rep
    is away, or a quotation raised under the wrong name.

    It changes who *owns* the document — which is what grants the right to
    submit it, mark it won, and edit it — without touching the customer.
    Visibility is wider than ownership (the rep who runs the account can
    always read it), so this is about who is answerable for the deal.
    """
    from app.api.v1.endpoints.customers import _REP_ROLES

    q = await _load(q_id, db)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Quotation not found")
    if q.status in ("won", "lost", "cancelled", "superseded"):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"A '{q.status}' quotation is closed — its owner is the record of "
            "who closed it.")

    rep = await db.get(User, payload.sales_pic_id)
    if not rep:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "That user doesn't exist")
    if not rep.is_active:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            f"{rep.full_name}'s account is disabled")
    if Role(rep.role) not in _REP_ROLES:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            f"{rep.full_name} is {rep.role} — only sales, a manager or a "
            "director can be put on a quotation")
    if q.sales_pic_id == payload.sales_pic_id:
        return await _load(q.id, db)

    old = await db.get(User, q.sales_pic_id) if q.sales_pic_id else None
    q.sales_pic_id = payload.sales_pic_id
    q.updated_by = user.id
    note = (payload.note or "").strip() or None

    # On the customer's timeline, where somebody looking at the account later
    # will find it, rather than only in an admin log.
    if q.customer_id:
        db.add(Activity(
            customer_id=q.customer_id,
            user_id=user.id,
            type="assignment",
            direction="internal",
            occurred_at=datetime.now(UTC),
            notes=(f"Quotation {q.number}: {old.full_name if old else 'nobody'}"
                   f" → {rep.full_name}" + (f" · {note}" if note else "")),
            meta={"quotation_id": str(q.id), "quotation_number": q.number,
                  "from_id": str(old.id) if old else None,
                  "from_name": old.full_name if old else "nobody",
                  "to_id": str(rep.id), "to_name": rep.full_name,
                  "note": note, "by_name": user.full_name},
        ))
    await audit_record(db, actor=user, action="reassign", entity="quotation",
                       entity_id=q.id,
                       before={"sales_pic_id": str(old.id) if old else None},
                       after={"sales_pic_id": str(rep.id), "note": note})
    await db.flush()
    return await _load(q.id, db)


@router.post("/{q_id}/submit", response_model=QuotationOut)
async def submit_quotation(
    q_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = await db.get(Quotation, q_id)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND)
    if not await _may_see(db, user, q):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    if q.status not in ("draft", "rejected"):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"A '{q.status}' quotation can't be submitted — only a draft or a "
            "rejected one.")
    # A rejected quotation goes straight back up once it has been fixed.
    # Revising it into a new -R2 draft is still there for a quote the
    # customer has already seen, but for one the director simply sent back
    # that minted a second document for no reason, and the reason it was
    # sent back lived nowhere the sales rep could read it. The note stays
    # on the row through the resubmission so the director sees what they
    # asked for the first time.
    was_rejected = q.status == "rejected"

    from app.core.config import settings

    if settings.QUOTATION_ALWAYS_DIRECTOR_APPROVAL:
        # Policy: every quotation needs the director's sign-off before it
        # leaves draft — no discount-based auto-approve shortcut.
        required_role = Role.DIRECTOR
        reason = "Director approval required for every quotation"
    else:
        rule = evaluate_discount(float(q.discount_pct))
        required_role = rule.required_role
        reason = rule.reason

    if required_role is None:
        q.status = "approved"
    else:
        q.status = "pending_approval"
        await request_approval(
            db,
            target_type="quotation",
            target_id=q.id,
            requested_by=user.id,
            required_role=required_role,
            reason=reason,
            payload={"discount_pct": float(q.discount_pct), "total": float(q.total)},
        )
    await audit_record(db, actor=user, action="resubmit" if was_rejected else "submit",
                       entity="quotation",
                       entity_id=q.id, before={"status": "draft"},
                       after={"status": q.status})
    await db.flush()
    return await _load(q.id, db)


@router.post("/{q_id}/unsubmit", response_model=QuotationOut)
async def unsubmit_quotation(
    q_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Pull an accidentally-submitted quotation back to draft.

    Only works while it's still `pending_approval` — once the director has
    approved or rejected it, the decision stands and the normal edit rules
    apply. Only the owner (or director) may withdraw. The pending
    ApprovalRequest is deleted so the director's queue can't decide a
    quotation that's already back in draft (a stale approve would have
    flipped it straight to 'approved').
    """
    q = await db.get(Quotation, q_id)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Quotation not found")
    if not await _may_see(db, user, q):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    if Role(user.role) not in (Role.SALES, Role.MANAGER, Role.DIRECTOR, Role.ADMIN):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    if q.status != "pending_approval":
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Only a pending-approval quotation can be unsubmitted — this "
            f"one is '{q.status}'."
            + (" It was already decided; edit it (draft/rejected rules) or "
               "ask the director." if q.status in ("approved", "rejected") else ""),
        )
    # Remove the live approval request so the approvals queue stays clean
    # and nobody can decide a withdrawn submission.
    reqs = (await db.scalars(
        select(ApprovalRequest).where(
            ApprovalRequest.target_type == "quotation",
            ApprovalRequest.target_id == q.id,
            ApprovalRequest.status == ApprovalStatus.PENDING.value,
        )
    )).all()
    for req in reqs:
        await db.delete(req)
    q.status = "draft"
    q.updated_by = user.id
    await audit_record(db, actor=user, action="unsubmit", entity="quotation",
                       entity_id=q.id, before={"status": "pending_approval"},
                       after={"status": "draft",
                              "withdrawn_requests": len(reqs)})
    await db.flush()
    return await _load(q.id, db)


@router.post("/{q_id}/revise", response_model=QuotationOut, status_code=201)
async def revise_quotation(
    q_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Post a revision: clone this quotation into a new editable draft.

    The clone carries everything (items, prices, discount, PR link, notes)
    with the number suffixed -R<version> and parent_id pointing back here.
    It walks the normal submit → director-approval path; when the revision
    is APPROVED, this original flips to 'superseded' so only one version
    of the offer is ever live. Draft/pending quotations can't be revised —
    a draft is directly editable, and a pending one should be unsubmitted.
    """
    q = await _load(q_id, db)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Quotation not found")
    if not await _may_see(db, user, q):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    if Role(user.role) not in (Role.SALES, Role.MANAGER, Role.DIRECTOR, Role.ADMIN):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    if q.status in ("draft", "pending_approval"):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "A draft is directly editable and a pending quotation should be "
            "unsubmitted — revisions are for approved/sent/rejected/lost quotes.",
        )
    _won_pending = await db.scalar(
        select(ApprovalRequest).where(
            ApprovalRequest.target_type == "quotation_won",
            ApprovalRequest.target_id == q.id,
            ApprovalRequest.status == ApprovalStatus.PENDING.value,
        )
    )
    if _won_pending:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "A Mark-won request is pending with the director — wait for that "
            "decision before posting a revision.",
        )
    if q.status in ("won", "cancelled", "superseded"):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"A '{q.status}' quotation can't be revised"
            + (" — revise its live revision instead."
               if q.status == "superseded" else " — the deal is closed."),
        )
    open_rev = await db.scalar(
        select(Quotation).where(
            Quotation.parent_id == q.id,
            Quotation.status.in_(["draft", "pending_approval"]),
        )
    )
    if open_rev:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Revision {open_rev.number} is already open — finish or "
            "discard that one first.",
        )

    new_version = int(q.version or 1) + 1
    base = q.number
    # Strip an existing -R<n> suffix so R2's revision is -R3, not -R2-R3.
    import re as _re2
    base = _re2.sub(r"-R\d+$", "", base)
    number = f"{base}-R{new_version}"
    clash = await db.scalar(
        select(func.count(Quotation.id)).where(Quotation.number == number)
    )
    if clash:
        number = f"{base}-R{new_version}-{datetime.now(UTC).strftime('%H%M%S')}"

    rev = Quotation(
        number=number,
        customer_id=q.customer_id,
        contact_id=q.contact_id,
        variant=q.variant,
        sales_pic_id=q.sales_pic_id,
        price_request_id=q.price_request_id,
        parent_id=q.id,
        version=new_version,
        discount_pct=q.discount_pct,
        tax_pct=q.tax_pct,
        valid_until=q.valid_until,
        notes=q.notes,
        status="draft",
        created_by=user.id, updated_by=user.id,
    )
    db.add(rev)
    await db.flush()
    items = [
        QuotationItem(
            quotation_id=rev.id,
            line_no=it.line_no,
            source=it.source,
            product_id=it.product_id,
            description=it.description,
            spec=it.spec,
            qty=it.qty,
            uom=it.uom,
            unit_price=it.unit_price,
            cost_estimate=it.cost_estimate,
        )
        for it in q.items
    ]
    db.add_all(items)
    _recalc(rev, items)
    await db.flush()
    await audit_record(db, actor=user, action="revise", entity="quotation",
                       entity_id=q.id,
                       after={"revision_id": str(rev.id), "number": number})
    return await _load(rev.id, db)


@router.post("/{q_id}/approve", response_model=QuotationOut)
async def approve_quotation(q_id: UUID, payload: QuotationDecide,
                            db: AsyncSession = Depends(get_db),
                            user: User = Depends(get_current_user)):
    if not can_approve_quotation(user):
        raise HTTPException(status.HTTP_403_FORBIDDEN,
                            "Only manager or director can approve quotations")
    q = await db.get(Quotation, q_id)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Quotation not found")
    # Only a SUBMITTED quotation can be approved. Allowing draft/rejected
    # meant no ApprovalRequest existed, so decide() never ran and the
    # director-approval gate (QUOTATION_ALWAYS_DIRECTOR_APPROVAL) was
    # bypassable by approving the quote before it was submitted.
    if q.status != "pending_approval":
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Cannot approve a quotation in status '{q.status}'",
        )
    # Decide any pending approval request that's still open; missing or
    # already-decided requests are fine — we still set the quotation status.
    req = await db.scalar(
        select(ApprovalRequest).where(
            ApprovalRequest.target_type == "quotation",
            ApprovalRequest.target_id == q_id,
            ApprovalRequest.status == ApprovalStatus.PENDING.value,
        )
    )
    if req:
        try:
            await decide(db, request_id=req.id, decider_id=user.id,
                         decider_role=Role(user.role), approve=True,
                         notes=payload.notes)
        except PermissionError as e:
            raise HTTPException(status.HTTP_403_FORBIDDEN, str(e)) from e
    q.status = "approved"
    # Fused pipeline: approving the quotation IS the sign-off that the deal
    # reached the 'quotation' stage — advance the customer automatically so
    # sales doesn't file a separate stage-move request for the same ground.
    if q.customer_id:
        from app.core.stage_playbook import bump_customer_stage
        from app.core.stage_tasks import ensure_stage_tasks
        cust = await db.get(Customer, q.customer_id)
        if cust and bump_customer_stage(cust, "quotation"):
            await ensure_stage_tasks(db, cust, "quotation")
    # A revision that gets approved replaces its parent: the old quotation
    # flips to 'superseded' so two versions of the same offer can't both
    # look live. Won/cancelled parents are history and stay untouched.
    if q.parent_id:
        parent = await db.get(Quotation, q.parent_id)
        if parent and parent.status not in ("won", "cancelled", "superseded"):
            parent.status = "superseded"
            await audit_record(db, actor=user, action="superseded",
                               entity="quotation", entity_id=parent.id,
                               after={"by": str(q.id), "by_number": q.number})
    await audit_record(db, actor=user, action="approve", entity="quotation",
                       entity_id=q.id, after={"status": "approved"})
    return await _load(q.id, db)


@router.post("/{q_id}/reject", response_model=QuotationOut)
async def reject_quotation(q_id: UUID, payload: QuotationDecide,
                           db: AsyncSession = Depends(get_db),
                           user: User = Depends(get_current_user)):
    if not can_approve_quotation(user):
        raise HTTPException(status.HTTP_403_FORBIDDEN,
                            "Only manager or director can reject quotations")
    q = await db.get(Quotation, q_id)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Quotation not found")
    if q.status not in ("pending_approval", "draft", "approved"):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"Cannot reject a quotation in status '{q.status}'",
        )
    req = await db.scalar(
        select(ApprovalRequest).where(
            ApprovalRequest.target_type == "quotation",
            ApprovalRequest.target_id == q_id,
            ApprovalRequest.status == ApprovalStatus.PENDING.value,
        )
    )
    if req:
        try:
            await decide(db, request_id=req.id, decider_id=user.id,
                         decider_role=Role(user.role), approve=False,
                         notes=payload.notes)
        except PermissionError as e:
            raise HTTPException(status.HTTP_403_FORBIDDEN, str(e)) from e
    q.status = "rejected"
    # On the quotation itself, not only in the audit log and the approval
    # request. The person who has to fix it is looking at this document.
    q.decision_notes = (payload.notes or "").strip() or None
    q.updated_by = user.id
    await audit_record(db, actor=user, action="reject", entity="quotation",
                       entity_id=q.id, after={"status": "rejected",
                                              "notes": payload.notes})
    return await _load(q.id, db)


@router.post("/{q_id}/won", response_model=QuotationOut)
async def mark_won(q_id: UUID, db: AsyncSession = Depends(get_db),
                   user: User = Depends(get_current_user)):
    q = await db.get(Quotation, q_id)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND)
    if not await _may_see(db, user, q):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    # Won and Lost are mutually exclusive terminal outcomes — once one is
    # set (or superseded by a revision), the other can't be clicked.
    if q.status in ("won", "lost", "cancelled", "superseded"):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"A '{q.status}' quotation can't be marked Won.",
        )
    # Won is a claim about the customer, so it needs the customer's own
    # document behind it. Their PO is that document: until one is on file
    # here, "we won this" is one person's word, and it is the word that
    # starts a project, posts revenue and moves the sales figures.
    #
    # This binds the director too. It is not a permission — it is what the
    # word means — and an escape hatch for the one person whose sign-off the
    # rule exists to inform would empty it out.
    from app.models.customer_po import CustomerPO
    evidence = await db.scalar(
        select(func.count(CustomerPO.id)).where(
            CustomerPO.quotation_id == q.id,
            CustomerPO.status.notin_(("rejected", "cancelled")),
        )
    )
    if not evidence:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "Enter the customer's PO first — a quotation is marked Won on the "
            "strength of their order, not before it.",
        )
    # Fused pipeline: no more "advance to negotiation first" bounce. The
    # director's Won approval IS the sign-off that the deal reached
    # negotiation — the customer's stage is bumped automatically when the
    # Won lands (direct director path below, or apply_to_target for the
    # approval-request path).
    cust = await db.get(Customer, q.customer_id) if q.customer_id else None
    # Only the director can flip a deal to Won directly. Everyone else files
    # an approval request; the Won + ledger posting happen on approval.
    if Role(user.role) != Role.DIRECTOR:
        existing = await db.scalar(
            select(ApprovalRequest).where(
                ApprovalRequest.target_type == "quotation_won",
                ApprovalRequest.target_id == q.id,
                ApprovalRequest.status == ApprovalStatus.PENDING.value,
            )
        )
        if not existing:
            await request_approval(
                db,
                target_type="quotation_won",
                target_id=q.id,
                requested_by=user.id,
                required_role=Role.DIRECTOR,
                reason=f"Mark quotation {q.number} as Won",
                payload={
                    "total": float(q.total or 0),
                    "customer_id": str(q.customer_id) if q.customer_id else None,
                    "quotation_number": q.number,
                },
            )
        await audit_record(db, actor=user, action="won_requested",
                           entity="quotation", entity_id=q.id,
                           after={"status": "pending_won_approval"})
        await db.flush()
        return JSONResponse(
            status_code=status.HTTP_202_ACCEPTED,
            content={
                "status": "pending_approval",
                "message": "Mark-won sent to the director for approval.",
            },
        )
    q.status = "won"
    # The director may mark Won directly while sales' own mark-won request is
    # still pending. Close it here — otherwise it sits in the approvals inbox
    # forever with nothing left to decide.
    stale = (await db.scalars(
        select(ApprovalRequest).where(
            ApprovalRequest.target_type == "quotation_won",
            ApprovalRequest.target_id == q.id,
            ApprovalRequest.status == ApprovalStatus.PENDING.value,
        )
    )).all()
    for req in stale:
        req.status = ApprovalStatus.APPROVED.value
        req.decided_by = user.id
        req.decided_at = datetime.now(UTC)
        req.decision_notes = "Marked Won directly by the director."
    # Projects no longer auto-spawn from a Won quotation — a Customer PO
    # has to be filed and approved by the director first. The Won flag
    # just means "the customer said yes, we're waiting for paperwork."
    # See app/api/v1/endpoints/customer_pos.py for the new gate.
    #
    # Fused pipeline: Won implies the deal reached negotiation — advance the
    # customer stage in the same stroke (forward-only, no-op if past it).
    if cust:
        from app.core.stage_playbook import bump_customer_stage
        from app.core.stage_tasks import ensure_stage_tasks
        if bump_customer_stage(cust, "negotiation"):
            await ensure_stage_tasks(db, cust, "negotiation")
    # Auto-post to the ledger (idempotent)
    try:
        await ledger.post_quotation(db, q)
    except Exception:
        # Don't block the won flow if posting fails (e.g. missing accounts)
        pass
    await audit_record(db, actor=user, action="won", entity="quotation",
                       entity_id=q.id, after={"status": "won", "total": float(q.total or 0)})
    return await _load(q.id, db)


@router.post("/{q_id}/lost", response_model=QuotationOut)
async def mark_lost(q_id: UUID, reason: str,
                    db: AsyncSession = Depends(get_db),
                    user: User = Depends(get_current_user)):
    # A lost deal without a reason is a hole in the lost-deal report, which is
    # the only place the company learns why it loses work. The composer already
    # insists on one; this is the rule itself, so it holds for anything calling
    # the API — `reason: str` alone accepts "" and " ".
    reason = (reason or "").strip()
    if len(reason) < 3:
        raise HTTPException(
            status.HTTP_400_BAD_REQUEST,
            "Say why the quotation was lost — the reason is what the "
            "lost-deal report is built from.",
        )
    q = await db.get(Quotation, q_id)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND)
    if not await _may_see(db, user, q):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    # Won and Lost are mutually exclusive: a won/closed quote can't flip
    # to lost, and while a Mark-won request sits in the director's queue
    # the outcome is theirs to decide — Lost is blocked until then.
    if q.status in ("won", "lost", "cancelled", "superseded"):
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            f"A '{q.status}' quotation can't be marked Lost.",
        )
    won_pending = await db.scalar(
        select(ApprovalRequest).where(
            ApprovalRequest.target_type == "quotation_won",
            ApprovalRequest.target_id == q.id,
            ApprovalRequest.status == ApprovalStatus.PENDING.value,
        )
    )
    if won_pending:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "A Mark-won request is pending with the director — wait for "
            "that decision (or ask the director to reject it) before "
            "marking the quotation Lost.",
        )
    q.status = "lost"
    q.notes = (q.notes or "") + f"\n[lost @ {datetime.now(UTC).isoformat()}] {reason}"
    await audit_record(db, actor=user, action="lost", entity="quotation",
                       entity_id=q.id, after={"status": "lost", "reason": reason})
    return await _load(q.id, db)


async def _log_export_activity(db: AsyncSession, q: Quotation, user: User,
                               fmt: str) -> None:
    """Every PDF/Excel export lands in the customer's activity timeline —
    who pulled which quotation, in which format, when."""
    db.add(Activity(
        customer_id=q.customer_id,
        user_id=user.id,
        type="export",
        direction="internal",
        occurred_at=datetime.now(UTC),
        notes=f"Exported quotation {q.number} as {fmt}",
        meta={"quotation_id": str(q.id), "format": fmt.lower()},
    ))
    await db.flush()


async def _signature_of(rep) -> bytes | None:
    """The signing rep's scanned signature, or None to leave a blank space."""
    from app.services.signature import load_for
    return await load_for(rep)


async def _may_see(db: AsyncSession, user: User, q: Quotation) -> bool:
    """Can this user open this quotation?

    Sales sees a quotation two ways: they are named on it, or the customer is
    theirs. The second matters because anyone may raise the paperwork — a
    director filing the price request behind it, an admin entering it — and
    the rep in charge of the account still has to be able to work the deal.
    """
    if Role(user.role) != Role.SALES:
        return True
    if q.sales_pic_id == user.id:
        return True
    cust = await db.get(Customer, q.customer_id) if q.customer_id else None
    return bool(cust and cust.sales_pic_id == user.id)


async def _load(q_id: UUID, db: AsyncSession) -> Quotation:
    q = await db.scalar(
        select(Quotation)
        .options(selectinload(Quotation.items))
        .where(Quotation.id == q_id)
    )
    if q is not None:
        # Enrich with lineage the detail page renders as click-through
        # chips: the source price request's number, the parent this quote
        # revises, and any newer revisions of this quote.
        pr_number = None
        if q.price_request_id:
            from app.models.price_request import PriceRequest
            pr = await db.get(PriceRequest, q.price_request_id)
            pr_number = pr.number if pr else None
        q.price_request_number = pr_number
        parent_number = None
        if q.parent_id:
            parent = await db.get(Quotation, q.parent_id)
            parent_number = parent.number if parent else None
        q.parent_number = parent_number
        revs = (await db.scalars(
            select(Quotation).where(Quotation.parent_id == q.id)
            .order_by(Quotation.version.asc())
        )).all()
        q.revisions = [
            {"id": str(r.id), "number": r.number, "status": r.status,
             "version": r.version, "created_at": r.created_at.isoformat() if r.created_at else None}
            for r in revs
        ]
        # Live director queues on this quote, so the UI can lock the
        # buttons that conflict (Mark lost while Won is pending, etc.).
        q.won_pending = bool(await db.scalar(
            select(ApprovalRequest.id).where(
                ApprovalRequest.target_type == "quotation_won",
                ApprovalRequest.target_id == q.id,
                ApprovalRequest.status == ApprovalStatus.PENDING.value,
            )
        ))
        q.edit_pending = bool(await db.scalar(
            select(ApprovalRequest.id).where(
                ApprovalRequest.target_type == "quotation_edit",
                ApprovalRequest.target_id == q.id,
                ApprovalRequest.status == ApprovalStatus.PENDING.value,
            )
        ))
    if q is not None and q.notes:
        # Safety net: any legacy quotation whose notes still carry a
        # [purchasing]/[director]/… side-channel line from the old copy
        # path gets scrubbed on the way out so it never renders on a
        # sales page or on the PDF/portal.
        from app.api.v1.endpoints.price_requests import strip_internal_notes
        cleaned = strip_internal_notes(q.notes)
        if cleaned != q.notes:
            q.notes = cleaned
    return q


@router.get("", response_model=list[QuotationOut])
async def list_quotations(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
    status_eq: str | None = None,
    customer_id: UUID | None = None,
    limit: int = 50,
):
    stmt = (
        select(Quotation)
        .options(selectinload(Quotation.items))
        .order_by(Quotation.created_at.desc())
        .limit(limit)
    )
    stmt = sales_scope(user, stmt, Quotation.sales_pic_id, Quotation.customer_id)
    if status_eq:
        stmt = stmt.where(Quotation.status == status_eq)
    if customer_id:
        stmt = stmt.where(Quotation.customer_id == customer_id)
    rows = (await db.scalars(stmt)).all()
    return [QuotationOut.model_validate(r) for r in rows]


@router.get("/stats")
async def stats(db: AsyncSession = Depends(get_db),
                user: User = Depends(get_current_user)):
    stmt = select(func.count(Quotation.id))
    # Counted the same way the list is scoped, or the number disagrees with
    # the rows underneath it.
    stmt = sales_scope(user, stmt, Quotation.sales_pic_id, Quotation.customer_id)
    total = await db.scalar(stmt)
    return {"total": total or 0}


def _idr(n) -> str:
    n = float(n or 0)
    sign = "-" if n < 0 else ""
    s = f"{abs(n):,.0f}".replace(",", ".")
    return f"{sign}Rp {s}"


async def _quotation_bundle(
    db: AsyncSession, q_id: UUID
) -> tuple[Quotation, list, Customer | None, CustomerContact | None, User | None]:
    q = await db.get(Quotation, q_id)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Quotation not found")
    items = (await db.scalars(
        select(QuotationItem).where(QuotationItem.quotation_id == q.id)
        .order_by(QuotationItem.line_no.asc())
    )).all()
    cust = await db.get(Customer, q.customer_id) if q.customer_id else None
    contact = await db.get(CustomerContact, q.contact_id) if q.contact_id else None
    sales = await db.get(User, q.sales_pic_id) if q.sales_pic_id else None
    return q, items, cust, contact, sales


def _pic_fields(cust: Customer | None, contact: CustomerContact | None) -> dict:
    """Return the addressed-to PIC details. Falls back to the customer
    record's primary PIC when no specific contact was selected."""
    if contact:
        return {
            "name":     contact.name or "—",
            "position": contact.position or "—",
            "email":    contact.email or "—",
            "phone":    contact.phone or contact.whatsapp or "—",
        }
    if cust:
        return {
            "name":     cust.pic_name or "—",
            "position": cust.pic_position or "—",
            "email":    cust.email or "—",
            "phone":    cust.phone or cust.whatsapp or "—",
        }
    return {"name": "—", "position": "—", "email": "—", "phone": "—"}


async def _chosen_contact(db: AsyncSession, cust, contact_id, fallback):
    """The contact the caller asked for, or the quotation's own.

    A contact id that belongs to some other customer is refused rather than
    ignored — printing a stranger's name and phone number on a quotation is
    the kind of mistake that reaches the customer before it reaches us.
    """
    if contact_id is None:
        return fallback
    ct = await db.get(CustomerContact, contact_id)
    if not ct or not cust or ct.customer_id != cust.id:
        raise HTTPException(status.HTTP_400_BAD_REQUEST,
                            "That contact doesn't belong to this customer")
    return ct


@router.get("/{q_id}/pdf-options")
async def quotation_pdf_options(
    q_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """The addresses and contacts this quotation could be printed against.

    Asked at download time rather than stored on the document: the same
    customer takes correspondence at a head office and delivery at a plant,
    and which one belongs on the sheet depends on what the quotation is for.
    """
    from app.services.print_address import address_options, contact_options

    q, _items, cust, _contact, _sales = await _quotation_bundle(db, q_id)
    if not sales_may_see(user, q.sales_pic_id, cust.sales_pic_id if cust else None):
        raise HTTPException(status.HTTP_403_FORBIDDEN)
    contacts = []
    if cust:
        contacts = (await db.scalars(
            select(CustomerContact).where(CustomerContact.customer_id == cust.id)
            .order_by(CustomerContact.is_primary.desc(), CustomerContact.name)
        )).all()
    return {
        "customer_name": cust.company_name if cust else "—",
        "addresses": address_options(cust),
        "pics": contact_options(cust, contacts),
        "default_address": "office",
    }


@router.get("/{q_id}/export.pdf")
async def export_quotation_pdf(
    q_id: UUID,
    address: str = "office",
    contact_id: UUID | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from app.services.print_address import resolve_address
    from app.services.quotation_pdf import build_quotation_pdf

    q, items, cust, contact, sales = await _quotation_bundle(db, q_id)
    if not sales_may_see(user, q.sales_pic_id, cust.sales_pic_id if cust else None):
        raise HTTPException(status.HTTP_403_FORBIDDEN)
    contact = await _chosen_contact(db, cust, contact_id, contact)
    await _log_export_activity(db, q, user, "PDF")
    pic = _pic_fields(cust, contact)
    addr = resolve_address(cust, address)

    # KODE BARANG comes from the linked product where there is one; a
    # free-text line falls back to its position on the sheet.
    codes: dict = {}
    product_ids = [it.product_id for it in items if it.product_id]
    if product_ids:
        for prod in (await db.scalars(
            select(Product).where(Product.id.in_(product_ids))
        )).all():
            codes[prod.id] = prod.code

    rows = [{
        "code": codes.get(it.product_id) or f"{it.line_no:03d}",
        "name": it.description or "",
        "qty": f"{float(it.qty or 0):g}",
        "uom": (it.uom or "").upper(),
        "unit_price": float(it.unit_price or 0),
        "line_total": float(it.line_total or 0)
                      or float(it.qty or 0) * float(it.unit_price or 0),
    } for it in items]

    subtotal = float(q.subtotal or 0)
    disc_amt = float(q.discount_amount or 0)
    tax_pct = float(q.tax_pct or 0)
    tax = (subtotal - disc_amt) * (tax_pct / 100.0)

    pdf = build_quotation_pdf(
        number=q.number or "—",
        issued=q.created_at.strftime("%d %b %Y") if q.created_at else "—",
        customer_name=(cust.company_name if cust else "—"),
        customer_address=addr.address,
        cp_name=pic["name"],
        cp_position=pic["position"],
        cp_email=pic["email"],
        rows=rows,
        subtotal=subtotal,
        discount=disc_amt,
        tax_pct=tax_pct,
        tax=tax,
        total=float(q.total or 0),
        notes=q.notes,
        signer_name=(sales.full_name if sales else "—"),
        signer_phone=((sales.phone if sales else None)
                      or (sales.whatsapp_id if sales else None) or "—"),
        # The address this rep corresponds from, when they have one. Their
        # login is an internal credential and has no business on a document
        # a customer replies to.
        signer_email=((sales.contact_email or sales.email) if sales else "—"),
        signer_signature=await _signature_of(sales),
    )
    return Response(
        content=pdf,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f'attachment; filename="quotation-{q.number}.pdf"',
        },
    )


@router.get("/{q_id}/export.xlsx")
async def export_quotation_excel(
    q_id: UUID,
    address: str = "office",
    contact_id: UUID | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from app.services.print_address import resolve_address
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    q, items, cust, contact, sales = await _quotation_bundle(db, q_id)
    if not sales_may_see(user, q.sales_pic_id, cust.sales_pic_id if cust else None):
        raise HTTPException(status.HTTP_403_FORBIDDEN)
    contact = await _chosen_contact(db, cust, contact_id, contact)
    await _log_export_activity(db, q, user, "Excel")
    pic = _pic_fields(cust, contact)
    addr = resolve_address(cust, address)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Quotation {q.number}"[:31]
    bold = Font(bold=True)
    title_font = Font(bold=True, size=16)
    header_fill = PatternFill("solid", fgColor="EEF0F4")
    thin = Side(border_style="thin", color="CBD1DC")
    box = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")

    ws["A1"] = "Transmisi Eng"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")
    ws["A2"] = f"Quotation {q.number}"
    ws["A2"].font = bold
    ws.merge_cells("A2:F2")

    row = 4
    ws.cell(row=row, column=1, value="Bill to / Addressed to").font = bold
    row += 1
    for label, value in [
        ("Customer",      cust.company_name if cust else "—"),
        (addr.label,      addr.address),
        ("PIC",           pic["name"]),
        ("Department",    pic["position"]),
        ("Email",         pic["email"]),
        ("Phone",         pic["phone"]),
        ("",              ""),
        ("Variant",       (q.variant or "—").title()),
        ("Status",        (q.status or "—").replace("_", " ").title()),
        ("Issued",        q.created_at.strftime("%Y-%m-%d") if q.created_at else "—"),
        ("Valid until",   q.valid_until.strftime("%Y-%m-%d") if q.valid_until else "—"),
    ]:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="Line items").font = bold
    row += 1
    headers = ["#", "Description", "Qty", "Unit", "Unit price (IDR)", "Line total (IDR)"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = bold
        cell.fill = header_fill
        cell.border = box
    row += 1
    for i, it in enumerate(items, start=1):
        line_total = float(it.qty or 0) * float(it.unit_price or 0)
        vals = [
            i,
            it.description or "",
            float(it.qty or 0),
            it.uom or "",
            float(it.unit_price or 0),
            line_total,
        ]
        for col_idx, v in enumerate(vals, start=1):
            cell = ws.cell(row=row, column=col_idx, value=v)
            cell.border = box
            if col_idx >= 3:
                cell.alignment = right
            if col_idx in (5, 6):
                cell.number_format = '"Rp" #,##0'
        row += 1

    # Totals block
    subtotal = float(q.subtotal or 0)
    disc_amt = float(q.discount_amount or 0)
    after = subtotal - disc_amt
    tax = after * (float(q.tax_pct or 0) / 100.0)
    total = float(q.total or 0)
    row += 1
    for label, value, fmt in [
        ("Subtotal",                                      subtotal, '"Rp" #,##0'),
        (f"Discount ({float(q.discount_pct or 0):.1f}%)", -disc_amt, '"Rp" #,##0'),
        (f"Tax ({float(q.tax_pct or 0):.1f}%)",           tax,      '"Rp" #,##0'),
        ("TOTAL",                                         total,    '"Rp" #,##0'),
    ]:
        ws.cell(row=row, column=5, value=label).font = bold
        cell = ws.cell(row=row, column=6, value=value)
        cell.font = bold
        cell.number_format = fmt
        cell.alignment = right
        row += 1

    # Column widths
    widths = [5, 50, 8, 8, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    if q.notes:
        row += 1
        ws.cell(row=row, column=1, value="Notes").font = bold
        row += 1
        ws.cell(row=row, column=1, value=q.notes)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    # ── Sales rep block at the bottom ────────────────────────────────────
    row += 2
    ws.cell(row=row, column=1, value="Issued by").font = bold
    row += 1
    for label, value in [
        ("Sales rep", sales.full_name if sales else "—"),
        ("Email",     (sales.contact_email or sales.email) if sales else "—"),
        ("Phone",     ((sales.phone if sales else None) or (sales.whatsapp_id if sales else None) or "—")),
    ]:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    buf = BytesIO()
    wb.save(buf)
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="quotation-{q.number}.xlsx"',
        },
    )


@router.get("/{q_id}", response_model=QuotationOut)
async def get_quotation(
    q_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = await _load(q_id, db)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Quotation not found")
    if not await _may_see(db, user, q):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    # Surface the sales rep's name so the detail page can show who owns the deal.
    if q.sales_pic_id:
        rep = await db.get(User, q.sales_pic_id)
        if rep:
            q.sales_pic_name = rep.full_name
    return q


# ─── Follow-ups ──────────────────────────────────────────────────────────────

class FollowUpIn(BaseModel):
    notes: str
    next_at: datetime | None = None
    next_channel: str = "dashboard"


async def _scoped_quote(q_id: UUID, db: AsyncSession, user: User) -> Quotation:
    q = await db.get(Quotation, q_id)
    if not q:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Quotation not found")
    if not await _may_see(db, user, q):
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Out of scope")
    return q


@router.post("/{q_id}/followup", status_code=201)
async def log_followup(
    q_id: UUID,
    payload: FollowUpIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = await _scoped_quote(q_id, db, user)
    if not payload.notes or not payload.notes.strip():
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Notes required.")

    # Sales follow-ups are routed through the director: the activity + reminder
    # are only created once the request is approved (see core/approval.py).
    if Role(user.role) == Role.SALES:
        req = await request_approval(
            db,
            target_type="followup",
            target_id=q.customer_id,
            requested_by=user.id,
            required_role=Role.DIRECTOR,
            reason=f"Follow-up on quotation {q.number}",
            payload={
                "source": "quotation",
                "quotation_id": str(q_id),
                "quotation_number": q.number,
                "notes": payload.notes,
                "next_at": payload.next_at.isoformat() if payload.next_at else None,
                "next_channel": payload.next_channel,
            },
        )
        await db.flush()
        return JSONResponse(
            status_code=status.HTTP_202_ACCEPTED,
            content={
                "status": "pending_approval",
                "approval_request_id": str(req.id),
                "message": "Follow-up sent to the director for approval.",
            },
        )

    activity = Activity(
        customer_id=q.customer_id,
        user_id=user.id,
        type="follow_up",
        direction="outbound",
        occurred_at=datetime.now(UTC),
        notes=payload.notes,
        meta={"quotation_id": str(q_id), "quotation_number": q.number},
    )
    db.add(activity)
    await db.flush()

    reminder_id: str | None = None
    if payload.next_at:
        rem = Reminder(
            customer_id=q.customer_id,
            user_id=user.id,
            kind="follow_up",
            due_at=payload.next_at,
            channel=payload.next_channel,
            message=f"Follow up on quotation {q.number}",
            status="pending",
        )
        db.add(rem)
        await db.flush()
        reminder_id = str(rem.id)

    return {
        "activity_id": str(activity.id),
        "reminder_id": reminder_id,
        "occurred_at": activity.occurred_at,
    }


@router.get("/{q_id}/followups")
async def list_followups(
    q_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = await _scoped_quote(q_id, db, user)

    # Activities: anything tagged with this quotation_id, plus generic
    # follow_up / quotation_sent / negotiation on the same customer after
    # the quotation was created.
    act_stmt = (
        select(Activity)
        .where(Activity.customer_id == q.customer_id)
        .order_by(Activity.occurred_at.desc())
        .limit(200)
    )
    activities = []
    for a in (await db.scalars(act_stmt)).all():
        tagged = bool(a.meta) and a.meta.get("quotation_id") == str(q_id)
        is_related_type = a.type in ("follow_up", "quotation_sent", "negotiation")
        after_quote = a.occurred_at >= q.created_at
        if tagged or (is_related_type and after_quote):
            activities.append({
                "id": str(a.id),
                "type": a.type,
                "direction": a.direction,
                "occurred_at": a.occurred_at,
                "notes": a.notes,
                "user_id": str(a.user_id) if a.user_id else None,
                "tagged": tagged,
            })

    # Pending reminders for the same customer
    rem_stmt = (
        select(Reminder)
        .where(Reminder.customer_id == q.customer_id, Reminder.status == "pending")
        .order_by(Reminder.due_at.asc())
    )
    reminders = [
        {
            "id": str(r.id),
            "kind": r.kind,
            "channel": r.channel,
            "due_at": r.due_at,
            "message": r.message,
        }
        for r in (await db.scalars(rem_stmt)).all()
    ]
    return {"activities": activities, "reminders": reminders}


@router.patch("/{q_id}/reminders/{reminder_id}/done")
async def followup_reminder_done(
    q_id: UUID,
    reminder_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    q = await _scoped_quote(q_id, db, user)
    rem = await db.get(Reminder, reminder_id)
    if not rem:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Reminder not found")
    # The reminder has to belong to THIS quotation's customer — otherwise
    # owning one quote let you close any other rep's follow-ups and stage
    # tasks just by knowing an id.
    if rem.customer_id != q.customer_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Reminder not found")
    rem.status = "done"
    return {"ok": True}


# ─── Account links + ledger posting ──────────────────────────────────────────


@router.get("/{q_id}/account-links")
async def get_account_links(
    q_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_ledger_viewer),
):
    q = await _scoped_quote(q_id, db, user)
    amounts = ledger.compute_amounts(q)

    out_links = []
    for role, default in ledger.DEFAULTS.items():
        no = getattr(q, f"account_{role}_no") or default
        acc = await db.scalar(select(Account).where(Account.account_no == no))
        out_links.append({
            "role": role,
            "account_no": no,
            "account_name": acc.name if acc else None,
            "account_type": acc.account_type if acc else None,
            "current_balance": float(acc.balance or 0) if acc else None,
            "amount_to_post": amounts.get(role, 0),
        })
    return {
        "links": out_links,
        "is_posted": q.is_posted,
        "posted_at": q.posted_at,
        "posted_snapshot": q.posted_snapshot,
        "amounts": amounts,
    }


@router.patch("/{q_id}/account-links", response_model=QuotationOut)
async def update_account_links(
    q_id: UUID,
    payload: QuotationAccountLinks,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_ledger_viewer),
):
    q = await _scoped_quote(q_id, db, user)
    if q.is_posted:
        raise HTTPException(
            status.HTTP_409_CONFLICT,
            "Cannot change account links after posting. Reverse first.",
        )
    data = payload.model_dump(exclude_unset=True)
    for k, v in data.items():
        setattr(q, k, v)
    await db.flush()
    return await _load(q.id, db)


@router.post("/{q_id}/post-ledger", response_model=QuotationOut)
async def post_to_ledger(
    q_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_ledger_viewer),
):
    q = await _scoped_quote(q_id, db, user)
    await ledger.post_quotation(db, q)
    return await _load(q.id, db)


@router.post("/{q_id}/reverse-ledger", response_model=QuotationOut)
async def reverse_from_ledger(
    q_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_ledger_viewer),
):
    q = await _scoped_quote(q_id, db, user)
    await ledger.reverse_quotation(db, q)
    return await _load(q.id, db)

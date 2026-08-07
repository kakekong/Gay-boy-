"""Undoing an import, without undoing anything else.

Importing hundreds of rows into a live system is only a reasonable thing to do
if getting them back out is one action. Ticking 87 customers by hand at nine at
night because a column mapped wrong is not a plan.

But "one action" must not mean "one careless action". By the time somebody
reaches for undo, staff may have started working on what arrived — a price
request against an imported customer, a customer PO against an imported
quotation. Deleting the parent takes that work with it, and *that* is the
damage worth engineering against, not the import itself.

So undo splits the run in two:

  removable    nothing has been built on it — it goes
  blocked      somebody has filed work against it since — it stays, and the
               preview names both the record and what is hanging off it

Taking the blocked ones as well needs a second, explicit yes. The difference
between reversing a mistake and causing a new one is exactly that yes.

The other property this pins is that undo is *narrow*: a second import run, and
records that were already in the system before either ran, must be untouched.
That is asserted against records built alongside, checked one at a time rather
than by counting — a count can be right while the wrong rows went.
"""
import asyncio, io, os, sys, uuid
os.environ.update(DATABASE_URL="postgresql+asyncpg://postgres@127.0.0.1:55432/transmisi_test",
    APP_ENV="dev", DEMO_SEED_PASSWORD="test-pass-123",
    STORAGE_LOCAL_DIR="/tmp/storage_test", JWT_SECRET="e2e-test-secret")
sys.path.insert(0, "/home/user/Gay-boy-/backend")
import httpx, logging; logging.disable(logging.INFO)
PASS, FAIL = [], []
def check(n, c, d=""):
    (PASS if c else FAIL).append(n); print(("  PASS " if c else "  FAIL ")+n+(f"  [{d}]" if d and not c else ""))
def J(r):
    try: return r.json()
    except Exception: return {"_": r.text[:220]}

XL = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CUST_HEADER = ("ID Pelanggan,Kategori,Nama,Kontak,No. Telp. Bisnis,Handphone,Email,"
               "Alamat Penagihan,Kota,Provinsi,Kode Pos,Alamat Pengiriman,"
               "Syarat Pembayaran,NPWP,Nama Wajib Pajak,Alamat (Pajak),Catatan")
ITEM_HEADER = ["No.", "Kategori Barang", "Kode Barang", "Nama Barang",
               "Jenis Barang", "Satuan", "Harga Beli", "Def. Hrg. Jual Satuan #1",
               "Batas Minimum Stok", "Kuantitas Saldo Awal", "Gudang Saldo Awal",
               "Pemasok Utama", "Merek Barang", "Catatan", "Non Aktif"]


def customers_csv(tag: str, n: int) -> bytes:
    rows = [f"C.{i:03d}{tag},Umum,PT Impor {tag}-{i},Pak {i},021-{i},,,"
            f"Jl. Uji {i},Jakarta,DKI,11440,,net 30,,,,"
            for i in range(1, n + 1)]
    return ("\n".join([CUST_HEADER, *rows]) + "\n").encode()


def items_xlsx(tag: str, n: int) -> bytes:
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.title = "Barang & Jasa"
    ws.append(ITEM_HEADER)
    for i in range(1, n + 1):
        ws.append([str(i), "CHAIN", f"SKU-{tag}-{i}", f"Part {tag} {i}", "INV",
                   "PCS", "0", "0", "0", "0", "", "", "", "", "TIDAK"])
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


async def main():
    from app.scripts.seed import ensure_schema; await ensure_schema()
    from app.main import app
    c = httpx.AsyncClient(transport=httpx.ASGITransport(app=app),
                          base_url="http://t/api/v1", timeout=180)

    async def login(e):
        r = await c.post("/auth/login", json={"email": e, "password": "test-pass-123"})
        return {"Authorization": f"Bearer {r.json()['access_token']}"}

    tag = uuid.uuid4().hex[:5]
    d = await login("director@demo.local")
    s1 = await login("sales1@demo.local")
    pur = await login("purchasing@demo.local")

    async def alive(kind, oid):
        path = {"customer": "/customers", "price_request": "/price-requests",
                "quotation": "/quotations", "inventory_item": None}[kind]
        if kind == "inventory_item":
            r = J(await c.get("/inventory", headers=d, params={"q": oid}))
            rows = r.get("data") if isinstance(r, dict) else r
            return bool(rows)
        return (await c.get(f"{path}/{oid}", headers=d)).status_code == 200

    # ── a customer typed in by hand, before any import ───────────────────────
    hand = J(await c.post("/customers", headers=s1, json={
        "company_name": f"PT Ditulis Tangan {tag}", "industry": "mining"}))["id"]

    # ══ run one ══════════════════════════════════════════════════════════════
    print("\n── importing, then undoing ──")
    data = customers_csv(tag, 6)
    up = lambda: {"file": (f"pelanggan-{tag}.csv", data, "text/csv")}
    r1 = J(await c.post("/imports/customers/commit", headers=d, files=up(),
                        data={"limit": 999, "confirm": "IMPORT"}))
    check("the import reports a run id", bool(r1.get("run_id")), str(r1)[:170])
    check("...and created what it said", r1.get("created") == 6, str(r1.get("created")))
    run1 = r1["run_id"]
    made1 = [x["id"] for x in r1["customers"]]

    runs = J(await c.get("/imports/runs", headers=d))
    mine = next((x for x in runs if x["id"] == run1), None)
    check("the run is listed", mine is not None, str(runs)[:180])
    check("...with its file, kind and count",
          mine and mine["kind"] == "customers" and mine["created_count"] == 6
          and f"pelanggan-{tag}" in (mine["filename"] or ""), str(mine))
    check("...and who ran it", mine and mine["by"] != "—", str(mine and mine["by"]))

    p = J(await c.get(f"/imports/runs/{run1}/undo-preview", headers=d))
    check("undo would remove all six", p["removable"] == 6 and p["blocked"] == 0,
          f"{p['removable']} / {p['blocked']}")
    check("...and nothing else comes with them", not p["dependents"],
          str(p["dependents"])[:160])

    # only a director
    r = await c.post(f"/imports/runs/{run1}/undo", headers=s1,
                     data={"confirm": "UNDO IMPORT"})
    check("sales cannot undo an import", r.status_code == 403, str(r.status_code))
    r = await c.get("/imports/runs", headers=s1)
    check("...nor even see the runs", r.status_code == 403, str(r.status_code))
    r = await c.post(f"/imports/runs/{run1}/undo", headers=d, data={"confirm": "yes"})
    check("the confirmation phrase is required", r.status_code == 400, str(r.status_code))
    check("...and nothing went in the meantime", await alive("customer", made1[0]))

    res = J(await c.post(f"/imports/runs/{run1}/undo", headers=d,
                         data={"confirm": "UNDO IMPORT"}))
    check("undo removes them", res.get("removed") == 6, str(res)[:180])
    gone = [not await alive("customer", i) for i in made1]
    check("...every one of them", all(gone), str(gone))
    check("the hand-typed customer is untouched", await alive("customer", hand))

    again = J(await c.get(f"/imports/runs/{run1}/undo-preview", headers=d))
    check("the run now shows nothing left", again["still_present"] == 0,
          str(again["still_present"]))
    res2 = J(await c.post(f"/imports/runs/{run1}/undo", headers=d,
                          data={"confirm": "UNDO IMPORT"}))
    check("undoing twice is harmless", res2.get("removed", 0) == 0, str(res2)[:150])
    listed = J(await c.get("/imports/runs", headers=d))
    check("...and the run is marked undone, not erased",
          any(x["id"] == run1 and x["undone_at"] for x in listed), "no undone_at")

    # ══ work filed against an import ═════════════════════════════════════════
    print("\n── when somebody has already used the imported data ──")
    data2 = customers_csv(f"{tag}b", 4)
    up2 = lambda: {"file": (f"pelanggan2-{tag}.csv", data2, "text/csv")}
    r2 = J(await c.post("/imports/customers/commit", headers=d, files=up2(),
                        data={"limit": 999, "confirm": "IMPORT"}))
    run2 = r2["run_id"]
    made2 = [x["id"] for x in r2["customers"]]
    check("a second run imports cleanly", r2.get("created") == 4, str(r2.get("created")))

    # Real work filed against one of them, since. Raised by the director
    # because an imported customer arrives with no sales rep, and a rep may
    # only raise a price request for a customer of their own.
    busy = made2[0]
    pr_res = J(await c.post("/price-requests", headers=d, json={
        "customer_id": busy,
        "items": [{"description": f"Chain {tag}", "qty": 5, "uom": "meter"}]}))
    check("work can be filed against an imported customer",
          bool(pr_res.get("id")), str(pr_res)[:180])
    pr = pr_res["id"]
    check("a price request now exists against an imported customer",
          await alive("price_request", pr))

    p2 = J(await c.get(f"/imports/runs/{run2}/undo-preview", headers=d))
    check("undo now reports one blocked record",
          p2["blocked"] == 1 and p2["removable"] == 3,
          f"removable={p2['removable']} blocked={p2['blocked']}")
    check("...and names what is hanging off it",
          any(x["type"] == "price_request" for x in p2["dependents"]),
          str(p2["dependents"])[:170])

    res3 = J(await c.post(f"/imports/runs/{run2}/undo", headers=d,
                          data={"confirm": "UNDO IMPORT"}))
    check("a plain undo removes only the untouched ones",
          res3.get("removed") == 3 and res3.get("left_alone") == 1, str(res3)[:180])
    check("the one with work filed against it stays", await alive("customer", busy))
    check("...and so does the work itself", await alive("price_request", pr))
    others = [not await alive("customer", i) for i in made2[1:]]
    check("...while the other three are gone", all(others), str(others))

    # taking it needs its own yes
    p3 = J(await c.get(f"/imports/runs/{run2}/undo-preview", headers=d))
    check("the preview still offers the blocked one", p3["blocked"] == 1, str(p3["blocked"]))
    r = await c.post(f"/imports/runs/{run2}/undo", headers=d,
                     data={"confirm": "UNDO IMPORT"})
    check("undoing again refuses rather than doing nothing quietly",
          r.status_code == 409, str(r.status_code))
    check("...explaining that work has been filed since",
          "filed against" in str(J(r)), str(J(r))[:170])

    res4 = J(await c.post(f"/imports/runs/{run2}/undo", headers=d,
                          data={"confirm": "UNDO IMPORT", "include_dependents": "true"}))
    check("with the box ticked it takes the rest", res4.get("removed") == 1, str(res4)[:180])
    check("...and the customer is gone", not await alive("customer", busy))
    check("...along with the price request that hung off it",
          not await alive("price_request", pr))
    check("the hand-typed customer is STILL untouched", await alive("customer", hand))

    # ══ undo is per-run, not per-import-ever ═════════════════════════════════
    print("\n── two runs at once ──")
    a = J(await c.post("/imports/items/commit", headers=d,
                       files={"file": ("a.xlsx", items_xlsx(f"{tag}A", 3), XL)},
                       data={"limit": 99, "confirm": "IMPORT"}))
    b = J(await c.post("/imports/items/commit", headers=d,
                       files={"file": ("b.xlsx", items_xlsx(f"{tag}B", 3), XL)},
                       data={"limit": 99, "confirm": "IMPORT"}))
    check("two item runs, two run ids", a["run_id"] != b["run_id"], "same id")
    res5 = J(await c.post(f"/imports/runs/{a['run_id']}/undo", headers=d,
                          data={"confirm": "UNDO IMPORT"}))
    check("undoing the first removes three parts", res5.get("removed") == 3, str(res5)[:150])
    gone_a = [not await alive("inventory_item", f"SKU-{tag}A-{i}") for i in (1, 2, 3)]
    check("...its parts are gone", all(gone_a), str(gone_a))
    still_b = [await alive("inventory_item", f"SKU-{tag}B-{i}") for i in (1, 2, 3)]
    check("...and the second run's parts are all still there", all(still_b), str(still_b))

    # ══ a run that imported quotations ═══════════════════════════════════════
    print("\n── quotations, which hang off customers ──")
    cust = J(await c.post("/customers", headers=s1, json={
        "company_name": f"PT Penawaran {tag}", "industry": "sugar"}))["id"]
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    qno = f"SQ.{tag}.900"
    ws = wb.create_sheet(title=qno[:31])
    ws.append(["Nomor", "Tanggal", "Pelanggan", "Nama Barang",
               "Kuantitas", "@Harga", "Total Harga"])
    ws.append([qno, "2024-01-02 00:00:00", f"PT Penawaran {tag}", "Chain", 2, 500_000, 1_000_000])
    ws.append(["", "", "", "", "", "Sub Total", 1_000_000])
    buf = io.BytesIO(); wb.save(buf)
    rq = J(await c.post("/imports/quotations/commit", headers=d,
                        files={"file": ("q.xlsx", buf.getvalue(), XL)},
                        data={"limit": 9, "confirm": "IMPORT"}))
    check("the quotation imports", rq.get("created") == 1, str(rq)[:160])
    qid = rq["quotations"][0]["id"]
    res6 = J(await c.post(f"/imports/runs/{rq['run_id']}/undo", headers=d,
                          data={"confirm": "UNDO IMPORT"}))
    check("undoing it removes the quotation", res6.get("removed") == 1, str(res6)[:150])
    check("...the quotation is gone", not await alive("quotation", qid))
    check("...but the customer it was filed against stays",
          await alive("customer", cust))

    # ══ and an unknown run ═══════════════════════════════════════════════════
    r = await c.get(f"/imports/runs/{uuid.uuid4()}/undo-preview", headers=d)
    check("an unknown run is a clean 404", r.status_code == 404, str(r.status_code))

    await c.aclose()
    print(f"\n{len(PASS)} passed, {len(FAIL)} failed")
    if FAIL:
        sys.exit(1)


asyncio.run(main())

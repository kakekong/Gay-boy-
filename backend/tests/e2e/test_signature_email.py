"""A contact address that is not the login, and a signature that fits.

Two things asked for together, and they belong together: both are about the
person as the customer sees them, rather than as the system authenticates
them.

**The email.** `User.email` is a credential. It was also what got printed in
the signature block of every quotation PDF and Excel export — so a customer
replying to a quote replied to `sales1@demo.local`. A separate
`contact_email` now carries the address they actually correspond from, and
the documents prefer it. Deliberately *not* unique and *never* accepted at
login, both of which are checked here: a contact detail that can be used to
sign in is a second credential nobody knows they have.

**The signature.** Both documents already left a gap above the name for a wet
signature. The scan now goes in that gap — scaled to fit the block it is
going into, not to a size baked into the image. The two blocks are different
shapes (the quotation's is 62mm wide and short, the customer PO's narrower
and taller), so the same upload has to come out right in both. That is what
this checks: the same image, in both documents, each still a valid PDF whose
page count did not change — because a signature that pushes the name onto a
second page is worse than no signature.
"""
import asyncio, io, os, sys, uuid
os.environ.update(DATABASE_URL="postgresql+asyncpg://postgres@127.0.0.1:55432/transmisi_test",
    APP_ENV="dev", DEMO_SEED_PASSWORD="test-pass-123",
    STORAGE_LOCAL_DIR="/tmp/storage_test", JWT_SECRET="e2e-test-secret")
sys.path.insert(0, "/home/user/Gay-boy-/backend")
import httpx, logging; logging.disable(logging.INFO)
PASS, FAIL = [], []
def check(n, c, d=""):
    (PASS if c else FAIL).append(n); print(("  PASS " if c else "  FAIL ")+n+(f"  [{d}]" if d and not c else ""))
def J(r):
    try: return r.json()
    except Exception: return {"_": r.text[:200]}


def png(w: int, h: int, transparent: bool = True) -> bytes:
    """A stand-in signature: a dark stroke on a transparent background."""
    from PIL import Image, ImageDraw
    img = Image.new("RGBA", (w, h), (0, 0, 0, 0) if transparent else (255, 255, 255, 255))
    d = ImageDraw.Draw(img)
    d.line([(int(w * .05), int(h * .7)), (int(w * .35), int(h * .2)),
            (int(w * .6), int(h * .8)), (int(w * .95), int(h * .3))],
           fill=(20, 20, 90, 255), width=max(2, h // 12))
    buf = io.BytesIO(); img.save(buf, "PNG")
    return buf.getvalue()


def pages(data: bytes) -> int:
    from pypdf import PdfReader
    return len(PdfReader(io.BytesIO(data)).pages)


def pdf_text(data: bytes) -> str:
    from pypdf import PdfReader
    return "\n".join(p.extract_text() or "" for p in PdfReader(io.BytesIO(data)).pages)


async def main():
    from app.scripts.seed import ensure_schema; await ensure_schema()
    from app.main import app
    c = httpx.AsyncClient(transport=httpx.ASGITransport(app=app),
                          base_url="http://t/api/v1", timeout=180)
    tag = uuid.uuid4().hex[:5]

    async def login(e, pw="test-pass-123"):
        r = await c.post("/auth/login", json={"email": e, "password": pw})
        return {"Authorization": f"Bearer {r.json()['access_token']}"}
    d = await login("director@demo.local")
    pur = await login("purchasing@demo.local")

    # ══ the contact address ══════════════════════════════════════════════════
    print("\n── an address separate from the login ──")
    rep_emp = J(await c.post("/employees", headers=d, json={
        "full_name": f"Rep Tanda {tag}", "intended_role": "sales"}))
    rep = J(await c.post("/users", headers=d, json={
        "email": f"rep-{tag}@demo.local", "full_name": f"Rep Tanda {tag}",
        "role": "sales", "password": "test-pass-123",
        "employee_id": rep_emp["id"],
        "contact_email": f"Rep.Tanda+{tag}@transmisi.co.id",
        "phone": "081234567890"}))
    check("a user can be created with both", rep.get("id") is not None, str(rep)[:150])
    got = J(await c.get(f"/users/{rep['id']}", headers=d))
    check("...the login is the login", got["email"] == f"rep-{tag}@demo.local",
          got["email"])
    check("...and the contact address is kept, lower-cased",
          got.get("contact_email") == f"rep.tanda+{tag}@transmisi.co.id",
          str(got.get("contact_email")))

    r = await c.post("/auth/login", json={"email": f"rep.tanda+{tag}@transmisi.co.id",
                                          "password": "test-pass-123"})
    check("the contact address cannot be used to sign in",
          r.status_code in (400, 401), str(r.status_code))
    r = await c.post("/auth/login", json={"email": f"rep-{tag}@demo.local",
                                          "password": "test-pass-123"})
    check("...and the real login still works", r.status_code == 200, str(r.status_code))

    r = await c.patch(f"/users/{rep['id']}", headers=d,
                      json={"contact_email": "not an address"})
    check("a malformed address is refused", r.status_code == 400, str(r.status_code))
    check("...and says which value was wrong", "not an address" in str(J(r)),
          str(J(r))[:130])
    await c.patch(f"/users/{rep['id']}", headers=d, json={"contact_email": ""})
    check("...and it can be cleared",
          J(await c.get(f"/users/{rep['id']}", headers=d)).get("contact_email") is None)
    await c.patch(f"/users/{rep['id']}", headers=d,
                  json={"contact_email": f"rep.tanda+{tag}@transmisi.co.id"})

    # Two people may share one mailbox — it is a contact detail, not an identity.
    rep2_emp = J(await c.post("/employees", headers=d, json={
        "full_name": f"Rep Dua {tag}", "intended_role": "sales"}))
    rep2 = J(await c.post("/users", headers=d, json={
        "email": f"rep2-{tag}@demo.local", "full_name": f"Rep Dua {tag}",
        "role": "sales", "password": "test-pass-123",
        "employee_id": rep2_emp["id"],
        "contact_email": f"rep.tanda+{tag}@transmisi.co.id"}))
    check("two people can share one contact address", rep2.get("id") is not None,
          str(rep2)[:150])
    clash_emp = J(await c.post("/employees", headers=d, json={
        "full_name": f"Clash {tag}", "intended_role": "sales"}))
    r = await c.post("/users", headers=d, json={
        "email": f"rep-{tag}@demo.local", "full_name": "Clash",
        "role": "sales", "employee_id": clash_emp["id"],
        "password": "test-pass-123"})
    check("...while a duplicate LOGIN is still refused",
          r.status_code in (400, 409), str(r.status_code))

    # ══ the signature ════════════════════════════════════════════════════════
    print("\n── the scan, and what it refuses ──")
    rp = await login(f"rep-{tag}@demo.local")
    files = {"file": ("sig.png", png(600, 200), "image/png")}
    r = await c.post(f"/users/{rep['id']}/signature", headers=rp, files=files)
    check("a rep can upload their own signature", r.status_code == 200,
          f"{r.status_code} {J(r)}"[:150])
    r = await c.get(f"/users/{rep['id']}/signature", headers=rp)
    check("...and read it back", r.status_code == 200 and len(r.content) > 100,
          f"{r.status_code} {len(r.content)}")

    r = await c.post(f"/users/{rep['id']}/signature", headers=d,
                     files={"file": ("x.png", b"not an image at all", "image/png")})
    check("a file that is not an image is refused", r.status_code == 400,
          str(r.status_code))
    check("...and says what to upload instead", "PNG" in str(J(r)), str(J(r))[:130])
    r = await c.post(f"/users/{rep['id']}/signature", headers=d,
                     files={"file": ("t.png", png(20, 8), "image/png")})
    check("one too small to print is refused", r.status_code == 400, str(r.status_code))
    check("...naming its size", "20×8" in str(J(r)) or "20" in str(J(r)),
          str(J(r))[:130])
    big = png(4000, 4000)
    if len(big) > 2 * 1024 * 1024:
        r = await c.post(f"/users/{rep['id']}/signature", headers=d,
                         files={"file": ("b.png", big, "image/png")})
        check("...and one far too large", r.status_code == 400, str(r.status_code))

    s2 = await login("sales2@demo.local")
    r = await c.post(f"/users/{rep['id']}/signature", headers=s2, files=files)
    check("nobody can set somebody else's signature", r.status_code == 403,
          str(r.status_code))
    r = await c.post(f"/users/{rep['id']}/signature", headers=d, files=files)
    check("...but the director can, for anyone", r.status_code == 200,
          str(r.status_code))

    # ══ on the documents ═════════════════════════════════════════════════════
    print("\n── on the two documents, each its own shape ──")
    me_rep = J(await c.get("/auth/me", headers=rp))
    cust = J(await c.post("/customers", headers=d, json={
        "company_name": f"PT Tanda {tag}", "industry": "mining"}))["id"]
    await c.post("/customers/reassign", headers=d, json={
        "customer_ids": [cust], "sales_pic_id": me_rep["id"]})

    pr = J(await c.post("/price-requests", headers=rp, json={
        "customer_id": cust,
        "items": [{"description": f"CHAIN {tag}", "qty": 40, "uom": "meter"}]}))["id"]
    await c.post(f"/price-requests/{pr}/submit", headers=rp)
    await c.post(f"/price-requests/{pr}/price", headers=pur, json={
        "items": [{"line_no": 1, "cost_price": 2_000_000, "basis": "unit"}]})
    await c.post(f"/price-requests/{pr}/approve", headers=d, json={
        "items": [{"line_no": 1, "sell_price": 2_556_061, "basis": "unit"}]})
    q = J(await c.post(f"/quotations/from-price-request/{pr}", headers=rp))

    r = await c.get(f"/quotations/{q['id']}/export.pdf", headers=rp)
    check("the quotation PDF builds with a signature on it",
          r.status_code == 200 and r.content[:4] == b"%PDF",
          f"{r.status_code} {r.content[:20]}")
    signed_pdf = r.content
    txt = pdf_text(signed_pdf)
    check("...printing the contact address, not the login",
          f"rep.tanda+{tag}@transmisi.co.id" in txt.lower(), txt[-400:])
    check("...and never the login", f"rep-{tag}@demo.local" not in txt.lower(),
          txt[-400:])
    check("...still one page — the image did not push the name off it",
          pages(signed_pdf) == 1, str(pages(signed_pdf)))

    excel = await c.get(f"/quotations/{q['id']}/export.xlsx", headers=rp)
    def xlsx_text(data: bytes) -> str:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data))
        return "\n".join(
            str(cell.value) for ws in wb.worksheets
            for row in ws.iter_rows() for cell in row if cell.value is not None)
    check("the Excel export carries the contact address too",
          excel.status_code == 200
          and f"rep.tanda+{tag}@transmisi.co.id" in xlsx_text(excel.content).lower(),
          str(excel.status_code))

    # the same scan on the other document, whose block is a different shape
    await c.post(f"/quotations/{q['id']}/submit", headers=rp)
    await c.post(f"/quotations/{q['id']}/approve", headers=d, json={})
    await c.post(f"/quotations/{q['id']}/won", headers=d)
    po = J(await c.post("/customer-pos", headers=rp, json={
        "customer_id": cust, "quotation_id": q["id"], "number": f"PO-SIG-{tag}",
        "po_date": "2026-08-01",
        "items": [{"description": "chain", "qty": 10, "unit_price": 140_000}]}))
    r = await c.get(f"/customer-pos/{po['id']}/export.pdf", headers=rp)
    check("the order confirmation builds with the same scan",
          r.status_code == 200 and r.content[:4] == b"%PDF",
          f"{r.status_code} {r.content[:20]}")
    check("...and it is still one page", pages(r.content) == 1, str(pages(r.content)))
    with_sig = len(r.content)

    # ══ and without one, nothing breaks ══════════════════════════════════════
    print("\n── and a document from somebody with no signature ──")
    r = await c.delete(f"/users/{rep['id']}/signature", headers=rp)
    check("the signature can be removed", r.status_code == 204, str(r.status_code))
    r = await c.get(f"/users/{rep['id']}/signature", headers=rp)
    check("...and is gone", r.status_code == 404, str(r.status_code))
    r = await c.get(f"/quotations/{q['id']}/export.pdf", headers=rp)
    check("the quotation still builds, with a blank space to sign",
          r.status_code == 200 and r.content[:4] == b"%PDF", str(r.status_code))
    check("...one page, as before", pages(r.content) == 1, str(pages(r.content)))
    check("...and smaller than the signed one, since no image is embedded",
          len(r.content) < len(signed_pdf),
          f"{len(r.content)} vs {len(signed_pdf)}")
    r2 = await c.get(f"/customer-pos/{po['id']}/export.pdf", headers=rp)
    check("the order confirmation too",
          r2.status_code == 200 and len(r2.content) < with_sig,
          f"{r2.status_code} {len(r2.content)} vs {with_sig}")

    # ══ a wildly-shaped scan still fits ══════════════════════════════════════
    print("\n── an awkwardly-shaped scan ──")
    for label, (w, h) in [("very wide", (2000, 120)), ("tall", (200, 900)),
                          ("square", (400, 400))]:
        await c.post(f"/users/{rep['id']}/signature", headers=d,
                     files={"file": ("s.png", png(w, h), "image/png")})
        rq = await c.get(f"/quotations/{q['id']}/export.pdf", headers=rp)
        rp_ = await c.get(f"/customer-pos/{po['id']}/export.pdf", headers=rp)
        check(f"a {label} signature fits both documents on one page",
              rq.status_code == 200 and pages(rq.content) == 1
              and rp_.status_code == 200 and pages(rp_.content) == 1,
              f"quote={pages(rq.content) if rq.status_code == 200 else rq.status_code} "
              f"po={pages(rp_.content) if rp_.status_code == 200 else rp_.status_code}")

    # Decide the PO this driver filed. Leaving it pending would park a live
    # approval request in the queue for whatever runs next — and the sheet
    # has to keep printing once it is approved, which is worth checking.
    r = await c.post(f"/customer-pos/{po['id']}/approve", headers=d, json={"notes": ""})
    check("the PO this driver filed can be approved and cleared",
          r.status_code == 200, f"{r.status_code} {J(r)}"[:140])
    r = await c.get(f"/customer-pos/{po['id']}/export.pdf", headers=rp)
    check("...and its sheet still prints afterwards",
          r.status_code == 200 and pages(r.content) == 1, str(r.status_code))

    await c.aclose()
    print(f"\n{len(PASS)} passed, {len(FAIL)} failed")
    if FAIL:
        sys.exit(1)


asyncio.run(main())

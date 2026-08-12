"""Purchasing module — PR → RFQ → PO → GR → QC → Payment.

Stubs scaffolded; full implementation follows the same pattern as quotations.
"""

from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, status
from pydantic import BaseModel
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.approval import request_approval, require_pr_approval
from app.core.audit import record as audit_record
from app.core.config import settings
from app.core.db import get_db
from app.core.deps import get_current_user
from app.core.permissions import Role, require, require_min
from app.models.purchasing import Supplier
from app.models.user import User

router = APIRouter(
    # Internal-only surface. External portal accounts (customer /
    # supplier, hierarchy tier 0) must never reach the CRM, pricing,
    # calendar or notification data — they have /portal/* instead.
    dependencies=[Depends(require_min(Role.SALES))]
)

# The supplier directory belongs to the department that deals with suppliers.
# Onboarding one was management-only for a while; it is not a decision anybody
# else is in a position to make — purchasing is the one talking to the vendor
# when the vendor first needs to exist, and making them ask somebody else to
# type a name in just meant the row got created late or not at all.
_supplier_editors = require(Role.ADMIN, Role.DIRECTOR, Role.MANAGER, Role.PURCHASING)


# ─── Suppliers ───────────────────────────────────────────────────────────────

class SupplierContactIn(BaseModel):
    name: str
    position: str | None = None
    phone: str | None = None
    whatsapp: str | None = None
    email: str | None = None
    is_primary: bool = False
    notes: str | None = None


class SupplierIn(BaseModel):
    name: str
    category: str | None = None
    rating: float = 0
    # Where they are. `warehouse_address` is where the goods are actually
    # collected from, which is regularly not the office on the letterhead.
    company_address: str | None = None
    warehouse_address: str | None = None
    # The company's own line and mailbox, not a person's.
    phone: str | None = None
    whatsapp: str | None = None
    email: str | None = None
    # The people. Submitted with the form so a supplier arrives complete
    # rather than needing a second visit to the page to be usable.
    contacts: list[SupplierContactIn] = []
    contact: dict = {}          # legacy blob; still accepted


class SupplierPatch(BaseModel):
    """Everything on the header that can be corrected after the fact.

    Deliberately not `SupplierIn`: contacts are their own rows with their own
    endpoints, and a PATCH that carried them would have to decide what an
    omitted list means. The scores (lead time, QC failure, volatility) are
    computed from delivery history and are not editable by hand.
    """
    name: str | None = None
    category: str | None = None
    rating: float | None = None
    company_address: str | None = None
    warehouse_address: str | None = None
    phone: str | None = None
    whatsapp: str | None = None
    email: str | None = None


def _supplier_contact_out(c) -> dict:
    return {
        "id": str(c.id),
        "supplier_id": str(c.supplier_id),
        "name": c.name,
        "position": c.position,
        "phone": c.phone,
        "whatsapp": c.whatsapp,
        "email": c.email,
        "is_primary": c.is_primary,
        "notes": c.notes,
        "created_at": c.created_at,
    }


@router.get("/suppliers")
async def list_suppliers(
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(get_current_user),
):
    rows = (await db.scalars(
        select(Supplier).order_by(Supplier.name.asc())
    )).all()
    return [
        {
            "id": str(s.id),
            "name": s.name,
            "category": s.category,
            "rating": float(s.rating or 0),
            "lead_time_days_avg": float(s.lead_time_days_avg or 0),
            "qc_fail_rate": float(s.qc_fail_rate or 0),
            # Enough to ring them from the directory without opening the row.
            "phone": s.phone or (s.contact or {}).get("phone"),
            "email": s.email or (s.contact or {}).get("email"),
            "company_address": s.company_address,
        }
        for s in rows
    ]


@router.get("/suppliers/{supplier_id}")
async def get_supplier(
    supplier_id: UUID,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(get_current_user),
):
    """Supplier detail with a recap of POs we've issued, the projects we
    source from them, incoming GR/QC history, and any files they've uploaded.
    Used by the supplier detail screen."""
    from app.models.attachment import Attachment
    from app.models.operation import Project
    from app.models.purchasing import (
        GoodsReceipt, QCReport, Supplier, SupplierContact, SupplierPO,
    )

    s = await db.get(Supplier, supplier_id)
    if not s:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Supplier not found")
    po_rows = (await db.scalars(
        select(SupplierPO)
        .where(SupplierPO.supplier_id == supplier_id)
        .order_by(SupplierPO.created_at.desc())
    )).all()
    open_pos = [p for p in po_rows if p.status in ("open", "pending_approval")]
    po_ids = [p.id for p in po_rows]
    po_number = {p.id: p.number for p in po_rows}

    # Projects this supplier feeds (distinct, via their POs)
    project_ids = {p.project_id for p in po_rows if p.project_id}
    projects = []
    if project_ids:
        for pr in (await db.scalars(
            select(Project).where(Project.id.in_(project_ids))
            .order_by(Project.created_at.desc())
        )).all():
            projects.append({
                "id": str(pr.id), "code": pr.code, "status": pr.status,
                "target_delivery": pr.target_delivery,
            })

    # Goods-receipt + QC history across this supplier's POs
    goods_receipts, qc_reports, files = [], [], []
    if po_ids:
        for gr in (await db.scalars(
            select(GoodsReceipt).where(GoodsReceipt.po_id.in_(po_ids))
            .order_by(GoodsReceipt.created_at.desc()).limit(50)
        )).all():
            goods_receipts.append({
                "id": str(gr.id), "po_number": po_number.get(gr.po_id),
                "received_at": gr.received_at, "status": gr.status,
                "items": gr.items,
            })
        for r in (await db.scalars(
            select(QCReport).where(QCReport.po_id.in_(po_ids))
            .order_by(QCReport.created_at.desc()).limit(50)
        )).all():
            qc_reports.append({
                "id": str(r.id), "po_number": po_number.get(r.po_id),
                "pass_qty": float(r.pass_qty or 0), "fail_qty": float(r.fail_qty or 0),
                "decision": r.decision, "findings": r.findings,
            })
        for a in (await db.scalars(
            select(Attachment).where(
                Attachment.owner_type == "supplier_po",
                Attachment.owner_id.in_(po_ids),
            ).order_by(Attachment.created_at.desc())
        )).all():
            files.append({
                "id": str(a.id), "filename": a.filename,
                "content_type": a.content_type, "size": a.size,
                "po_number": po_number.get(a.owner_id),
                "uploaded_at": a.created_at,
                "download_url": f"/api/v1/attachments/{a.id}/download",
            })

    contacts = (await db.scalars(
        select(SupplierContact)
        .where(SupplierContact.supplier_id == supplier_id)
        .order_by(SupplierContact.is_primary.desc(),
                  SupplierContact.created_at.asc())
    )).all()

    return {
        "id": str(s.id),
        "name": s.name,
        "category": s.category,
        "rating": float(s.rating or 0),
        "lead_time_days_avg": float(s.lead_time_days_avg or 0),
        "qc_fail_rate": float(s.qc_fail_rate or 0),
        "price_volatility": float(s.price_volatility or 0),
        "company_address": s.company_address,
        "warehouse_address": s.warehouse_address,
        # Fall back to the legacy blob so suppliers created before the columns
        # existed still show the number somebody typed into it.
        "phone": s.phone or (s.contact or {}).get("phone"),
        "whatsapp": s.whatsapp or (s.contact or {}).get("whatsapp"),
        "email": s.email or (s.contact or {}).get("email"),
        "contacts": [_supplier_contact_out(x) for x in contacts],
        "contact": s.contact or {},
        "po_count": len(po_rows),
        "open_po_count": len(open_pos),
        "lifetime_value": float(sum(float(p.total or 0) for p in po_rows)),
        "purchase_orders": [
            {
                "id": str(p.id),
                "number": p.number,
                "status": p.status,
                "po_date": p.po_date,
                "total": float(p.total or 0),
                "project_id": str(p.project_id) if p.project_id else None,
            }
            for p in po_rows
        ],
        "projects": projects,
        "goods_receipts": goods_receipts,
        "qc_reports": qc_reports,
        "files": files,
    }


@router.post("/suppliers", status_code=201)
async def create_supplier(
    payload: SupplierIn,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_supplier_editors),
):
    if not payload.name.strip():
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Name required")
    existing = await db.scalar(select(Supplier).where(Supplier.name == payload.name.strip()))
    if existing:
        raise HTTPException(status.HTTP_409_CONFLICT, "Supplier with this name already exists")
    from app.models.purchasing import SupplierContact

    s = Supplier(
        name=payload.name.strip(),
        category=payload.category,
        rating=payload.rating,
        company_address=payload.company_address,
        warehouse_address=payload.warehouse_address,
        phone=payload.phone,
        whatsapp=payload.whatsapp,
        email=payload.email,
        contact=payload.contact or {},
    )
    db.add(s)
    await db.flush()

    # The PICs typed into the same form. A supplier with nobody to ring is a
    # row you have to come back and finish, so the form carries them.
    for c in payload.contacts:
        if not (c.name or "").strip():
            continue
        db.add(SupplierContact(
            supplier_id=s.id,
            name=c.name.strip(),
            position=c.position,
            phone=c.phone,
            whatsapp=c.whatsapp,
            email=c.email,
            is_primary=c.is_primary,
            notes=c.notes,
        ))
    await db.flush()
    return {"id": str(s.id), "name": s.name}


@router.patch("/suppliers/{supplier_id}")
async def update_supplier(
    supplier_id: UUID,
    payload: SupplierPatch,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_supplier_editors),
):
    """Correct the header: address, company line, category, rating.

    Suppliers move warehouses and change switchboards, and until now the row
    was write-once — the only way to fix a typo in an address was to create a
    second supplier, which splits the PO history in two.
    """
    s = await db.get(Supplier, supplier_id)
    if not s:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Supplier not found")
    data = payload.model_dump(exclude_unset=True)

    if "name" in data:
        new_name = (data.pop("name") or "").strip()
        if not new_name:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Name required")
        if new_name != s.name:
            clash = await db.scalar(
                select(Supplier).where(Supplier.name == new_name,
                                       Supplier.id != s.id)
            )
            if clash:
                raise HTTPException(status.HTTP_409_CONFLICT,
                                    "Supplier with this name already exists")
            s.name = new_name

    for k, v in data.items():
        setattr(s, k, v)
    await db.flush()
    return {"id": str(s.id), "name": s.name}


# ─── Supplier PICs (multiple contacts per supplier) ──────────────────────────
#
# Same shape as the customer's contacts, deliberately: the two directories are
# the same job seen from opposite ends, and a purchasing officer who has used
# one should not have to learn the other.


async def _supplier_or_404(supplier_id: UUID, db: AsyncSession) -> Supplier:
    s = await db.get(Supplier, supplier_id)
    if not s:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Supplier not found")
    return s


@router.get("/suppliers/{supplier_id}/contacts")
async def list_supplier_contacts(
    supplier_id: UUID,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(get_current_user),
):
    from app.models.purchasing import SupplierContact

    await _supplier_or_404(supplier_id, db)
    rows = (await db.scalars(
        select(SupplierContact)
        .where(SupplierContact.supplier_id == supplier_id)
        .order_by(SupplierContact.is_primary.desc(),
                  SupplierContact.created_at.asc())
    )).all()
    return [_supplier_contact_out(c) for c in rows]


@router.post("/suppliers/{supplier_id}/contacts", status_code=201)
async def create_supplier_contact(
    supplier_id: UUID,
    payload: SupplierContactIn,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_supplier_editors),
):
    from app.models.purchasing import SupplierContact

    await _supplier_or_404(supplier_id, db)
    if not payload.name.strip():
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Name required")
    c = SupplierContact(
        supplier_id=supplier_id,
        name=payload.name.strip(),
        position=payload.position,
        phone=payload.phone,
        whatsapp=payload.whatsapp,
        email=payload.email,
        is_primary=payload.is_primary,
        notes=payload.notes,
    )
    db.add(c)
    await db.flush()
    return _supplier_contact_out(c)


@router.patch("/suppliers/{supplier_id}/contacts/{contact_id}")
async def update_supplier_contact(
    supplier_id: UUID,
    contact_id: UUID,
    payload: SupplierContactIn,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_supplier_editors),
):
    from app.models.purchasing import SupplierContact

    await _supplier_or_404(supplier_id, db)
    c = await db.get(SupplierContact, contact_id)
    if not c or c.supplier_id != supplier_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Contact not found")
    for k, v in payload.model_dump(exclude_unset=True).items():
        setattr(c, k, v)
    await db.flush()
    return _supplier_contact_out(c)


@router.delete("/suppliers/{supplier_id}/contacts/{contact_id}", status_code=204)
async def delete_supplier_contact(
    supplier_id: UUID,
    contact_id: UUID,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_supplier_editors),
):
    from app.models.purchasing import SupplierContact

    await _supplier_or_404(supplier_id, db)
    c = await db.get(SupplierContact, contact_id)
    if not c or c.supplier_id != supplier_id:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Contact not found")
    await db.delete(c)
    return None


# ─── Supplier POs ────────────────────────────────────────────────────────────

class PoCreate(BaseModel):
    supplier_id: UUID
    # The job this order is for. On an order covering several jobs — one
    # vendor, one truck, three customers — this is the first of them and the
    # per-line `project_id` on each item is the real answer.
    project_id: UUID
    po_date: str | None = None  # ISO date
    quoted_lead_days: int | None = None
    # When this shipment lands. Per PO, because a job split across three
    # vendors arrives in three deliveries.
    eta: str | None = None      # ISO date
    items: list[dict] = []
    total: float = 0
    # What the figures are denominated in. Defaults to rupiah because most
    # orders are local; an overseas vendor's PO must say USD or CNY, and the
    # printed order carries it on every money column.
    currency: str = "IDR"
    number: str | None = None  # auto-generated if missing
    price_request_id: UUID | None = None  # source the buying price from this PR
    # Build the lines straight off a supplier's answered quote, keeping every
    # line's project and price-request origin.
    supplier_price_request_id: UUID | None = None


# Who may work supplier purchase orders. Admin is out: a supplier PO is the
# procurement side of a job — the vendor, what we paid them, when they ship —
# and admin run the customer side. Every PO screen, export and detail hangs off
# this one dependency, so removing them here closes all of them at once.
_purchasing_or_director = require(Role.PURCHASING, Role.MANAGER, Role.DIRECTOR)


@router.get("/po/prefill")
async def po_prefill(
    project_id: UUID | None = None,
    price_request_id: UUID | None = None,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
):
    """PO line items + total pre-filled from a price request.

    Purchasing already entered the buying (cost) price per line on the price
    request, so a PO shouldn't make them retype it. Pass ``price_request_id``
    to use a specific request, or ``project_id`` to resolve the request the
    project sources from (its direct link, falling back to its quotation's).
    Returns PO-ready lines using the *cost* price as the unit price — never
    the selling price or the customer (purchasing stays blind to that side).
    """
    from app.models.operation import Project
    from app.models.price_request import PriceRequest
    from app.models.quotation import Quotation

    pr = None
    if price_request_id:
        pr = await db.get(PriceRequest, price_request_id)
    elif project_id:
        project = await db.get(Project, project_id)
        if not project:
            raise HTTPException(status.HTTP_404_NOT_FOUND, "Project not found")
        pr_id = project.price_request_id
        # Fall back to the price request behind the project's quotation — covers
        # projects created before the direct link was recorded.
        if not pr_id and project.quotation_id:
            quote = await db.get(Quotation, project.quotation_id)
            pr_id = quote.price_request_id if quote else None
        if pr_id:
            pr = await db.get(PriceRequest, pr_id)
    if not pr:
        return {"price_request_id": None, "items": [], "total": 0}

    items, total, uncosted = [], 0.0, 0
    for it in (pr.items or []):
        qty = float(it.get("qty") or 0)
        unit_cost = float(it.get("cost_price") or 0)
        amount = unit_cost * qty
        total += amount
        if not unit_cost:
            uncosted += 1
        items.append({
            "line_no": it.get("line_no"),
            "description": it.get("description"),
            "qty": qty,
            "uom": it.get("uom"),
            "spec": it.get("spec"),
            "unit_price": unit_cost,   # buying price per unit
            "amount": amount,
            "costed": bool(unit_cost),
        })
    return {
        "price_request_id": str(pr.id),
        "price_request_number": pr.number,
        "items": items,
        "total": total,
        # How many lines have no cost on them. An imported request, or one
        # still being worked, prefills a column of Rp 0 — and the panel used to
        # announce that as "pulled from purchasing's costing", which is how a
        # purchase order for nothing gets raised without anyone noticing.
        "uncosted": uncosted,
    }


@router.get("/po/from-quote/{spr_id}")
async def po_prefill_from_quote(
    spr_id: UUID,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
):
    """PO lines built from a supplier's answered quote.

    The other prefill reads a customer price request and gives you its costs.
    This one reads the quote *that vendor actually gave*, which is the right
    source once the asking has been done — and it carries each line's job with
    it, so an order covering three customers' work arrives with every line
    already pointing at the right project instead of being sorted out by hand.
    """
    from app.models.operation import Project
    from app.models.price_request import PriceRequest
    from app.models.purchasing import SupplierPriceRequest
    from app.models.quotation import Quotation

    spr = await db.get(SupplierPriceRequest, spr_id)
    if not spr:
        raise HTTPException(status.HTTP_404_NOT_FOUND,
                            "Supplier price request not found")

    # Which project each source price request ended up as. A PR reaches a
    # project through its quotation, or directly.
    project_for_pr: dict[str, dict] = {}
    for sid in (spr.source_pr_ids or []):
        pr = await db.get(PriceRequest, UUID(str(sid)))
        if not pr:
            continue
        proj = await db.scalar(
            select(Project).where(Project.price_request_id == pr.id,
                                  Project.is_deleted.is_(False))
        )
        if proj is None and pr.quotation_id:
            proj = await db.scalar(
                select(Project).where(Project.quotation_id == pr.quotation_id,
                                      Project.is_deleted.is_(False))
            )
        if proj is None:
            q = await db.scalar(
                select(Quotation).where(Quotation.price_request_id == pr.id))
            if q:
                proj = await db.scalar(
                    select(Project).where(Project.quotation_id == q.id,
                                          Project.is_deleted.is_(False)))
        project_for_pr[str(sid)] = {
            "price_request_number": pr.number,
            "project_id": str(proj.id) if proj else None,
            "project_code": proj.code if proj else None,
        }

    items, total = [], 0.0
    for it in (spr.items or []):
        qty = float(it.get("qty") or 0)
        unit = float(it.get("quoted_price") or 0)
        amount = qty * unit
        total += amount
        src = project_for_pr.get(str(it.get("source_pr_id")), {})
        items.append({
            "line_no": it.get("line_no"),
            "description": it.get("description"),
            "qty": qty,
            "uom": it.get("uom"),
            "spec": it.get("spec"),
            "unit_price": unit,
            "amount": amount,
            # Where it is going, and where it came from — both travel with the
            # line so the PO can be built without re-deriving either.
            "project_id": src.get("project_id"),
            "project_code": src.get("project_code"),
            "source_pr_id": it.get("source_pr_id"),
            "source_pr_number": it.get("source_pr_number") or src.get("price_request_number"),
            "source_line_no": it.get("source_line_no"),
            "quote_number": spr.number,
        })

    projects = sorted({(i["project_id"], i["project_code"]) for i in items
                       if i["project_id"]})
    return {
        "supplier_price_request_id": str(spr.id),
        "number": spr.number,
        "supplier_id": str(spr.supplier_id),
        "quoted_lead_days": spr.quoted_lead_days,
        "items": items,
        "total": total,
        "projects": [{"id": p, "code": code} for p, code in projects],
        # Lines whose job could not be resolved — usually the deal has not
        # been won yet, so there is no project to point at. Said plainly
        # rather than silently dropped.
        "unassigned_lines": [i["line_no"] for i in items if not i["project_id"]],
    }


@router.get("/po/for-project/{project_id}")
async def pos_for_project(
    project_id: UUID,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(get_current_user),
):
    """Every supplier order feeding one project, as shipments.

    A job that needed three vendors arrives in three deliveries, and the
    project's own `est_arrive_*` fields cannot hold three answers. So each PO
    is a shipment with its own ETA and its own share of the items, numbered in
    the order they are expected — "shipment 1, 2, 3" — and the latest of them
    is the date the whole job is actually complete.
    """
    from datetime import date as date_t

    from app.models.operation import Project
    from app.models.purchasing import GoodsReceipt, Supplier, SupplierPO

    project = await db.get(Project, project_id)
    if not project or project.is_deleted:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Project not found")

    rows = list((await db.scalars(
        select(SupplierPO).where(or_(
            SupplierPO.project_id == project_id,
            SupplierPO.project_ids.contains([str(project_id)]),
        ))
    )).all())
    sups = {s.id: s for s in (await db.scalars(
        select(Supplier).where(
            Supplier.id.in_({r.supplier_id for r in rows}))))} if rows else {}

    def sort_key(p):
        return (p.eta or date_t.max, p.created_at)

    # Admin need the shipments — the dates are their job — but not who is
    # sending them. They are barred from the supplier PO itself, and a card
    # listing "Shipment 2 · PT Rantai" would hand back the vendor anyway. The
    # money goes with it: a per-project total is what we paid the vendor.
    hide_supplier = Role(_u.role) == Role.ADMIN

    rows.sort(key=sort_key)
    shipments, latest = [], None
    for n, po in enumerate(rows, 1):
        mine = [i for i in (po.items or [])
                if not i.get("project_id")
                or str(i.get("project_id")) == str(project_id)]
        received = await db.scalar(
            select(func.count(GoodsReceipt.id)).where(GoodsReceipt.po_id == po.id))
        if po.eta and (latest is None or po.eta > latest):
            latest = po.eta
        sup = sups.get(po.supplier_id)
        shipments.append({
            "shipment_no": n,
            # No id either — an id is a lookup away from the name, and the PO
            # page it points at is closed to them.
            "po_id": None if hide_supplier else str(po.id),
            "number": None if hide_supplier else po.number,
            "status": po.status,
            "supplier_id": None if hide_supplier else str(po.supplier_id),
            "supplier_name": None if hide_supplier else (sup.name if sup else None),
            "eta": po.eta,
            "quoted_lead_days": po.quoted_lead_days,
            "po_date": po.po_date, "currency": po.currency or "IDR",
            "is_shared": len(po.project_ids or []) > 1,
            "other_projects": [c for c in (po.project_ids or [])
                               if str(c) != str(project_id)],
            "received": bool(received),
            "items": [{
                "description": i.get("description"),
                "qty": i.get("qty"),
                "uom": i.get("uom"),
                "supplier_name": None if hide_supplier else (sup.name if sup else None),
            } for i in mine],
            "total_for_project": None if hide_supplier else sum(
                float(i.get("amount") or 0) or
                float(i.get("qty") or 0) * float(i.get("unit_price") or 0)
                for i in mine),
        })

    return {
        "project_id": str(project.id),
        "project_code": project.code,
        "shipments": shipments,
        # Counted off the orders, not off the scrubbed rows — "3 deliveries
        # from 2 suppliers" is a useful thing for admin to know, and it names
        # nobody. Doing it from `shipments` would collapse every hidden id to
        # one None and report a single supplier.
        "supplier_count": len({r.supplier_id for r in rows}),
        # The one date that answers "when is this job actually here": the last
        # shipment to land. Offered rather than written onto the project, so
        # the director's promised date stays theirs.
        "last_eta": latest,
        "all_received": bool(shipments) and all(s["received"] for s in shipments),
    }


@router.get("/po/price-request-options")
async def po_price_request_options(
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
):
    """Approved price requests purchasing can link a PO to manually, when a
    project isn't auto-linked. Cost-only — no customer, no selling price."""
    from app.models.price_request import PriceRequest

    rows = (await db.scalars(
        select(PriceRequest)
        .where(PriceRequest.status == "approved", PriceRequest.is_deleted.is_(False))
        .order_by(PriceRequest.created_at.desc())
    )).all()
    out = []
    for pr in rows:
        total = sum(
            float(it.get("cost_price") or 0) * float(it.get("qty") or 0)
            for it in (pr.items or [])
        )
        out.append({
            "id": str(pr.id),
            "number": pr.number,
            "line_count": len(pr.items or []),
            "total_cost": total,
        })
    return out


@router.get("/po")
async def list_pos(
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
    supplier_id: UUID | None = None,
    project_id: UUID | None = None,
    customer_id: UUID | None = None,
):
    """All supplier POs.

    Rows are enriched with supplier / project / customer / sales-rep
    context so the list view, the PO recap and the per-customer PO
    section can render in one round-trip. Use ?customer_id=… to scope
    to a single customer's POs (joined via project.customer_id).
    """
    from app.models.crm import Customer
    from app.models.operation import Project
    from app.models.purchasing import Supplier, SupplierPO

    stmt = select(SupplierPO).order_by(SupplierPO.created_at.desc())
    if supplier_id:
        stmt = stmt.where(SupplierPO.supplier_id == supplier_id)
    if project_id:
        stmt = stmt.where(SupplierPO.project_id == project_id)
    if customer_id:
        # Join via project to filter by customer.
        stmt = (
            stmt.join(Project, Project.id == SupplierPO.project_id)
            .where(Project.customer_id == customer_id)
        )
    rows = (await db.scalars(stmt)).all()

    # Batch-load related rows so we don't N+1 supplier / project / customer
    # / sales-rep lookups inside the list comprehension.
    supplier_ids = {r.supplier_id for r in rows if r.supplier_id}
    project_ids  = {r.project_id  for r in rows if r.project_id}
    suppliers: dict[UUID, Supplier] = {}
    projects:  dict[UUID, Project]  = {}
    if supplier_ids:
        for s in (await db.scalars(
            select(Supplier).where(Supplier.id.in_(supplier_ids))
        )).all():
            suppliers[s.id] = s
    if project_ids:
        for p in (await db.scalars(
            select(Project).where(Project.id.in_(project_ids))
        )).all():
            projects[p.id] = p

    customer_ids = {p.customer_id for p in projects.values() if p.customer_id}
    customers: dict[UUID, Customer] = {}
    if customer_ids:
        for c in (await db.scalars(
            select(Customer).where(Customer.id.in_(customer_ids))
        )).all():
            customers[c.id] = c

    sales_ids = {c.sales_pic_id for c in customers.values() if c.sales_pic_id}
    sales_users: dict[UUID, User] = {}
    if sales_ids:
        for u in (await db.scalars(
            select(User).where(User.id.in_(sales_ids))
        )).all():
            sales_users[u.id] = u

    # Purchasing works suppliers + parts; the customer behind a project is
    # deliberately hidden from them (they see the project code as the order
    # reference instead).
    hide_customer = Role(_u.role) == Role.PURCHASING
    out = []
    for r in rows:
        proj = projects.get(r.project_id) if r.project_id else None
        cust = customers.get(proj.customer_id) if proj else None
        sales = sales_users.get(cust.sales_pic_id) if cust and cust.sales_pic_id else None
        sup = suppliers.get(r.supplier_id)
        out.append({
            "id": str(r.id), "number": r.number, "status": r.status,
            "supplier_id": str(r.supplier_id),
            "supplier_name": sup.name if sup else None,
            "project_id": str(r.project_id) if r.project_id else None,
            "project_code": proj.code if proj else None,
            "customer_id": None if hide_customer else (str(cust.id) if cust else None),
            "customer_name": (
                (proj.code if proj else "—") if hide_customer
                else (cust.company_name if cust else None)
            ),
            "sales_pic_id": None if hide_customer else (str(sales.id) if sales else None),
            "sales_pic_name": None if hide_customer else (sales.full_name if sales else None),
            "po_date": r.po_date, "eta": r.eta,
            "currency": r.currency or "IDR", "total": float(r.total or 0),
            "quoted_lead_days": r.quoted_lead_days,
            "items": r.items, "created_at": r.created_at,
        })
    return out


@router.post("/po", status_code=201)
async def create_po(
    payload: PoCreate,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_purchasing_or_director),
):
    """Create a supplier PO. Every step in the PO lifecycle is gated on
    director approval: when a non-director submits this, the PO is
    created with status='pending_approval' and an ApprovalRequest is
    filed for the director. The director sees it in /approvals and can
    flip it to 'open' (or 'cancelled') from there. Director themselves
    short-circuit and the PO comes up 'open' immediately.
    """
    from datetime import date as date_t

    from app.models.operation import Project
    from app.models.purchasing import Supplier, SupplierPO

    supplier = await db.get(Supplier, payload.supplier_id)
    if not supplier:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Unknown supplier")
    project = await db.get(Project, payload.project_id)
    if not project:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Unknown project")

    number = payload.number
    if not number:
        prefix = f"PO-{date_t.today().strftime('%y%m%d')}-"
        # One past the highest issued today, not a count of today's rows: a
        # count walks backwards the moment one is deleted and hands the next
        # PO a number that is still in use.
        from app.services.numbering import _next_suffix
        number = f"{prefix}{await _next_suffix(db, SupplierPO.number, prefix):03d}"

    po_date_parsed = None
    if payload.po_date:
        try:
            po_date_parsed = date_t.fromisoformat(payload.po_date)
        except ValueError:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "po_date must be YYYY-MM-DD")

    # Link the PO to the price request it sources against: an explicit choice
    # wins, else the project's direct link, else the project's quotation's link.
    price_request_id = payload.price_request_id or project.price_request_id
    if not price_request_id and project.quotation_id:
        from app.models.quotation import Quotation
        quote = await db.get(Quotation, project.quotation_id)
        price_request_id = quote.price_request_id if quote else None

    eta_parsed = None
    if payload.eta:
        try:
            eta_parsed = date_t.fromisoformat(payload.eta)
        except ValueError:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "eta must be YYYY-MM-DD")

    # Lines carry the job they belong to. A line that does not name one is
    # for the PO's own project, which is every ordinary single-job order.
    items = [dict(it) for it in (payload.items or [])]
    project_cache: dict[str, Project] = {str(project.id): project}
    for it in items:
        pid = str(it.get("project_id") or payload.project_id)
        if pid not in project_cache:
            other = await db.get(Project, UUID(pid))
            if not other:
                raise HTTPException(status.HTTP_400_BAD_REQUEST,
                                    f"Line names project {pid}, which does not exist")
            project_cache[pid] = other
        it["project_id"] = pid
        it["project_code"] = project_cache[pid].code
    project_ids = sorted({it["project_id"] for it in items}) or [str(project.id)]

    is_director = Role(user.role) == Role.DIRECTOR
    po = SupplierPO(
        number=number,
        supplier_id=payload.supplier_id,
        project_id=payload.project_id,
        project_ids=project_ids,
        price_request_id=price_request_id,
        po_date=po_date_parsed,
        eta=eta_parsed,
        quoted_lead_days=payload.quoted_lead_days,
        currency=(payload.currency or "IDR").strip().upper()[:8] or "IDR",
        total=payload.total,
        items=items,
        status="open" if is_director else "pending_approval",
    )
    db.add(po)
    await db.flush()

    if not is_director:
        await request_approval(
            db,
            target_type="supplier_po",
            target_id=po.id,
            requested_by=user.id,
            required_role=Role.DIRECTOR,
            reason=f"Create PO {po.number} ({supplier.name}, project {project.code})",
            payload={"action": "create"},
        )

    # The supplier PO is now the trigger that moves the project to the
    # purchasing stage (we dropped the separate purchase-request gate as
    # redundant: the cost decision already happened on the price request,
    # the director approves the PO itself, and there's no RFQ comparison
    # in this workflow). Only advance forward — never regress a project
    # already past purchasing.
    from app.models.operation import advance_project_status
    advance_project_status(project, "purchasing")
    await db.flush()

    return {
        "id": str(po.id), "number": po.number,
        "supplier_id": str(po.supplier_id),
        "project_id": str(po.project_id),
        "status": po.status,
        # Echoed back because it is normalised on the way in ("usd" → "USD"),
        # and a caller that cannot see what was stored cannot tell.
        "currency": po.currency,
        "eta": po.eta,
        "total": float(po.total or 0),
        "pending_approval": not is_director,
    }


@router.get("/po/{po_id}")
async def get_po(
    po_id: UUID,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
):
    """Full PO detail with supplier and project context for the detail page."""
    from app.models.crm import Customer
    from app.models.operation import Project
    from app.models.purchasing import Supplier, SupplierPO

    po = await db.get(SupplierPO, po_id)
    if not po:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "PO not found")
    supplier = await db.get(Supplier, po.supplier_id) if po.supplier_id else None
    project = await db.get(Project, po.project_id) if po.project_id else None
    price_request = None
    if po.price_request_id:
        from app.models.price_request import PriceRequest
        pr = await db.get(PriceRequest, po.price_request_id)
        if pr:
            price_request = {"id": str(pr.id), "number": pr.number}
    # Sales rep in charge (via the project's customer) for the detail page —
    # hidden from purchasing, who must stay blind to the customer side.
    sales_rep = None
    if project and project.customer_id and Role(_u.role) != Role.PURCHASING:
        cust = await db.get(Customer, project.customer_id)
        if cust and cust.sales_pic_id:
            rep = await db.get(User, cust.sales_pic_id)
            if rep:
                sales_rep = {"id": str(rep.id), "name": rep.full_name}
    return {
        "id": str(po.id),
        "number": po.number,
        "status": po.status,
        "supplier_id": str(po.supplier_id),
        "supplier_name": supplier.name if supplier else None,
        "supplier_category": supplier.category if supplier else None,
        "sales_pic_id": sales_rep["id"] if sales_rep else None,
        "sales_pic_name": sales_rep["name"] if sales_rep else None,
        "price_request_id": price_request["id"] if price_request else None,
        "price_request_number": price_request["number"] if price_request else None,
        "project_id": str(po.project_id) if po.project_id else None,
        "project_code": project.code if project else None,
        "project_status": project.status if project else None,
        "project_target_delivery": project.target_delivery if project else None,
        "project_actual_delivery": project.actual_delivery if project else None,
        "po_date": po.po_date,
        "eta": po.eta,
        "currency": po.currency or "IDR",
        "quoted_lead_days": po.quoted_lead_days,
        "total": float(po.total or 0),
        "items": po.items,
        "project_ids": [str(x) for x in (po.project_ids or [])],
        "created_at": po.created_at,
    }


# ─── Printing a supplier PO ──────────────────────────────────────────────────
#
# A purchase order that only exists on a screen is one somebody has to retype
# into an email. Two formats because they are used differently: the PDF is
# what gets sent to the vendor, the spreadsheet is what gets pasted into a
# stock sheet or a payment schedule.


async def _po_print_bundle(po_id: UUID, db: AsyncSession) -> dict:
    """Everything the two exporters need, gathered once."""
    from app.models.operation import Project
    from app.models.purchasing import Supplier, SupplierContact, SupplierPO

    po = await db.get(SupplierPO, po_id)
    if not po:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "PO not found")
    supplier = await db.get(Supplier, po.supplier_id) if po.supplier_id else None
    project = await db.get(Project, po.project_id) if po.project_id else None
    pic = None
    if supplier:
        pic = await db.scalar(
            select(SupplierContact)
            .where(SupplierContact.supplier_id == supplier.id)
            .order_by(SupplierContact.is_primary.desc(),
                      SupplierContact.created_at.asc())
            .limit(1)
        )
    return {"po": po, "supplier": supplier, "project": project, "pic": pic}


@router.get("/po/{po_id}/export.pdf")
async def export_po_pdf(
    po_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_purchasing_or_director),
):
    from fastapi.responses import Response

    from app.services.signature import load_for
    from app.services.supplier_po_pdf import build_supplier_po_pdf

    b = await _po_print_bundle(po_id, db)
    po, supplier, project, pic = b["po"], b["supplier"], b["project"], b["pic"]

    # Where the goods come from: the warehouse if they gave one, because that
    # is the gate the truck actually pulls up to, otherwise the office.
    supplier_address = ""
    if supplier:
        supplier_address = (supplier.warehouse_address
                            or supplier.company_address or "")

    rows = []
    for it in (po.items or []):
        qty = float(it.get("qty") or 0)
        unit = float(it.get("unit_price") or 0)
        rows.append({
            "description": it.get("description"),
            "qty": qty, "uom": it.get("uom"),
            "unit_price": unit,
            "amount": float(it.get("amount") or qty * unit),
        })

    pdf = build_supplier_po_pdf(
        number=po.number,
        po_date=po.po_date.strftime("%d %b %Y") if po.po_date else "—",
        supplier_name=supplier.name if supplier else "—",
        supplier_address=supplier_address,
        supplier_pic=(pic.name if pic else ((supplier.contact or {}).get("name")
                                            if supplier else "")) or "",
        supplier_phone=(pic.phone if pic and pic.phone
                        else (supplier.phone if supplier else "")) or "",
        supplier_email=(pic.email if pic and pic.email
                        else (supplier.email if supplier else "")) or "",
        # Not a literal any more. This is the address a vendor ships to, so a
        # value invented in code is a container sent to the wrong gate; it
        # comes from COMPANY_WAREHOUSE_ADDRESS, and says it is unset rather
        # than guessing when nobody has configured one.
        ship_to_label=settings.COMPANY_WAREHOUSE_LABEL,
        ship_to_address=(settings.COMPANY_WAREHOUSE_ADDRESS.strip()
                         or "— delivery address not set, please confirm with us —"),
        project_code=project.code if project else None,
        lead_days=po.quoted_lead_days,
        rows=rows,
        currency=po.currency or "IDR",
        total=float(po.total or 0),
        # A supplier PO has no keterangan field of its own yet; when it
        # gains one this is where it prints.
        notes=None,
        issued_by=user.full_name or "—",
        issuer_signature=await load_for(user),
    )
    await audit_record(db, actor=user, action="export", entity="supplier_po",
                       entity_id=po.id, after={"format": "pdf"})
    return Response(
        content=pdf, media_type="application/pdf",
        headers={"Content-Disposition":
                 f'attachment; filename="PO-{po.number}.pdf"'},
    )


@router.get("/po/{po_id}/export.xlsx")
async def export_po_excel(
    po_id: UUID,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_purchasing_or_director),
):
    from io import BytesIO

    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    b = await _po_print_bundle(po_id, db)
    po, supplier, project, pic = b["po"], b["supplier"], b["project"], b["pic"]

    wb = Workbook()
    ws = wb.active
    ws.title = f"PO {po.number}"[:31]
    bold = Font(bold=True)
    thin = Side(border_style="thin", color="CBD1DC")
    boxed = Border(left=thin, right=thin, top=thin, bottom=thin)
    right = Alignment(horizontal="right")
    head_fill = PatternFill("solid", fgColor="EEF0F4")

    ws["A1"] = "PT. Transmisi Enjinering"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:E1")
    ws["A2"] = f"Purchase Order {po.number}"
    ws["A2"].font = bold
    ws.merge_cells("A2:E2")

    row = 4
    for label, value in (
        ("Supplier", supplier.name if supplier else "—"),
        ("Address", (supplier.warehouse_address or supplier.company_address or "—")
         if supplier else "—"),
        ("PIC", (pic.name if pic else "—")),
        ("Phone", (pic.phone if pic and pic.phone
                   else (supplier.phone if supplier else "")) or "—"),
        ("PO date", str(po.po_date) if po.po_date else "—"),
        ("Project", project.code if project else "—"),
        ("Lead time (days)", po.quoted_lead_days if po.quoted_lead_days is not None else "—"),
        ("Currency", po.currency or "IDR"),
        ("Status", po.status),
    ):
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    cur = po.currency or "IDR"
    # Rupiah has no minor unit in practice; a dollar or yuan price without its
    # cents is a rounded number pretending to be exact.
    money_fmt = "#,##0" if cur.upper() == "IDR" else "#,##0.00"
    for col, header in enumerate(
            ["#", "Description", "Qty", "UoM", f"Unit price ({cur})",
             f"Amount ({cur})"], start=1):
        c = ws.cell(row=row, column=col, value=header)
        c.font = bold
        c.fill = head_fill
        c.border = boxed
    row += 1
    for i, it in enumerate((po.items or []), 1):
        qty = float(it.get("qty") or 0)
        unit = float(it.get("unit_price") or 0)
        values = [i, it.get("description") or "—", qty, it.get("uom") or "",
                  unit, float(it.get("amount") or qty * unit)]
        for col, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=col, value=v)
            c.border = boxed
            if col in (3, 5, 6):
                c.alignment = right
                if col in (5, 6):
                    c.number_format = money_fmt
        row += 1

    ws.cell(row=row, column=5, value=f"TOTAL {cur}").font = bold
    tot = ws.cell(row=row, column=6, value=float(po.total or 0))
    tot.font = bold
    tot.number_format = money_fmt
    tot.alignment = right

    for col, width in zip("ABCDEF", (6, 46, 10, 10, 16, 18)):
        ws.column_dimensions[col].width = width

    buf = BytesIO()
    wb.save(buf)
    await audit_record(db, actor=user, action="export", entity="supplier_po",
                       entity_id=po.id, after={"format": "xlsx"})
    return Response(
        content=buf.getvalue(),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition":
                 f'attachment; filename="PO-{po.number}.xlsx"'},
    )


class POPatch(BaseModel):
    number: str | None = None
    po_date: str | None = None        # ISO YYYY-MM-DD
    eta: str | None = None            # ISO YYYY-MM-DD — when the vendor says it lands
    quoted_lead_days: int | None = None
    currency: str | None = None       # IDR | USD | CNY | …
    total: float | None = None
    status: str | None = None         # open | received | closed | cancelled
    items: list | None = None


@router.patch("/po/{po_id}")
async def update_po(
    po_id: UUID,
    payload: POPatch,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_purchasing_or_director),
):
    """Update an existing supplier PO. Every change is gated on director
    approval: when a non-director submits this, the PO is left untouched
    and an ApprovalRequest is filed with the proposed changes in its
    payload. The director sees it in /approvals and the changes are
    applied when they approve. Director themselves apply immediately.

    Renaming the PO number is allowed but the new value must be unique —
    a conflict returns 409 so the UI can show "that number's already
    used" instead of letting the DB raise an opaque IntegrityError.
    """
    from datetime import date as date_t

    from app.models.purchasing import SupplierPO

    po = await db.get(SupplierPO, po_id)
    if not po:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "PO not found")

    data = payload.model_dump(exclude_unset=True)

    # Validate without mutating — same checks regardless of approval path,
    # so we never queue a doomed approval the director can't apply later.
    if "number" in data:
        new_num = (data["number"] or "").strip()
        if not new_num:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "PO number cannot be empty")
        if new_num != po.number:
            clash = await db.scalar(
                select(SupplierPO).where(
                    SupplierPO.number == new_num, SupplierPO.id != po_id
                )
            )
            if clash:
                raise HTTPException(
                    status.HTTP_409_CONFLICT,
                    f"PO number '{new_num}' is already used by another PO",
                )
        data["number"] = new_num
    if "po_date" in data and data["po_date"] not in (None, ""):
        try:
            date_t.fromisoformat(data["po_date"])
        except ValueError:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "po_date must be YYYY-MM-DD")
    if "eta" in data and data["eta"] not in (None, ""):
        try:
            date_t.fromisoformat(data["eta"])
        except ValueError:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "eta must be YYYY-MM-DD")

    # The vendor's promised arrival applies straight away, unlike everything
    # else on this PO. It is not a money decision — it moves whenever the
    # supplier calls to say the truck slipped a week — and the shipment list on
    # the project page is only worth reading if purchasing can keep it current.
    # Queuing it behind the director would mean the dates staff actually plan
    # around live in WhatsApp again. It is audited instead, and it changes no
    # customer-facing date: the project's target/actual delivery stay gated in
    # operation.py, where the customer sees them.
    if "eta" in data:
        raw = data.pop("eta")
        before = po.eta
        po.eta = None if raw in (None, "") else date_t.fromisoformat(raw)
        await audit_record(
            db, actor=user, action="update", entity="supplier_po", entity_id=po.id,
            before={"eta": before.isoformat() if before else None},
            after={"eta": po.eta.isoformat() if po.eta else None},
        )

    is_director = Role(user.role) == Role.DIRECTOR
    # `data` can be empty now if the ETA was the only change — filing an
    # approval for "no fields" would put a meaningless row in the director's
    # queue, so fall through to the apply path and return the updated PO.
    if not is_director and data:
        # Drop the field outright if it would clear an existing date — the
        # rule everywhere else in the app is "null doesn't wipe protected
        # dates". The director can still set a date explicitly.
        if data.get("po_date") in (None, ""):
            data.pop("po_date", None)
        await request_approval(
            db,
            target_type="supplier_po",
            target_id=po.id,
            requested_by=user.id,
            required_role=Role.DIRECTOR,
            reason=f"Update PO {po.number}: {', '.join(sorted(data.keys())) or 'no fields'}",
            payload={"action": "update", "changes": data},
        )
        # IMPORTANT: return (don't raise) — raising rolls back the session via
        # get_db's exception handler, which would discard the approval we just
        # filed. The PO is intentionally left unchanged until the director signs
        # off; the changes ride in the approval payload.
        return {
            "id": str(po.id), "number": po.number, "status": po.status,
            "supplier_id": str(po.supplier_id),
            "project_id": str(po.project_id) if po.project_id else None,
            "po_date": po.po_date, "eta": po.eta,
            "currency": po.currency or "IDR", "total": float(po.total or 0),
            "quoted_lead_days": po.quoted_lead_days, "items": po.items,
            "pending_approval": True,
            "detail": "Submitted for director approval; changes will apply once approved.",
        }

    # Director path: apply immediately.
    if "number" in data:
        po.number = data["number"]
    if "po_date" in data:
        raw = data["po_date"]
        po.po_date = None if raw in (None, "") else date_t.fromisoformat(raw)
    if "quoted_lead_days" in data:
        po.quoted_lead_days = data["quoted_lead_days"]
    if "currency" in data and data["currency"]:
        po.currency = data["currency"].strip().upper()[:8]
    if "total" in data and data["total"] is not None:
        po.total = data["total"]
    if "status" in data and data["status"]:
        po.status = data["status"]
    if "items" in data and data["items"] is not None:
        po.items = data["items"]

    await db.flush()
    return {
        "id": str(po.id), "number": po.number, "status": po.status,
        "supplier_id": str(po.supplier_id),
        "project_id": str(po.project_id) if po.project_id else None,
        "po_date": po.po_date, "eta": po.eta,
            "currency": po.currency or "IDR", "total": float(po.total or 0),
        "quoted_lead_days": po.quoted_lead_days,
        "items": po.items,
    }


# ─── Purchase Requests (PR) ──────────────────────────────────────────────────
# A PR signals demand against a project. It commits no money, so any of the
# procurement-board roles can raise and close one — no director approval gate.

def _seq_number(db_count: int, prefix: str) -> str:
    from datetime import date as date_t
    ts = date_t.today().strftime("%y%m%d")
    return f"{prefix}-{ts}-{db_count + 1:03d}"


class PRIn(BaseModel):
    project_id: UUID | None = None
    items: list[dict] = []
    notes: str | None = None


class PRPatch(BaseModel):
    status: str | None = None          # open | closed
    items: list[dict] | None = None
    notes: str | None = None


@router.get("/pr")
async def list_pr(
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
    project_id: UUID | None = None,
    status_filter: str | None = None,
):
    from app.models.operation import Project
    from app.models.purchasing import PurchaseRequest

    stmt = select(PurchaseRequest).order_by(PurchaseRequest.created_at.desc())
    if project_id:
        stmt = stmt.where(PurchaseRequest.project_id == project_id)
    if status_filter:
        stmt = stmt.where(PurchaseRequest.status == status_filter)
    rows = (await db.scalars(stmt)).all()

    project_ids = {r.project_id for r in rows if r.project_id}
    requester_ids = {r.requested_by for r in rows if r.requested_by}
    projects: dict[UUID, Project] = {}
    requesters: dict[UUID, User] = {}
    if project_ids:
        for p in (await db.scalars(select(Project).where(Project.id.in_(project_ids)))).all():
            projects[p.id] = p
    if requester_ids:
        for u in (await db.scalars(select(User).where(User.id.in_(requester_ids)))).all():
            requesters[u.id] = u

    return [
        {
            "id": str(r.id), "number": r.number, "status": r.status,
            "project_id": str(r.project_id) if r.project_id else None,
            "project_code": projects[r.project_id].code if r.project_id in projects else None,
            "requested_by": str(r.requested_by) if r.requested_by else None,
            "requested_by_name": (
                requesters[r.requested_by].full_name if r.requested_by in requesters else None
            ),
            "items": r.items, "notes": r.notes, "created_at": r.created_at,
        }
        for r in rows
    ]


@router.post("/pr", status_code=201)
async def create_pr(
    payload: PRIn,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(_purchasing_or_director),
):
    from datetime import date as date_t

    from app.models.operation import Project
    from app.models.purchasing import PurchaseRequest

    if payload.project_id:
        if not await db.get(Project, payload.project_id):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "Unknown project")
    count = await db.scalar(
        select(func.count(PurchaseRequest.id)).where(
            PurchaseRequest.number.like(f"PR-{date_t.today():%y%m%d}-%")
        )
    ) or 0
    pr = PurchaseRequest(
        number=_seq_number(count, "PR"),
        project_id=payload.project_id,
        requested_by=user.id,
        items=payload.items or [],
        notes=payload.notes,
        status="open",
    )
    db.add(pr)
    await db.flush()
    # Every PR is director-gated: a non-director's request parks at
    # pending_approval until the director opens it from /approvals.
    pending = await require_pr_approval(db, pr=pr, requester=user)
    return {"id": str(pr.id), "number": pr.number, "status": pr.status,
            "pending_approval": pending}


@router.patch("/pr/{pr_id}")
async def update_pr(
    pr_id: UUID,
    payload: PRPatch,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
):
    from app.models.purchasing import PurchaseRequest

    pr = await db.get(PurchaseRequest, pr_id)
    if not pr:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "PR not found")
    data = payload.model_dump(exclude_unset=True)
    if "status" in data and data["status"]:
        if data["status"] not in ("open", "closed"):
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "status must be open or closed")
        # A PR awaiting the director's approval can't be opened from here —
        # that would bypass the gate require_pr_approval put it behind.
        if pr.status == "pending_approval" and Role(_u.role) != Role.DIRECTOR:
            raise HTTPException(
                status.HTTP_409_CONFLICT,
                "This purchase request is awaiting director approval — it "
                "can't be opened or closed until that decision is made.",
            )
        pr.status = data["status"]
    if "items" in data and data["items"] is not None:
        pr.items = data["items"]
    if "notes" in data:
        pr.notes = data["notes"]
    await db.flush()
    return {"id": str(pr.id), "number": pr.number, "status": pr.status}


# ─── RFQ (request for quotation) ─────────────────────────────────────────────
# One RFQ row = one supplier's quote against a PR. Listing all RFQs for a PR
# gives the price/lead-time comparison the director uses to award a PO.

class RFQIn(BaseModel):
    supplier_id: UUID
    quoted_lines: list[dict] = []
    quoted_lead_days: int | None = None


@router.get("/rfq")
async def list_rfq(
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
    pr_id: UUID | None = None,
):
    from app.models.purchasing import RFQ, PurchaseRequest, Supplier

    stmt = select(RFQ).order_by(RFQ.created_at.desc())
    if pr_id:
        stmt = stmt.where(RFQ.pr_id == pr_id)
    rows = (await db.scalars(stmt)).all()
    sup_ids = {r.supplier_id for r in rows}
    pr_ids = {r.pr_id for r in rows}
    sups: dict[UUID, Supplier] = {}
    prs: dict[UUID, PurchaseRequest] = {}
    if sup_ids:
        for s in (await db.scalars(select(Supplier).where(Supplier.id.in_(sup_ids)))).all():
            sups[s.id] = s
    if pr_ids:
        pr_q = select(PurchaseRequest).where(PurchaseRequest.id.in_(pr_ids))
        for p in (await db.scalars(pr_q)).all():
            prs[p.id] = p
    return [
        {
            "id": str(r.id), "pr_id": str(r.pr_id),
            "pr_number": prs[r.pr_id].number if r.pr_id in prs else None,
            "supplier_id": str(r.supplier_id),
            "supplier_name": sups[r.supplier_id].name if r.supplier_id in sups else None,
            "quoted_lines": r.quoted_lines, "quoted_lead_days": r.quoted_lead_days,
            "quoted_total": float(
                sum(float(line.get("amount", 0) or 0) for line in (r.quoted_lines or []))
            ),
            "status": r.status, "created_at": r.created_at,
        }
        for r in rows
    ]


@router.post("/pr/{pr_id}/rfq", status_code=201)
async def spawn_rfq(
    pr_id: UUID,
    payload: RFQIn,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
):
    from app.models.purchasing import RFQ, PurchaseRequest, Supplier

    if not await db.get(PurchaseRequest, pr_id):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Unknown PR")
    if not await db.get(Supplier, payload.supplier_id):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Unknown supplier")
    rfq = RFQ(
        pr_id=pr_id,
        supplier_id=payload.supplier_id,
        quoted_lines=payload.quoted_lines or [],
        quoted_lead_days=payload.quoted_lead_days,
        status="open",
    )
    db.add(rfq)
    await db.flush()
    return {"id": str(rfq.id), "pr_id": str(pr_id), "supplier_id": str(payload.supplier_id)}


# ─── Goods Receipt (GR) ──────────────────────────────────────────────────────

class GRIn(BaseModel):
    received_at: str | None = None     # ISO date
    items: list[dict] = []
    status: str = "received"


@router.get("/gr")
async def list_gr(
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
    po_id: UUID | None = None,
):
    from app.models.purchasing import GoodsReceipt, SupplierPO

    stmt = select(GoodsReceipt).order_by(GoodsReceipt.created_at.desc())
    if po_id:
        stmt = stmt.where(GoodsReceipt.po_id == po_id)
    rows = (await db.scalars(stmt)).all()
    po_ids = {r.po_id for r in rows}
    pos: dict[UUID, SupplierPO] = {}
    if po_ids:
        for p in (await db.scalars(select(SupplierPO).where(SupplierPO.id.in_(po_ids)))).all():
            pos[p.id] = p
    return [
        {
            "id": str(r.id), "po_id": str(r.po_id),
            "po_number": pos[r.po_id].number if r.po_id in pos else None,
            "received_at": r.received_at, "items": r.items, "status": r.status,
            "created_at": r.created_at,
        }
        for r in rows
    ]


@router.post("/po/{po_id}/gr", status_code=201)
async def goods_receipt(
    po_id: UUID,
    payload: GRIn,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
):
    from datetime import date as date_t

    from app.models.purchasing import GoodsReceipt, SupplierPO

    po = await db.get(SupplierPO, po_id)
    if not po:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Unknown PO")
    received = None
    if payload.received_at:
        try:
            received = date_t.fromisoformat(payload.received_at)
        except ValueError:
            raise HTTPException(status.HTTP_400_BAD_REQUEST, "received_at must be YYYY-MM-DD")
    gr = GoodsReceipt(
        po_id=po_id,
        received_at=received or date_t.today(),
        items=payload.items or [],
        status=payload.status or "received",
    )
    db.add(gr)
    # Receiving goods moves the PO to 'received' so the board reflects progress.
    if po.status in ("open", "pending_approval"):
        po.status = "received"
    await db.flush()
    return {"id": str(gr.id), "po_id": str(po_id), "status": gr.status}


# ─── QC (incoming inspection) ────────────────────────────────────────────────

class QCIn(BaseModel):
    pass_qty: float = 0
    fail_qty: float = 0
    findings: str | None = None
    decision: str = "accepted"         # accepted | rejected | conditional


@router.get("/qc")
async def list_qc(
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
    po_id: UUID | None = None,
):
    from app.models.purchasing import QCReport, SupplierPO

    stmt = select(QCReport).order_by(QCReport.created_at.desc())
    if po_id:
        stmt = stmt.where(QCReport.po_id == po_id)
    rows = (await db.scalars(stmt)).all()
    po_ids = {r.po_id for r in rows}
    pos: dict[UUID, SupplierPO] = {}
    if po_ids:
        for p in (await db.scalars(select(SupplierPO).where(SupplierPO.id.in_(po_ids)))).all():
            pos[p.id] = p
    return [
        {
            "id": str(r.id), "po_id": str(r.po_id),
            "po_number": pos[r.po_id].number if r.po_id in pos else None,
            "supplier_id": str(pos[r.po_id].supplier_id) if r.po_id in pos else None,
            "pass_qty": float(r.pass_qty or 0), "fail_qty": float(r.fail_qty or 0),
            "findings": r.findings, "decision": r.decision, "created_at": r.created_at,
        }
        for r in rows
    ]


@router.post("/po/{po_id}/qc", status_code=201)
async def qc(
    po_id: UUID,
    payload: QCIn,
    db: AsyncSession = Depends(get_db),
    _u: User = Depends(_purchasing_or_director),
):
    """Record an incoming-QC result for a PO and refresh the supplier's
    rolling QC fail rate (total failed / total inspected across all their
    POs) so the supplier directory reflects real quality history."""
    from app.models.purchasing import QCReport, Supplier, SupplierPO

    po = await db.get(SupplierPO, po_id)
    if not po:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Unknown PO")
    if payload.decision not in ("accepted", "rejected", "conditional"):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Invalid decision")
    report = QCReport(
        po_id=po_id,
        pass_qty=payload.pass_qty,
        fail_qty=payload.fail_qty,
        findings=payload.findings,
        decision=payload.decision,
    )
    db.add(report)
    await db.flush()

    # Recompute the supplier's QC fail rate across every QC report for their POs.
    supplier_po_ids = (await db.scalars(
        select(SupplierPO.id).where(SupplierPO.supplier_id == po.supplier_id)
    )).all()
    if supplier_po_ids:
        agg = (await db.execute(
            select(
                func.coalesce(func.sum(QCReport.pass_qty), 0),
                func.coalesce(func.sum(QCReport.fail_qty), 0),
            ).where(QCReport.po_id.in_(supplier_po_ids))
        )).one()
        total_pass, total_fail = float(agg[0]), float(agg[1])
        inspected = total_pass + total_fail
        if inspected > 0:
            supplier = await db.get(Supplier, po.supplier_id)
            if supplier:
                supplier.qc_fail_rate = round(total_fail / inspected, 4)
    await db.flush()
    return {
        "id": str(report.id), "po_id": str(po_id), "decision": report.decision,
        "pass_qty": float(report.pass_qty), "fail_qty": float(report.fail_qty),
    }

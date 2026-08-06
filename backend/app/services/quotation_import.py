"""Read Accurate's "Rincian Penawaran Penjualan" export — one sheet per quote.

This export is shaped unlike the others: 137 worksheets, each one a single
quotation, each with the same seven columns. The number, date and customer
appear only on the first line; the rest of the lines leave them blank and
belong to the row above. The last line of every sheet is a `Sub Total` row
rather than an item.

Three things in the real file are worth naming, because each would otherwise
import as plausible-looking nonsense — or not import at all.

**Shifted rows.** On 24 sheets the item name is empty in Accurate, and the
export drops the column rather than leaving it blank, so everything slides one
place left: the quantity lands in `Nama Barang`, the price in `Kuantitas`, the
line total in `@Harga`, and `Total Harga` is empty. Read literally that is "an
item called 200, of which there are 68000". These rows are put back rather than
thrown away, because throwing them away silently lowers what the quotation was
worth — and the shift is safe to undo precisely because the numbers still
multiply out once they are in the right columns.

**Line discounts.** A dozen lines state a total that is 2% under quantity ×
price. That is a discount, not corruption, so the stated total is taken as
authoritative and the discount is reported. An importer that "fixed" the
arithmetic here would quietly raise the price of work already quoted.

**A subtotal to check against.** Each sheet states its own total, so the lines
can be checked against what Accurate thought the quotation was worth. That is
what makes the two repairs above safe to make rather than merely plausible:
if a repair were wrong, the sheet's own total would disagree, and the
quotation says so instead of importing quietly.
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from datetime import date, datetime

HEADERS = ("nomor", "tanggal", "pelanggan", "nama barang",
           "kuantitas", "@harga", "total harga")


def _norm(s) -> str:
    return re.sub(r"\s+", " ", str(s if s is not None else "").strip().lower())


def _clean(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _num(v) -> float | None:
    """A cell that is meant to be money or a count, or None if it isn't one."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


# Accurate writes a third of this file's dates as Excel dates and the rest as
# Indonesian text — "09 Mei 2024". `strptime` has no locale to fall back on
# here, so the twelve month names are simply listed.
ID_MONTHS = {
    "jan": 1, "feb": 2, "peb": 2, "mar": 3, "apr": 4, "mei": 5, "jun": 6,
    # August alone appears three ways in this one file: Agu, Ags and Agt.
    "jul": 7, "agu": 8, "ags": 8, "agt": 8, "sep": 9, "okt": 10, "nov": 11,
    "des": 12,
}


def _date(v) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _clean(v)
    if not s:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    m = re.match(r"^\s*(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})\s*$", s)
    if m:
        month = ID_MONTHS.get(m.group(2)[:3].lower())
        if month:
            try:
                return date(int(m.group(3)), month, int(m.group(1)))
            except ValueError:
                return None
    return None


@dataclass
class MappedLine:
    line_no: int
    description: str
    qty: float
    unit_price: float
    line_total: float
    recovered: bool = False          # the shifted-column repair was applied
    discount_pct: float = 0.0        # implied by a stated total under qty x price


def _adds_up(qty, price, total) -> bool:
    """Do these three numbers describe the same line? Tolerant of rounding."""
    if qty is None or price is None or total is None:
        return False
    return abs(qty * price - total) <= max(1.0, abs(total) * 0.005)


@dataclass
class MappedQuotation:
    sheet: str
    number: str
    quote_date: date | None
    customer_name: str
    lines: list[MappedLine] = field(default_factory=list)
    stated_subtotal: float | None = None
    computed_subtotal: float = 0.0
    dropped_rows: int = 0
    warnings: list[str] = field(default_factory=list)


def _clean_customer(raw: str) -> str:
    """Accurate glues its own column headings onto some exported names.

    One customer comes out as `Sugar Group Companies Mata Uang   Indonesian
    Rupiah`. Everything from "Mata Uang" on is the export's, not the name's.
    """
    s = re.split(r"\bMata Uang\b", raw or "", flags=re.I)[0]
    return re.sub(r"\s+", " ", s).strip()


def read_workbook(data: bytes) -> list[tuple[str, list[list]]]:
    """Every sheet, as (name, rows). Cell values are kept as Excel typed them
    so dates stay dates and numbers stay numbers."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    out = [(name, [list(r) for r in wb[name].iter_rows(values_only=True)])
           for name in wb.sheetnames]
    wb.close()
    return out


def map_quotations(sheets: list[tuple[str, list[list]]]) -> tuple[list[MappedQuotation], list[str]]:
    problems: list[str] = []
    out: list[MappedQuotation] = []
    unrecognised: list[str] = []

    for sheet_name, rows in sheets:
        if not rows:
            continue
        header = [_norm(c) for c in rows[0]]
        if tuple(header[:7]) != HEADERS:
            unrecognised.append(sheet_name)
            continue

        body = [r for r in rows[1:] if any(_clean(c) for c in r)]
        if not body:
            continue

        head = body[0]
        number = _clean(head[0])
        if not number:
            unrecognised.append(sheet_name)
            continue

        q = MappedQuotation(
            sheet=sheet_name,
            number=number,
            quote_date=_date(head[1]),
            customer_name=_clean_customer(_clean(head[2]) or ""),
        )
        if not q.customer_name:
            q.warnings.append("no customer named in the sheet")
        if not q.quote_date:
            q.warnings.append("no readable date — the import date will be used")

        for row in body:
            # The closing line states the sheet's own total; keep it to check
            # the arithmetic against, but it is not an item.
            if _norm(row[5] if len(row) > 5 else "") == "sub total":
                q.stated_subtotal = _num(row[6] if len(row) > 6 else None)
                continue

            desc = _clean(row[3] if len(row) > 3 else None)
            qty = _num(row[4] if len(row) > 4 else None)
            price = _num(row[5] if len(row) > 5 else None)
            total = _num(row[6] if len(row) > 6 else None)

            if not desc:
                q.dropped_rows += 1
                continue

            recovered = False
            discount = 0.0

            if _adds_up(qty, price, total):
                pass                                  # the ordinary case

            elif total is None and _adds_up(_num(desc), qty, price):
                # Shifted left by one: the item had no name. Slide it back.
                # `_adds_up` on the shifted reading is the proof, not a guess —
                # three unrelated numbers do not multiply out by accident.
                qty, price, total = _num(desc), qty, price
                desc = "(no item name in the export)"
                recovered = True

            elif (qty is not None and price is not None and total is not None
                  and qty and price and 0 < total < qty * price):
                # A stated total below quantity x price is a line discount.
                # The document is what it is: keep the total it states.
                discount = round((1 - total / (qty * price)) * 100, 2)

            else:
                q.dropped_rows += 1
                q.warnings.append(
                    f"a line could not be read ({desc[:28]!r}: qty {qty}, "
                    f"price {price}, total {total}) — left out")
                continue

            q.lines.append(MappedLine(
                line_no=len(q.lines) + 1,
                description=desc,
                qty=qty,
                unit_price=price,
                line_total=total,
                recovered=recovered,
                discount_pct=discount,
            ))
            if recovered:
                q.warnings.append(
                    f"line {len(q.lines)} had no item name and the export "
                    f"dropped a column; read back as {qty:g} x {price:,.0f}")
            if discount:
                q.warnings.append(
                    f"line {len(q.lines)} carries a {discount:g}% discount "
                    f"({desc[:24]}) — the quoted total is kept as stated")

        q.computed_subtotal = round(sum(x.line_total for x in q.lines), 2)
        if q.stated_subtotal is not None and q.lines:
            gap = abs(q.stated_subtotal - q.computed_subtotal)
            if gap > max(1.0, abs(q.stated_subtotal) * 0.005):
                q.warnings.append(
                    f"lines add up to {q.computed_subtotal:,.0f} but the sheet "
                    f"says {q.stated_subtotal:,.0f}")
        if not q.lines:
            q.warnings.append("no usable lines on this sheet")
        out.append(q)

    if unrecognised:
        problems.append(
            f"{len(unrecognised)} sheet(s) do not look like a quotation and were "
            f"skipped: {', '.join(unrecognised[:5])}"
            + ("…" if len(unrecognised) > 5 else ""))
    if not out:
        problems.append(
            "No quotations found. Expected one worksheet per quotation with "
            "the columns Nomor, Tanggal, Pelanggan, Nama Barang, Kuantitas, "
            "@Harga, Total Harga.")
    return out, problems

"""Parse an uploaded spreadsheet (or fall back to the AI doc parser for
PDF/image) into draft quotation line items, so the new-quotation form can be
auto-filled. Deterministic for .xlsx/.csv — no LLM needed; the user reviews
and edits everything before the quotation is created."""

import csv
import io
import re

from loguru import logger

_DESC_KEYS = (
    "description", "deskripsi", "uraian", "keterangan", "item", "items",
    "produk", "product", "barang", "jasa", "nama", "name",
)
_QTY_KEYS = ("qty", "quantity", "jumlah", "kuantitas", "qnty", "volume", "vol", "banyak")
_UOM_KEYS = ("uom", "unit", "satuan", "sat")
# Checked before UOM so "unit price" / "harga satuan" land here, not on UOM.
_PRICE_KEYS = (
    "unit price", "unitprice", "unit_price", "harga satuan", "hargasatuan",
    "harga", "price", "rate",
)
_COST_KEYS = ("cost estimate", "cost_estimate", "cost", "hpp", "modal")
# Rows whose description is really a summary line — skip them.
_SKIP_DESC = (
    "total", "subtotal", "sub total", "grand total", "ppn", "pajak", "tax",
    "discount", "diskon", "dpp", "terbilang",
)
_MAX_ITEMS = 200


def _num(v) -> float:
    """Best-effort number parse tolerant of 'Rp', spaces and ID/EN thousands
    separators ('1.500.000,50' and '1,500,000.50' both work)."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return 0.0
    s = s.replace("Rp", "").replace("rp", "").replace(" ", "")
    if "." in s and "," in s:
        # Whichever separator comes last is the decimal one.
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # ID convention: comma is the decimal sep; multiple commas = thousands.
        s = s.replace(",", "") if s.count(",") > 1 else s.replace(",", ".")
    elif "." in s:
        # Dots are thousands separators when repeated, or when the dot groups
        # digits in 3s ("1.500.000", "250.000") — common Indonesian currency.
        if s.count(".") > 1 or len(s.rsplit(".", 1)[-1]) == 3:
            s = s.replace(".", "")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s) if s and s not in (".", "-") else 0.0
    except ValueError:
        return 0.0


def _rows_from_xlsx(blob: bytes) -> list[list]:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 2000:
            break
        rows.append(list(row))
    wb.close()
    return rows


def _rows_from_csv(blob: bytes) -> list[list]:
    text = blob.decode("utf-8-sig", errors="ignore")
    try:
        dialect = csv.Sniffer().sniff(text[:2048], delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    return [list(r) for r in csv.reader(io.StringIO(text), dialect)]


def _match(header: str, keys: tuple[str, ...]) -> bool:
    return any(k in header for k in keys)


def _find_header(rows: list[list]) -> tuple[int, dict] | None:
    """Find the header row in the first 15 rows and map column indices."""
    for idx, row in enumerate(rows[:15]):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        has_desc = any(_match(c, _DESC_KEYS) for c in cells)
        has_price = any(_match(c, _PRICE_KEYS) for c in cells)
        if not (has_desc and has_price):
            continue
        cols: dict = {}
        for ci, c in enumerate(cells):
            if not c:
                continue
            # Order matters: price/qty/cost/uom before desc.
            if "price" not in cols and _match(c, _PRICE_KEYS):
                cols["price"] = ci
            elif "qty" not in cols and _match(c, _QTY_KEYS):
                cols["qty"] = ci
            elif "cost" not in cols and _match(c, _COST_KEYS):
                cols["cost"] = ci
            elif "uom" not in cols and _match(c, _UOM_KEYS):
                cols["uom"] = ci
            elif "desc" not in cols and _match(c, _DESC_KEYS):
                cols["desc"] = ci
        if "desc" in cols and "price" in cols:
            return idx, cols
    return None


def _item(line_no: int, desc: str, qty: float, uom: str, price: float, cost: float) -> dict:
    return {
        "line_no": line_no,
        "source": "custom",
        "description": desc,
        "spec": {},
        "qty": qty if qty > 0 else 1.0,
        "uom": (uom or "pcs").strip() or "pcs",
        "unit_price": price,
        "cost_estimate": cost,
    }


def _parse_rows(rows: list[list]) -> dict:
    warnings: list[str] = []
    found = _find_header(rows)
    if found:
        start, cols = found
        data_rows = rows[start + 1:]
    else:
        # No recognizable header — assume Description, Qty, UoM, Unit price.
        warnings.append(
            "No header row detected — assumed columns: Description, Qty, UoM, "
            "Unit price. Please check the lines below."
        )
        cols = {"desc": 0, "qty": 1, "uom": 2, "price": 3}
        data_rows = rows

    def cell(row, key):
        ci = cols.get(key)
        if ci is None or ci >= len(row):
            return None
        return row[ci]

    items: list[dict] = []
    skipped = 0
    for row in data_rows:
        if not any(c is not None and str(c).strip() for c in row):
            continue
        desc = cell(row, "desc")
        desc = str(desc).strip() if desc is not None else ""
        if not desc:
            skipped += 1
            continue
        if any(desc.lower().startswith(k) or desc.lower() == k for k in _SKIP_DESC):
            continue
        items.append(_item(
            line_no=len(items) + 1,
            desc=desc,
            qty=_num(cell(row, "qty")),
            uom=str(cell(row, "uom") or "").strip(),
            price=_num(cell(row, "price")),
            cost=_num(cell(row, "cost")),
        ))
        if len(items) >= _MAX_ITEMS:
            warnings.append(f"Only the first {_MAX_ITEMS} lines were imported.")
            break

    if skipped:
        warnings.append(f"{skipped} row(s) without a description were skipped.")
    if not items:
        warnings.append(
            "No line items could be read from the file. Add them manually, or "
            "use columns named like Description / Qty / UoM / Unit price."
        )
    return {"items": items, "warnings": warnings}


async def parse_upload(blob: bytes, *, filename: str | None, content_type: str | None) -> dict:
    name = (filename or "").lower()
    ct = (content_type or "").lower()

    if name.endswith(".xlsx") or "spreadsheetml" in ct:
        try:
            return {"source": "xlsx", **_parse_rows(_rows_from_xlsx(blob))}
        except Exception as exc:
            logger.warning(f"xlsx parse failed: {exc}")
            return {"source": "xlsx", "items": [], "warnings": [
                "Couldn't read that Excel file. Make sure it's a real .xlsx "
                "(not .xls) and try again."
            ]}
    if name.endswith(".xls"):
        return {"source": "xls", "items": [], "warnings": [
            "Old .xls files aren't supported — open it in Excel and 'Save As' "
            ".xlsx, then upload again."
        ]}
    if name.endswith(".csv") or "csv" in ct:
        try:
            return {"source": "csv", **_parse_rows(_rows_from_csv(blob))}
        except Exception as exc:
            logger.warning(f"csv parse failed: {exc}")
            return {"source": "csv", "items": [], "warnings": ["Couldn't read that CSV file."]}

    # PDF / image / text → best-effort via the AI document parser.
    try:
        from app.ai import doc_parser
        parsed = await doc_parser.parse(None, blob, filename=filename, content_type=content_type)
        raw_items = ((parsed or {}).get("extracted") or {}).get("items") or []
        items = []
        for r in raw_items:
            desc = str((r or {}).get("description") or "").strip()
            if not desc:
                continue
            items.append(_item(
                line_no=len(items) + 1,
                desc=desc,
                qty=_num(r.get("qty")),
                uom=str(r.get("uom") or "").strip(),
                price=_num(r.get("unit_price")),
                cost=0.0,
            ))
        warnings = []
        if not items:
            warnings.append(
                "Couldn't pull line items from that file automatically — add "
                "them manually below."
            )
        elif parsed and parsed.get("needs_review"):
            warnings.append("Low confidence — please double-check every line.")
        return {"source": "document", "items": items, "warnings": warnings}
    except Exception as exc:
        logger.warning(f"document parse fallback failed: {exc}")
        return {"source": "unknown", "items": [], "warnings": [
            "Unsupported file type. Upload an .xlsx or .csv with your line items."
        ]}

"""Read an Accurate customer export and map it onto our Customer records.

The export is the "Daftar Pelanggan" sheet: ~80 columns, most of them
accounting settings we have nowhere to put and no need for. Six of them carry
the information this app actually works from.

The one that isn't obvious is **Kategori**. Accurate uses it as a free-text
grouping and this company has been using it to record *whose customer it is* —
"Customer Candra", "Customer Gora", "customer Kantor". That is our
`sales_pic_id`, and it is the single most valuable column in the file: without
it every imported customer lands unassigned and the sales scoping that the rest
of the app depends on has nothing to work with.

Nothing here writes. Parsing and mapping are separated from committing so the
same code can answer "what would happen" before anyone commits to it.
"""

from __future__ import annotations

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field

# Accurate writes the same sheet with slightly different headers depending on
# version and locale, so match on a normalised form rather than exact text.
COLUMNS = {
    "external_code": ["id pelanggan", "customer id"],
    "category": ["kategori"],
    "company_name": ["nama", "name"],
    "pic_name": ["kontak", "contact"],
    "phone_business": ["no. telp. bisnis", "no telp bisnis", "business phone"],
    "mobile": ["handphone", "mobile phone", "mobile"],
    "email": ["email"],
    "billing_address": ["alamat penagihan", "billing address"],
    "city": ["kota", "city"],
    "province": ["provinsi", "province"],
    "postcode": ["kode pos", "zip code"],
    "delivery_address": ["alamat pengiriman", "shipping address"],
    "payment_terms": ["syarat pembayaran", "terms of payment"],
    "tax_id": ["npwp"],
    "tax_name": ["nama wajib pajak", "tax name"],
    "tax_address": ["alamat (pajak)", "tax address"],
    "notes": ["catatan", "notes"],
}

# Industry is not in the export. These are the words that appear in Indonesian
# company names for the sectors this business actually sells into — a guess
# offered in the preview, never applied silently.
INDUSTRY_HINTS: list[tuple[str, tuple[str, ...]]] = [
    ("sugar", ("gula", "sugar", "pg.", "pg ", "pabrik gula", "manis", "tebu")),
    ("cement", ("semen", "cement", "conbloc")),
    ("fertilizer", ("pupuk", "fertilizer", "pusri", "kaltim")),
    ("pltu", ("pltu", "power", "energi", "energy", "listrik")),
    ("pulp_paper", ("pulp", "kertas", "paper")),
    ("food", ("food", "makanan", "indofood", "mayora", "wings", "coconut",
              "furnindo", "diamond cold", "gandum")),
    ("mining", ("mining", "tambang", "coal", "batubara", "bara", "nikel",
                "nickel", "adaro", "kideco", "vale", "mineral")),
]


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFKD", (s or "")).strip().lower()
    return re.sub(r"\s+", " ", s)


def _clean(v: str | None) -> str | None:
    """Trim, drop Accurate's markdown escaping, and treat blanks as absent."""
    if v is None:
        return None
    s = str(v).replace("\\_", "_").replace("\\[", "[").replace("\\]", "]").strip()
    return s or None


def guess_industry(company_name: str) -> str:
    name = _norm(company_name)
    for industry, words in INDUSTRY_HINTS:
        if any(w in name for w in words):
            return industry
    return "other"


def parse_payment_terms(raw: str | None) -> dict:
    """`net 30` / `C.O.D` / `net 45` -> something the app can reason about."""
    s = _norm(raw or "")
    if not s:
        return {}
    if "c.o.d" in s or s == "cod":
        return {"raw": raw, "kind": "cod", "days": 0}
    m = re.search(r"(\d+)", s)
    if "net" in s and m:
        return {"raw": raw, "kind": "net", "days": int(m.group(1))}
    return {"raw": raw}


def sales_rep_from_category(category: str | None) -> str | None:
    """`Customer Candra` -> `candra`. `Umum` / `customer Kantor` -> nobody.

    "Umum" means general/unassigned and "Kantor" means the office rather than a
    person, so both correctly resolve to no rep rather than to a bad guess.
    """
    s = _norm(category or "")
    if not s:
        return None
    s = re.sub(r"^customer\s+", "", s).strip()
    if s in ("", "umum", "kantor", "office", "general"):
        return None
    return s


@dataclass
class MappedCustomer:
    row_no: int
    external_code: str | None
    company_name: str
    industry: str
    industry_guessed: bool
    pic_name: str | None = None
    phone: str | None = None
    whatsapp: str | None = None
    email: str | None = None
    company_address: str | None = None
    delivery_address: str | None = None
    tax_id: str | None = None
    tax_name: str | None = None
    tax_address: str | None = None
    is_pkp: bool = False
    payment_terms: dict = field(default_factory=dict)
    sales_rep_hint: str | None = None
    notes: str | None = None
    warnings: list[str] = field(default_factory=list)


def _header_index(header: list[str]) -> dict[str, int]:
    norm = [_norm(h) for h in header]
    out: dict[str, int] = {}
    for field_name, candidates in COLUMNS.items():
        for cand in candidates:
            if cand in norm:
                out[field_name] = norm.index(cand)
                break
    return out


def read_rows(data: bytes, filename: str = "") -> tuple[list[str], list[list[str]]]:
    """Header + rows, from .xlsx or .csv. Only the first sheet is read."""
    if filename.lower().endswith((".xlsx", ".xlsm")) or data[:2] == b"PK":
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = [["" if c is None else str(c) for c in r]
                for r in ws.iter_rows(values_only=True)]
        wb.close()
        if not rows:
            return [], []
        return rows[0], rows[1:]

    text = data.decode("utf-8-sig", errors="replace")
    rdr = list(csv.reader(io.StringIO(text)))
    if not rdr:
        return [], []
    return rdr[0], rdr[1:]


def map_customers(header: list[str], rows: list[list[str]]) -> tuple[list[MappedCustomer], list[str]]:
    """Map raw rows onto customer records. Returns (mapped, file-level problems)."""
    idx = _header_index(header)
    problems: list[str] = []
    if "company_name" not in idx:
        problems.append(
            "Could not find the customer name column. Expected a header called "
            "'Nama'. Is this the Daftar Pelanggan export?")
        return [], problems
    if "category" not in idx:
        problems.append(
            "No 'Kategori' column — customers will import without a sales rep, "
            "and each one will have to be assigned by hand afterwards.")

    def cell(row: list[str], key: str) -> str | None:
        i = idx.get(key)
        if i is None or i >= len(row):
            return None
        return _clean(row[i])

    out: list[MappedCustomer] = []
    for n, row in enumerate(rows, start=2):      # row 1 is the header
        name = cell(row, "company_name")
        if not name:
            continue                              # blank spacer row
        warnings: list[str] = []

        industry = guess_industry(name)
        rep = sales_rep_from_category(cell(row, "category"))
        if not rep:
            warnings.append("no sales rep in Kategori — will import unassigned")

        # Accurate exports a leading apostrophe on numeric-looking text.
        tax_id = (cell(row, "tax_id") or "").lstrip("'") or None

        addr = cell(row, "billing_address")
        city, prov, post = cell(row, "city"), cell(row, "province"), cell(row, "postcode")
        if addr:
            tail = ", ".join(x for x in (city, prov, post) if x)
            if tail and _norm(tail) not in _norm(addr):
                addr = f"{addr}, {tail}"

        if name.startswith("["):
            warnings.append("name looks like an Accurate duplicate marker")

        out.append(MappedCustomer(
            row_no=n,
            external_code=cell(row, "external_code"),
            company_name=name,
            industry=industry,
            industry_guessed=True,
            pic_name=cell(row, "pic_name"),
            phone=cell(row, "phone_business") or cell(row, "mobile"),
            whatsapp=cell(row, "mobile"),
            email=cell(row, "email"),
            company_address=addr,
            delivery_address=cell(row, "delivery_address"),
            tax_id=tax_id,
            tax_name=cell(row, "tax_name"),
            tax_address=cell(row, "tax_address"),
            is_pkp=bool(tax_id),
            payment_terms=parse_payment_terms(cell(row, "payment_terms")),
            sales_rep_hint=rep,
            notes=cell(row, "notes"),
            warnings=warnings,
        ))
    return out, problems
